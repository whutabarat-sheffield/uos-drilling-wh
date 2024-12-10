import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import cm
from mpl_toolkits.axes_grid1 import make_axes_locatable
import os
import scaleogram as scg
from glob import glob
from abyss.dataparser import loadSetitecXls
import multiprocessing as mp
from abyss.energy_estimation import get_energy_estimation

def combineLegends(*ax,loc=0):
    '''
        Combine legends from multiple axes together

        Meant to be used when multiple axes have the same signals but you only
        want one legend for easier viewing.

        Places legend on the first axis

        Inputs:
            ax : Iterable collection of matplotlib axes
            loc : Where to put the legend. See matplotlib.legend

        Returns list of axes
    '''
    lns = []
    labels = []
    for aa in ax:
        for ll in aa.lines:
            lns.append(ll)
            labels.append(ll.get_label())
    ax[0].legend(lns,labels,loc=loc)
    return ax

def getEnergyEst(path):
    data = loadSetitecXls(path,'auto_data')
    if 'I Torque Empty (A)' in data:
        tqe = data['I Torque Empty (A)'].values
    else:
        tqe = np.zeros(data['I Torque (A)'].values.shape)
    return get_energy_estimation(Itorque=data["I Torque (A)"].values,
                                       Itorque_empty=tqe,
                                       Ithrust=data["I Thrust (A)"].values)
                                     
def plotAllEnergyEst(path,**kwargs):
    '''
        Estimate energy for each found file and plot it as a scatter plot

        Inputs:
            path : String or list of paths
            title : Axes title. Default Energy Estimation (J)

        Returns figure
    '''
    if isinstance(path,str):
        path = glob(path)
    ee = [getEnergyEst(f) for f in path]
    f,ax = plt.subplots()
    ax.plot(ee,'x')
    ax.set(xlabel='File Index',ylabel="Energy Estimate (J)",title=kwargs.get("title","Energy Estimate (J)"))
    return f

def plotMaxTorque(path,**kwargs):
    '''
        Plot the max torque of each found file

        Inputs:
            path : String or list of paths
            title : Axes title. Default Max Torque (A)

        Returns figure
    '''
    if isinstance(path,str):
        path = glob(path)
    # lists to hold values
    maxtq = []
    for fn in path:
        # load file and get torque
        data = loadSetitecXls(fn,'auto_data')
        tq = data['I Torque (A)'].values.flatten()
        if 'I Torque Empty (A)' in data:
            tq += data['I Torque Empty (A)'].values.flatten()
        maxtq.append(tq.max())
    # plot
    f,ax = plt.subplots()
    ax.scatter(maxtq,'x')
    ax.set(xlabel="File Index",ylabel="Max Torque (A)",title=kwargs.get("title","Max Torque (A)"))
    return f

def plotMaxTorquePos(path,**kwargs):
    '''
        Plot the max torque of each found file against the location of the max torque

        Inputs:
            path : String or list of paths
            title : Axes title. Default Max Torque (A) + location (mm)

        Returns figure
    '''
    if isinstance(path,str):
        path = glob(path)
    # lists to hold values
    pos = []
    maxtq = []
    for fn in path:
        # load file and get torque
        data = loadSetitecXls(fn,'auto_data')
        tq = data['I Torque (A)'].values.flatten()
        if 'I Torque Empty (A)' in data:
            tq += data['I Torque Empty (A)'].values.flatten()
        # find index of max
        ii = tq.argmax()
        # append to lists
        pos.append(data['Position (mm)'].values[ii])
        maxtq.append(tq.max())
    # plot
    f,ax = plt.subplots()
    ax.scatter(pos,maxtq,'x')
    ax.set(xlabel="Position (mm)",ylabel="Max Torque (A)",title=kwargs.get("title","Max Torque (A) + location (mm)"))
    return f

def plotMaxThrust(path,**kwargs):
    '''
        Plot the max thrust of each found file

        Inputs:
            path : String or list of paths
            title : Axes title. Default Max Thrust (A)

        Returns figure
    '''
    if isinstance(path,str):
        path = glob(path)
    # lists to hold values
    maxtq = []
    for fn in path:
        # load file and get torque
        data = loadSetitecXls(fn,'auto_data')
        tq = data['I Thrust (A)'].values.flatten()
        if 'I Thrust Empty (A)' in data:
            tq += data['I Thrust Empty (A)'].values.flatten()
        maxtq.append(tq.max())
    # plot
    f,ax = plt.subplots()
    ax.scatter(maxtq,'x')
    ax.set(xlabel="File Index",ylabel="Max Thrust (A)",title=kwargs.get("title","Max Thrust (A)"))
    return f

def plotMaxThrustPos(path,**kwargs):
    '''
        Plot the max Thrust of each found file against the location of the max torque

        Inputs:
            path : String or list of paths
            title : Axes title. Default Max Thrust (A) + location (mm)

        Returns figure
    '''
    if isinstance(path,str):
        path = glob(path)
    # lists to hold values
    pos = []
    maxtq = []
    for fn in path:
        # load file and get torque
        data = loadSetitecXls(fn,'auto_data')
        tq = data['I Thrust (A)'].values.flatten()
        if 'I Thrust Empty (A)' in data:
            tq += data['I Thrust Empty (A)'].values.flatten()
        # find index of max
        ii = tq.argmax()
        # append to lists
        pos.append(data['Position (mm)'].values[ii])
        maxtq.append(tq.max())
    # plot
    f,ax = plt.subplots()
    ax.scatter(pos,maxtq,'x')
    ax.set(xlabel="Position (mm)",ylabel="Max Thrust (A)",title=kwargs.get("title","Max Thrust (A) + location (mm)"))
    return f

def plotChangeTypes(path,t=0.2,**kwargs):
    '''
        For each file determine if the transition from step code 0 to 1 was caused
        by signal or DEP

        The Setitec program sets certain triggers to trigger changes between step codes.
        They can be broadly classified as either distance or signal (thrust or torque) related
        trigger.

        This program compares the final position of step 0 against the DEP trigger of step code 0.
        If they're more than t mm appart, then the trigger was the signals rather than DEP

        Creates a scatter plot for each file indicating if the change was Signal or DEP

        Inputs:
            path : String path or list of paths
            t : Threshold for check in mm. Default 0.2 mm
            title : Axis title. Default f"Trigger Type, sc=(0,1), t={t:.2f}mm"

        Returns figure
    '''
    from dataparser import getProgramTriggers, getStepCodeFinalPos
    if isinstance(path,str):
        path = glob(path)
    changes = []
    for fn in path:
        # get trigger for step code 0
        trig = getProgramTriggers(fn)[0.0][0][1]
        # get step code final position for step code 0
        pos = abs(getStepCodeFinalPos(fn,0)[0.0])
        # if the trigger is within 0.5mm of the final position value of step code 0
        # then it was triggered by position rather than torque
        changes.append(int(abs(trig-pos)<=t))
    f,ax = plt.subplots()
    ax.plot(changes,'x')
    ax.set(xlabel="File Index",ylabel="Trigger Change Type",title=kwargs.get("title",f"Trigger Type, sc=(0,1), t={t:.2f}mm"))
    ax.set_yticks([0,1])
    ax.set_yticklabels(['Signal','DEP'])
    return f

def plotMaxTorqueMaxThrust(path,**kwargs):
    '''
        Plot Max Torque against max Thrust

        Inputs:
            path : string or list of paths
            title : Axis title. Default Max Torque (A) vs Max Thrust (A)

        Return figure
    '''
    if isinstance(path,str):
        path = glob(path)
    maxtq = []
    maxtt = []
    for fn in path:
        df = loadSetitecXls(fn,'auto_data')
        maxtq.append((df['I Torque (A)'].values.flatten() + (df['I Torque (Empty A)'].values.flatten()[-1] if 'I Torque (Empty A)' in df else np.zeros(df.shape[0]))).max())
        maxtt.append((df['I Thrust (A)'].values.flatten() + (df['I Thrust (Empty A)'].values.flatten()[-1] if 'I Thrust (Empty A)' in df else np.zeros(df.shape[0]))).max())
    f,ax = plt.subplots()
    ax.plot(maxtq,maxtt,'x')
    ax.set(xlabel="Max Torque (A)",ylabel="Max Thrust (A)",title=kwargs.get("title","Max Torque (A) vs Max Thrust (A)"))
    return f

def plotAllChangeTypes(path,t=0.2,**kwargs):
    '''
        For each file determine if the transition from between step codes was caused
        by signal or DEP
        
        See plotChangeTypes.

        Inputs:
            path : String path or list of paths
            t : Threshold for check in mm. Default 0.2 mm
            title : Axis title. Default f"Trigger Type, t={t:.2f}mm"

        Returns figure
    '''
    if isinstance(path,str):
        path = glob(path)
    changes = {}
    # list of step codes to search for
    sc = list(range(10))
    for fn in path:
        all_pos = getStepCodeFinalPos(fn,sc)
        all_trigs = getProgramTriggers(fn)
        for k,v in all_pos.items():
            # get trigger for step code k
            trig = all_trigs[k][0][1]
            # get step code final position for step code 0
            pos = abs(v)
            # if the trigger is within 0.5mm of the final position value of step code 0
            # then it was triggered by position rather than torque
            if not (k in changes):
                changes[k] = []
            changes[k].append(int(abs(trig-pos)<=t))
    f,ax = plt.subplots()
    for kk,cc in changes.items():
        ax.plot(cc,'x',label=f"SC {kk}")
    ax.legend()
    ax.set(xlabel="File Index",ylabel="Trigger Change Type",title=kwargs.get("title",f"Trigger Type, t={t:.2f}mm"))
    ax.set_yticks([0,1])
    ax.set_yticklabels(['Signal','DEP'])
    return f

def plotAllStepCodeStartPos(path,scs='all',**kwargs):
    '''
        Plot the starting position for each step code found for each file

        Loads each file, filters it to each unique step code and saves the first position value.
        The purpose of this is to look for abnormal values that might be too high or low inidicating an issue

        Input:
            path : String or list of paths
            scs : Single or list of floating numerical step codes to search for.
                Also accepts string all to indicate all step codes in the file
                Default 1.
            title : Axis title. Default Step Code First Positions (mm)

        Returns figure
    '''
    if isinstance(path,str):
        path = glob(path)
    pos = {}
    for fn in path:
        # get first pos for each step code
        pfn = getStepCodeStartPos(fn,scs)
        # update dictionary with values
        for kk,vv in pfn.items():
            if not (kk in pos):
                pos[kk] = []
            pos[kk].append(vv)
    # plot each step code with a separate color and label
    f,ax =plt.subplots()
    for kk,vv in pos.items():
        ax.plot(len(vv)*[int(kk),],vv,'x',label=f"SC {int(kk)}")
    # set x label ticks to the tick locatins rather than auto generates
    ax.set_xticks(list(pos.keys()))
    ax.set(xlabel="Step Code",ylabel="First Position of Step Code (mm)",title=kwargs.get("title","Step Code First Positions (mm)"))
    ax.legend()
    return f

def plotAllStepCodeFinalPos(path,scs='all',**kwargs):
    '''
        Plot the final position for each step code found for each file

        Loads each file, filters it to each unique step code and saves the first position value.
        The purpose of this is to look for abnormal values that might be too high or low inidicating an issue

        Input:
            path : String or list of paths
            scs : Single or list of floating numerical step codes to search for.
                Also accepts string all to indicate all step codes in the file
                Default 1.
            title : Axis title. Default Step Code First Positions (mm)

        Returns figure
    '''
    if isinstance(path,str):
        path = glob(path)
    pos = {}
    for fn in path:
        # get first pos for each step code
        pfn = getStepCodeFinalPos(fn,scs)
        # update dictionary with values
        for kk,vv in pfn.items():
            if not (kk in pos):
                pos[kk] = []
            pos[kk].append(vv)
    # plot each step code with a separate color and label
    f,ax =plt.subplots()
    for kk,vv in pos.items():
        ax.plot(len(vv)*[int(kk),],vv,'x',label=f"SC {int(kk)}")
    # set x label ticks to the tick locatins rather than auto generates
    ax.set_xticks(list(pos.keys()))
    ax.set(xlabel="Step Code",ylabel="Final Position of Step Code (mm)",title=kwargs.get("title","Step Code First Positions (mm)"))
    ax.legend()
    return f

def plotStepCodeFinalSignal(path,scs='all',**kwargs):
    '''
        Plot the final signal values for each step code found for each file

        Loads each file, filters it to each unique step code and saves the first signal values.
        The purpose of this is to look for abnormal values that might be too high or low inidicating an issue

        Input:
            path : String or list of paths
            scs : Single or list of floating numerical step codes to search for.
                Also accepts string all to indicate all step codes in the file
                Default 1.
            title : Axis title. Default "Step Code Final Signal (A)"

        Returns figure
    '''
    if isinstance(path,str):
        path = glob(path)
    pos = {}
    for fn in path:
        # get first pos for each step code
        pfn = getStepCodeFinalSignal(fn,scs)
        # update dictionary with values
        for kk,vv in pfn.items():
            if not (kk in pos):
                pos[kk] = []
            pos[kk].append(vv)
    # plot each step code with a separate color and label
    f,ax =plt.subplots(ncols=2)
    for kk,vv in pos.items():
        sig = [v[0] for v in vv]
        ax[0].plot(len(sig)*[int(kk),],sig,'x',label=f"SC {int(kk)}")
        sig = [v[1] for v in vv]
        ax[1].plot(len(sig)*[int(kk),],sig,'x',label=f"SC {int(kk)}")

    # set x label ticks to the tick locatins rather than auto generates
    ax.set_xticks(list(pos.keys()))
    ax[0].set(xlabel="Step Code",ylabel="Final Torque Value (A)",title="Torque (A)")
    ax[1].set(xlabel="Step Code",ylabel="Final Thrust Value (A)",title="Thrust (A)")
    f.suptitle(kwargs.get("title","Step Code Final Signal (A)"))
    ax.legend()
    return f

def plotStepChangeAgainstDEP(path,scs='all'):
    '''
        Plot the final value of each step code against the DEP trigger for the step code

        The purpose is to look for which step code would be triggered by signal rather than DEP

        Inputs:
            path : string or list of paths
            scs : Single or list of floating numerical step codes to search for.
                Also accepts string all to indicate all step codes in the file
                Default 1.

        Return figure
    '''
    if isinstance(path,str):
        path = glob(path)
    acutal_pos = {}
    dep = {}
    for fn in path:
        # get final positions of each step code
        pfn = getStepCodeFinalPos(fn,scs)
        # get program triggers for each step code
        trig = getProgramTriggers(fn)
        # for each key# update dictionaries with postions
        for kk,vv in pfn.items():
            if not (kk in acutal_pos):
                acutal_pos[kk] = []
            acutal_pos[kk].append(vv)

            if not (kk in dep):
                dep[kk] = []
            dep[kk].append(trig[kk][0][1])
    f,ax =plt.subplots()
    for kk,vv in acutal_pos.items():
        ax.plot(dep[kk],acutal_pos[kk],'x',label=f"SC {int(kk)}")
    ax.set(xlabel="Trigger DEP (mm)",ylabel="Final Step Code Position (mm)",title="LEDU20100001, 20110019_20110019_ST_ Step Code Last Pos vs DEP")
    ax.legend()
    return f
  

def plotBatchSetitecSpectrogram(path,fclip=None,clip_mode='gt',opath='',**kwargs):
    '''
        Batch plot Spectrograms for the several Setitec files in a target folder

        Clip the plotted frequency range using fclip and clip_mode. Fclip is the thresholding
        frequency in Hz. clip_mode controls whether the threshold is greater than or equal to (gt) or
        less than or equal to (lt) the target threshold

        Inputs:
            path : Wildcard path to Setitec XLS files.
            fclip : Thresholding frequency. Default None.
            clip_mode : How the thresholding is performed. If gt then it's greater than or equal to fclip.
                        If lt then it's greater than or equal to fclip. Anything else raises a ValueError.
                        Default gt.
            opath : Where to save the figures to. Default local.
            max_title : Figure suptitle for the max response plot
            resp_title : Figure suptitle for the max signal plot

        Returns max STFT response plot and max response vs max signal response figure
    '''
    from scipy.signal import spectrogram
    if not (clip_mode in ['gt','lt']):
        raise ValueError(f"Unsupported clip_mode {clip_mode}! Only supported modes are 'gt' and 'lt'")
    max_ss_tt = []
    max_ss_tq = []
    max_ss_freq_tt = []
    max_ss_freq_tq = []

    maxtq = []
    maxtt = []
    for fn in glob(path):
        fname = os.path.splitext(os.path.basename(fn))[0]
        data = loadSetitecXls(fn,"auto_data")
        fig,ax = plt.subplots(ncols=2,constrained_layout=True)
        signal = data['I Torque (A)'].values.flatten() + data['I Torque Empty (A)'].values.flatten()
        f,t,Sxx = spectrogram(signal,100.0,nperseg=2048)

        if fclip:
            ii = np.where(f>=fclip if clip_mode == 'gt' else f<=fclip)[0]
            f = f[ii]
            Sxx = Sxx[ii,:]
        max_ss_tq.append(Sxx.max())
        max_ss_freq_tq.append(f[np.unravel_index(Sxx.argmax(),Sxx.shape)[0]])
        #print(np.unravel_index(Sxx.argmax(),Sxx.shape))
        maxtq.append(signal.max())
        
        ax[0].pcolormesh(t,f,Sxx,shading='gouraud',cmap='hot')
        ax[0].set(xlabel='Time (s)',ylabel='Freq (Hz)',title="Torque")
        tax = ax[0].twinx()
        tax.plot(np.linspace(t.min(),t.max(),signal.shape[0]),signal,'y-')
        tax.set_ylabel("Torque (A)")
        
        signal = data['I Thrust (A)'].values.flatten()
        if 'I Thrust Empty (A)' in data:
            signal += data['I Thrust Empty (A)'].values.flatten()
        f,t,Sxx = spectrogram(signal,100.0,nperseg=2048)

        if fclip:
            ii = np.where(f>=fclip if clip_mode == 'gt' else f<=fclip)[0]
            f = f[ii]
            Sxx = Sxx[ii,:]
        max_ss_tt.append(Sxx.max())
        max_ss_freq_tt.append(f[np.unravel_index(Sxx.argmax(),Sxx.shape)[0]])
        maxtt.append(signal.max())

        ax[1].pcolormesh(t,f,Sxx,shading='gouraud',cmap='hot')
        ax[1].set(xlabel='Time (s)',ylabel='Freq (Hz)',title="Thrust")
        tax = ax[1].twinx()
        tax.plot(np.linspace(t.min(),t.max(),signal.shape[0]),signal,'y-')
        tax.set_ylabel("Thrust (A)")

        fig.suptitle(f"{fname}, fclip{'>=' if clip_mode == 'gt' else '<='}{fclip} Hz")
        fig.savefig(os.path.join(opath,f"{fname}-spectrogram-fclip-{clip_mode}-{fclip}.png"))
        plt.close(fig)
    fA,ax = plt.subplots(ncols=2,nrows=2,constrained_layout=True)
    ax[0,0].plot(max_ss_tq)
    ax[0,0].set(xlabel="Hole Number",ylabel="Max STFT Response (V**2/Hz)",title="Torque")
    ax[0,1].plot(max_ss_tt)
    ax[0,1].set(xlabel="Hole Number",ylabel="Max STFT Response (V**2/Hz)",title="Thrust")
    
    ax[1,0].plot(max_ss_freq_tt)
    ax[1,0].set(xlabel="Hole Number",ylabel="Max STFT Response Freq (Hz)",title="Torque")
    ax[1,1].plot(max_ss_freq_tt)
    ax[1,1].set(xlabel="Hole Number",ylabel="Max STFT Response Freq (Hz)",title="Thrust")
    fA.suptitle(kwargs.get("max_title",f"Max STFT Response, fclip{'>=' if clip_mode == 'gt' else '<='}{fclip} Hz"))

    fB,ax = plt.subplots(ncols=2,constrained_layout=True)
    ax[0].plot(maxtq,max_ss_tq,'bx')
    ax[0].set(xlabel="Max Torque (A)",ylabel="Max STFT Response (V**2/Hz)",title="Torque")
    ax[1].plot(maxtt,max_ss_tt,'rx')
    ax[1].set(xlabel="Max Thrust (A)",ylabel="Max STFT Response (V**2/Hz)",title="Thrust")
    fB.suptitle(kwargs.get("resp_title",f"Max Signal vs STFT Response,fclip{'>=' if clip_mode == 'gt' else '<='}{fclip} Hz"))
    return fA,fB
                    
def findSignalLims(path,kk,lim=np.max):
    '''
        Apply function lim to key kk in Setitec XLS file path and return the result

        This is to be used inside a multiprocessing pool or thread.

        Inputs:
            path : File path to Setitec XLS file
            kk : Target key inside Setitec run data
            lim : Function to apply to values to extract some info

        Returns result of lim on Series inside run data
    '''
    return lim(loadSetitecXls(path,"auto_data")[kk].values.flatten())

def plotSetitecHistory(path,cmap="binary",as_3d=False,cmap_str="Light to Dark",**kwargs):
    '''
        Plot the Setitec torque and thrust istory using a colormap to show where the line is in the history

        The filenames are sorted by local headcount.

        When as_3d flag is False, the history is plotted on two axes and the colormapping is set by index.
        When as_3d flag is True, the colormapping is based on thr respective signal value

        The purpose of the cmap_str is to provide some quick context the colors. For example, the default colormap is binary which when used to show
        history transitions from light for the early holes to dark.

        If the user doesn't want to plot everything, they can use the keys keyword to specific what they want plotted. The default is all which plots
        I Torque (A), I Torque Empty (A), I Thrust (A), I Thrust Empty (A),I Torque (A)+ I Torque Empty (A) and I Thrust (A)+ I Thrust Empty (A). The user
        can specify a particular key by passing a string and setting search_keys to False. This only plots that specific column

        e.g. specific key
        plotSetitecHistory("8B Life Test/*.xls",key="I Torque (A)",search_keys=False)

        Only I Torque (A) is plotted.

        If search_keys is True then all columns that contain the search term are plotted

        e.g. search key
        plotSetitecHistory("8B Life Test/*.xls",key="Torque",search_keys=True)

        ALL columns containing Torque are plotted

        The keyword vmax controls how the different lines are colour mapped in 3D. By default vmax is set to max_each, which means each line are colour mapped separately.
        When vmax is max, then it's colour mapped according to the global max of the dataset. The user can also set a specific number as the upper limit.

        Inputs:
            path : Wildcard path to Setitec XLS files
            cmap : String name of Matplotlib colormap. Default binary.
            as_3d : Flag to plot the data as 3d scatter plot with hole number along x-axis
            cmap_str : String used in figure title to indicate how the colormap changes so the user can understand the change in signal features over time.
            keys : Which keys to plot. Can be a single string or a list of keys to try. Default all which is Torque, Torque Empty, Thrust, Thrust Empty and appropriate sums
            search_keys : Flag to use the string given in keys as a search term to filter the column names. Default True.
            vmax : Max value used for the colormapping. Can be either max, max_each or a specific number. Default max_each.
            use_mp : Use multiprocessing when vmax is max to calculate the max signal value to use when generating colors. When True, then number of processes is set to 3.
                    User can specify number of processes too. Default True.

        Returns figure object
    '''
    keys =kwargs.get('keys','all')
    if not keys:
        raise ValueError(f"Target keys have to be non-empty! Recevied {keys}")
    # sort the paths according to local hole number
    if isinstance(path,str):
        paths = sorted(glob(path),key=lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[-1])
    else:
        paths = path
    # create figure
    f = plt.figure(constrained_layout=True)
    # if keys is all then create the array
    if keys == 'all':
        # create first axis
        ax = [f.add_subplot(3,2,1,projection='3d' if as_3d else None),]
        # create other axes using first axis to share x with
        for i in range(2,7):
            ax.append(f.add_subplot(3,2,i,projection='3d' if as_3d else None,sharex=ax[0]))
        ax = np.asarray(ax).reshape((3,2))    
    # single key or search term
    elif isinstance(keys,str):
        # load first file
        data = loadSetitecXls(paths[0],"auto_data")
        # get column names
        cn = data.keys()
        # if using the key as a search term
        if kwargs.get("search_keys",True):
            # find columns that contain the target phrase
            pts = keys.split(' ')
            print(keys)
            keys = list(filter(lambda x : all([c in x for c in pts]),cn))
            # if it failed to find any search terms
            if len(keys)==0:
                raise ValueError(f"Failed to find column names containing {key}!")
        else:
            keys = [keys,]
    # if keys is now a list
    if isinstance(keys,(list,tuple)):
        kk = len(keys)
        ax = [f.add_subplot(1,kk,1,projection='3d' if as_3d else None),]
        if kk>1:
            for i in range(2,kk+1):
                ax.append(f.add_subplot(1,kk,i,projection='3d' if as_3d else None,sharex=ax[0]))
        ax = np.asarray(ax) 
    # get the target colormap
    cmap = cm.get_cmap(cmap)
    # get number of paths
    # used in colormaps
    nf = len(paths)
    # get color max mode
    vmax = kwargs.get("vmax","max_each")
    # if vmax is the global maximum
    if as_3d and vmax == 'max':
        # if use_mp is False then iterate over the files sequentially
        if kwargs.get("use_mp",False):
            clim = max([loadSetitecXls(fn,"auto_data")[kk].values.flatten().max() for fn in paths])
        # if use_mp is True then use multiprocessing to find the max
        # if True, then cores set to 3
        # if something else, then treat as a number
        else:
            cdict = {kk: max(mp.Pool(3 if isinstance(ump,bool) else ump).starmap(findSignalLims,[(fn,kk) for fn in paths])) for kk in keys}
    # sort colormaps by global headcount
    # iterate over to use index as way of iterating over colormap
    for fi,fn in enumerate(paths,1):
        # load file        
        data = loadSetitecXls(fn,"auto_data")
        # create hole number vector
        hn = fi*np.ones(data.shape[0],np.int16)
        # get position data
        pos = np.abs(data['Position (mm)'].values.flatten())
        # extract the signals
        if keys == "all":
            # get thrust and torque data
            tq = data['I Torque (A)'].values.flatten()
            tqe = data['I Torque Empty (A)'].values.flatten()
            tt = data['I Thrust (A)'].values.flatten()
            tte = data['I Thrust Empty (A)'].values.flatten()
            # plot the torque data with and without empty
            if as_3d:
                ax[0,0].scatter(hn,pos,tq,color=cmap(tq/tq.max()))
                ax[1,0].scatter(hn,pos,tq+tqe,color=cmap((tq+tqe)/(tq+tqe).max()))
                ax[2,0].scatter(hn,pos,tqe,color=cmap(tqe/tqe.max()))
                # plot the thrust data with and without empty
                ax[0,1].scatter(hn,pos,tt,color=cmap(tt/tt.max()))
                ax[1,1].scatter(hn,pos,tt+tte,color=cmap((tt+tte)/(tt+tte.max())))
                ax[2,1].scatter(hn,pos,tte,color=cmap(tte/tte.max()))
            else:
                ax[0,0].plot(pos,tq,c=cmap(fi/nf))
                ax[1,0].plot(pos,tq+tqe,c=cmap(fi/nf))
                ax[2,0].plot(pos,tqe,c=cmap(fi/nf))
                # plot the thrust data with and without empty
                ax[0,1].plot(pos,tt,c=cmap(fi/nf))
                ax[1,1].plot(pos,tt+tte.flatten(),c=cmap(fi/nf))
                ax[2,1].plot(pos,tte,c=cmap(fi/nf))
        # extract specific keys        
        else:
            if isinstance(vmax,float):
                clim = vmax
            # iterate over keys
            for ki,kk in enumerate(keys):
                # extract vector
                signal = data[kk].values.flatten()
                if vmax == 'max_each':
                    clim = signal.max()
                elif vmax == 'max':
                    clim = cdict[kk]
                # if plotting as 3d
                if as_3d:
                    col = cmap(signal/clim)
                    np.nan_to_num(col,copy=False)
                    ax[ki].scatter(hn,pos,signal,color=col)
                # if plotting as 2d
                else:
                    ax[ki].plot(pos,signal,color=cmap(fi/nf))
    # set the labels for 3d
    if as_3d:
        if keys == "all":
            ax[0,0].set(xlabel="Hole Number",ylabel="Position (mm)",zlabel="Torque (A)",title="Torque (No Empty)")
            ax[1,0].set(xlabel="Hole Number",ylabel="Position (mm)",zlabel="Torque (A)",title="Torque + Empty")
            ax[2,0].set(xlabel="Hole Number",ylabel="Position (mm)",zlabel="Torque (A)",title="Torque Empty Only")

            ax[0,1].set(xlabel="Hole Number",ylabel="Position (mm)",zlabel="Thrust (A)",title="Thrust (No Empty)")
            ax[1,1].set(xlabel="Hole Number",ylabel="Position (mm)",zlabel="Thrust (A)",title="Thrust + Empty")
            ax[2,1].set(xlabel="Hole Number",ylabel="Position (mm)",zlabel="Thrust (A)",title="Thrust Empty Only")
        else:
            for kk,aa in zip(keys,ax):
                aa.set(xlabel="Hole Number",ylabel="Position (mm)",zlabel=kk,title=kk)
    # set the labels for 2d
    else:
        if keys == "all":
            ax[0,0].set(xlabel="Position (mm)",ylabel="Torque (A)",title="Torque (No Empty)")
            ax[1,0].set(xlabel="Position (mm)",ylabel="Torque (A)",title="Torque + Empty")
            ax[2,0].set(xlabel="Position (mm)",ylabel="Torque (A)",title="Torque Empty Only")

            ax[0,1].set(xlabel="Position (mm)",ylabel="Thrust (A)",title="Thrust (No Empty)")
            ax[1,1].set(xlabel="Position (mm)",ylabel="Thrust (A)",title="Thrust + Empty")
            ax[2,1].set(xlabel="Position (mm)",ylabel="Thrust (A)",title="Thrust Empty Only")
        else:
            for kk,aa in zip(keys,ax):
                aa.set(xlabel="Position (mm)",ylabel=kk,title=kk)
    # set the figure title    
    f.suptitle(f"{kwargs.get('title','Setitec Data History')} ({cmap_str})")
    return f

def plotSetitecBurrEntropy(path,burr,**kwargs):
    from dataparser import loadSetitecXls
    from scipy.stats import entropy
    ent = [entropy(np.unique(loadSetitecXls(fn)[-1]['I Torque (A)'].values.flatten(),return_counts=True)[1],base=None) for fn in sorted(glob(fn),key=lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[3])]
    f,ax = plt.subplots()
    ax.plot(ent,burr,'bx')
    ax.set(xlabel="Entropy",ylabel="Burr Height (microns)",title=kwargs.get("title","Signal Entropy"))
    return f

def plotSetitecEntropy(path,**kwargs):
    '''
        Plot the entropy of the Torque Signal over the lifetime of the tool

        Entropy is calculated from unique value passed to scipy.stats.entropy
        Files are sorted by the integer at the end of the filename

        Inputs:
            path : Wildcard path to where a set of Setitec XLS files are located
            title : Axes title. Default Signal Energy

        Returns generated figure object
    '''
    from dataparser import loadSetitecXls
    from scipy.stats import entropy
    ent = [entropy(np.unique(loadSetitecXls(fn)[-1]['I Torque (A)'].values.flatten(),return_counts=True)[1],base=None) for fn in sorted(glob(fn),key=lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[3])]
    f,ax = plt.subplots()
    ax.plot(ent,'bx')
    ax.set(xlabel="Hole Number",ylabel="Entropy",title=kwargs.get("title","Signal Entropy"))
    return f

def plotSetitecSignalEnergy(path,**kwargs):
    '''
        Plot the signal energy of the Torque Signal over the lifetime of the tool

        Energy is calculated as sum(signal * signal).
        Files are sorted by the integer at the end of the filename

        Inputs:
            path : Wildcard path to where a set of Setitec XLS files are located
            title : Axes title. Default Signal Energy

        Returns generated figure object
    '''
    from dataparser import loadSetitecXls
    energy = [np.sum(loadSetitecXls(fn)[-1]['I Torque (A)'].values.flatten()**2) for fn in sorted(glob(fn),key=lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[3])]
    f,ax = plt.subplots()
    ax.plot(energy,'bx')
    ax.set(xlabel="Hole Number",ylabel="Signal Energy",title=kwargs.get("title","Signal Energy"))
    return f

def plotCompareSetitecEnergy(*args,**kwargs):
    '''
        Plot the signal energy of the Torque Signal for multiple directories
        on the same axis

        Energy is calculated as sum(signal * signal). 
        Files are sorted by the integer at the end of the filename. Meant to compare
        the results from different experiments for shared features.

        If no labels are given, letters of the alphabet are assigned (i.e. A,B C etc.)

        Inputs:
            *args: Several wildcard paths for different Setitec sets
            title : Axes title. Default Signal Energy
            labels : List of labels for each dataset.

        Returns generated figure object
    '''
    from dataparser import loadSetitecXls
    def _energy(path):
        return [np.sum(loadSetitecXls(fn,version="auto_data")['I Torque (A)'].values.flatten()**2) for fn in sorted(glob(path),key=lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[3])]
    if kwargs.get('labels',[]):
        ee = {ll : _energy(fn) for ll,fn in zip(kwargs.get('labels',[]),args)}
    else:
        ee = {chr(65+i) : _entropy(fn) for i,fn in enumerate(args)}

    f,ax = plt.subplots()
    for kk,vv in ee.items():
        ax.plot(np.linspace(0.0,1.0,len(vv)),vv,label=kk)
    ax.set(xlabel="Normalized Tool Life",ylabel="Signal Energy",title=kwargs.get("title","Compare Signal Energy"))
    ax.legend()
    return f

def plotCompareSetitecEntropy(*args,normy=False,**kwargs):
    '''
        Plot the signal entropy of the Torque Signal for multiple directories
        on the same axis

        Entropy is calculated from unique value passed to scipy.stats.entropy
        Files are sorted by the integer at the end of the filename. Meant to compare
        the results from different experiments for shared features.

        If no labels are given, letters of the alphabet are assigned (i.e. A,B C etc.)

        Inputs:
            *args: Several wildcard paths for different Setitec sets
            title : Axes title. Default Compare Signal Entropy

        Returns generated figure object
    '''
    from dataparser import loadSetitecXls
    from scipy.stats import entropy
    def _entropy(path):
        return [entropy(np.unique(loadSetitecXls(fn,version="auto_data")['I Torque (A)'].values.flatten(),return_counts=True)[1],base=None) for fn in sorted(glob(path),key=lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[3])]
    if kwargs.get('labels',[]):
        ee = {ll : _entropy(fn) for ll,fn in zip(kwargs.get('labels',[]),args)}
    else:
        ee = {chr(65+i) : _entropy(fn) for i,fn in enumerate(args)}

    f,ax = plt.subplots()
    for kk,vv in ee.items():
        ax.plot(np.linspace(0.0,1.0,len(vv)),vv,label=kk)
    ax.set(xlabel="Normalized Tool Life",ylabel="Signal Entropy",title=kwargs.get("title","Compare Signal Entropy"))
    ax.legend()
    return f

def plotSetitecMaxEmptyHistory(path):
    '''
        Plot the max Setitec Thrust and Torque Empty Values

        The filenames are sorted by global headcount.

        Inputs:
            path : Wildcard path to Setitec XLS files

        Returns figure objects
    '''
    # create empty lists to hold max values
    tq_empty = []
    tt_empty = []
    # sort files by global head count
    for fn in sorted(glob(path),key=lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[3]):
        # load data files
        data = loadSetitecXls(fn)[-1]
        # find the max torque and thrust empty files
        tq_empty.append(data['I Torque Empty (A)'].values.flatten().max())
        tt_empty.append(data['I Thrust Empty (A)'].values.flatten().max())
    # create axes
    f,ax = plt.subplots()
    # plot the max torque empty values as a blue line
    ax.plot(tq_empty,'b-',label="Torque Empty")
    # set the labels
    ax.set(xlabel="Hole Count",ylabel="Torque Empty (A)")
    # create a twin axes for the thrust data    
    cax = ax.twinx()
    # plot the max thrust empty values as a red line
    cax.plot(tt_empty,'r-',label="Thrust Empty")
    # set y label
    cax.set_ylabel("Thrust Empty (A)")
    # set figure title
    f.suptitle("Setitec Max Empty Signal History (Torque=Blue,Thrust=Red)")
    return f

def plotSetitecMaxHistory(path):
    '''
        Plot the max Setitec Thrust and Torque Values

        The filenames are sorted by global headcount.

        Inputs:
            path : Wildcard path to Setitec XLS files

        Returns figure objects
    '''
    # create empty lists to hold max values
    tq_empty = []
    tt_empty = []
    # sort files by global head count
    for fn in sorted(glob(path),key=lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[3]):
        # load data files
        data = loadSetitecXls(fn)[-1]
        # find the max torque and thrust empty files
        tq_empty.append(data['I Torque (A)'].values.flatten().max())
        tt_empty.append(data['I Thrust (A)'].values.flatten().max())
    # create axes
    f,ax = plt.subplots()
    # plot the max torque empty values as a blue line
    ax.plot(tq_empty,'b-',label="Torque Empty")
    # set the labels
    ax.set(xlabel="Hole Count",ylabel="Torque Empty (A)")
    # create a twin axes for the thrust data    
    cax = ax.twinx()
    # plot the max thrust empty values as a red line
    cax.plot(tt_empty,'r-',label="Thrust Empty")
    # set y label
    cax.set_ylabel("Thrust Empty (A)")
    # set figure title
    f.suptitle("Setitec Max Signal History (Torque=Blue,Thrust=Red)")
    return f

def plotSetitecArray(program,nr=2,sep_plots=False):
    '''
        Plot the Setitec array on a set of subplots

        Inputs:
            program : Panda Dataframe read using processSetitecXls from dataparser
            nr : Number of rows to plot the data on. Number of columns inferred
            sep_plots : Flag to create separate plot for each column

        Returns figure object or list of figure objects if sep_plots is True
    '''
    # get the default figure size
    fz = plt.rcParams.get('figure.figsize')
    # if there are going to be more than 10 columns therefore more than 10 subplots
    # increase figure size
    if program.shape[1]>10:
        fz = [ss*2.7 for ss in fz]
    # get num cols to plot
    nc = (program.shape[1]-1)//nr
    xl = program.columns[0]
    # get x data
    xdata = program[program.columns[0]].to_numpy()
    # create separate plots
    if sep_plots:
        figs = []
        for cc in program.columns[1:]:
            f,ax = plt.subplots(figsize=fz)
            ax.plot(xdata,program[cc].to_numpy(),label=cc)[0]
            ax.set_xlabel(xl)
            ax.set_ylabel(cc)
            ax.set_title(f"Program for {cc}")
            ax.legend(loc="upper right")
            figs.append(f)
        return figs
    else:
        # create figure
        f,ax = plt.subplots(nrows=nr,ncols=nc,figsize=fz)
        # iterate over the other program columns, 
        for cc,aa in zip(program.columns[1:],ax.ravel()):
            # make a copy of the xdata
            xdata_cp = np.copy(xdata)
            # get the ydata
            data = program[cc].to_numpy()
            # plot data
            aa.plot(xdata_cp,data,label=cc)
            aa.set_xlabel(xl)
            aa.set_ylabel(cc)
            aa.set_title(cc)
        plt.subplots_adjust(left=0.05, bottom=0.05, right=0.95, top=0.95, wspace=0.2, hspace=0.3)
        return f

def hasMultipleTools(sheet):
    '''
        Checks if the measurement DataFrame contaisn data about multiple tools

        Checks for presence of NTOOL fields

        Returns bool
    '''
    # convert to numpy array
    # get first row as it contains the actual column names
    sheet_data = sheet.values[0].astype(str)
    return 'NTOOL' in sheet_data

def plotMeasurementsXLS(path,use_sp=True):
    '''
        Plot measurement matrix spreadsheet

        Inputs:
            path : File path to measurement XLXS file
            use_sp : Flag to use subplots for each measurement. Default True.
    '''
    from openpyxl import load_workbook
    # load workbook
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    # list of figure objects to return to user
    figs = []
    # iterate over workbook sheetname
    # iterate over sheets
    for sn in wb.sheetnames:
        # get sheet data
        sheet_data = wb[sn]
        # convert to Pandas DataFrame
        sheet_data = pd.DataFrame(sheet_data.values)
        # check if empty
        if sheet_data.empty:
            print(f"Skipping sheet {sn} as it's empty!")
            continue
        # get actual col titles
        cnames = sheet_data.values[0].astype(str)
        # for summary files where all measurements are in a single file
        # it has a separate columns for tool ID and cycle ID
        ntidx = -1
        ncyidx = -1
        # check if multiple tools are stored in the file
        if 'NTOOL' in cnames:
            # get columns index of NTOOL
            ntidx = np.where(cnames == 'NTOOL')[0][0]
            # get index of NCYCLE
            ncyidx = np.where(cnames == 'NCYCLE')[0][0]
            # get the unique tools
            utools = np.unique(sheet_data[ntidx][1:]).astype(str)
        # if not
        else:
            # create blank array of tools
            utools = [""]
            # set the measurement array to plot to current sheet data
            meas = sheet_data.iloc[1:]
            ### create cycle codes ###
            cycles = []
            offset = 0
            idx = 0
            for i in range(1,meas.shape[0]+1):
                idx += 1
                cycle = offset + idx
                cycles.append(cycle)
                if (i%5)==0:
                    idx = 0
                    offset += 100
        # get index of CFRP_0
        cfrpidx = np.where(cnames == 'CFRP_0')[0][0]
        cnames = cnames[cfrpidx:]
        # for each of the unique tools
        for ut in utools:
            # if the index was assigned
            if ntidx != -1:
                # get the data for the unique tool
                ut_data = sheet_data.loc[sheet_data[ntidx] == ut]
                # get the cycle data to act as the x-axis
                cycles = ut_data.iloc[:,ncyidx].values.astype(np.int64)
                # get measurement data
                meas = ut_data.iloc[:,cfrpidx:-1]
                meas
            # if using subplots for each measurement
            if use_sp:
                # create subplots
                f,ax = plt.subplots(ncols=5)
                # iterate over measurements and axes
                for cc,aa,mtitle in zip(meas.columns,ax.ravel(),cnames):
                    if mtitle == 'NOTES':
                        continue
                    # measurement data
                    mdata = meas[cc].values.astype(np.float64)
                    # plot measurement
                    aa.plot(cycles,mdata,'x')
                    aa.set_xlabel("Cycle Number")
                    aa.set_ylabel("Measurement (mm)")
                    aa.set_title(mtitle)
                plt.subplots_adjust(left=0.05, bottom=0.05, right=0.95, top=0.95, wspace=0.2, hspace=0.3)
            else:
                # create single plot
                f,ax = plt.subplots()
                # iterate over measurements
                for cc in meas.columns:
                    # get measurement title
                    mtitle = meas[cc][0]
                    if mtitle == 'NOTES':
                        continue
                    mdata = meas[cc][1:].values.astype(np.float64)
                    # plot values
                    ax.plot(cycles,mdata,'x',label=mtitle)
                # display legend
                plt.legend(loc="upper right")
                # assign plotting axis label
                ax.set_xlabel("Cycle Number")
                ax.set_ylabel("Diameter (mm)")
            # assign title
            f.suptitle(f"Measurements from tool {ut}")
            # store in list
            figs.append(f)
    return figs

def plotTDMS(path,use_subs=True,use_times=True):
    '''
        Plot the data stored in a TDMS file

        Loads a TDMS file and iterates over the channels and plots them.

        Inputs:
            path : Path to TDMS file or TdmsFile file object
            use_sep : Flag to use subplots for each measurement. If False, a separate Figure is created.
            use_times : Flag to use time index for plot instead of data index.

        Returns list of matplotlib.Figure objects
    '''
    import nptdms
    # if the user gave a path open the file
    if isinstance(path,str):
        file = nptdms.TdmsFile(path)
    # if the user gave the file object directly
    elif isinstance(path,nptdms.tdms.TdmsFile):
        file = path
    # list for figure objects
    figs = {}
    # iterate over upper groups
    for gg in file.groups():
        # if using separate subplots
        # create subplots and axes array
        if use_subs:
            ww,hh = plt.rcParams.get('figure.figsize')
            fz = [ww*4.0,hh*2.7]
            f,ax = plt.subplots(nrows=2,ncols=4,tight_layout=True,figsize=fz)
            ax = ax.flatten()
            # set figure title
            f.suptitle(f"Measurements for {gg.name}")
            plt.subplots_adjust(left=0.05, bottom=0.05, right=0.95, top=0.95, wspace=0.2, hspace=0.3)
            # append figure to list
            figs[gg.name] = f
        else:
            ax = [None,]*len(gg.channels())
            figs[gg.name] = {}
        # iterate over channels
        for cc,units,aa in zip(gg.channels(),[cc.properties["unit_string"] for cc in gg.channels()],ax):
            # convert to dataframe
            data = cc.as_dataframe(time_index=use_times)
            # if using separate subplots
            if use_subs:
                # plot data
                aa.plot(data.index.values,data.values.flatten())
                # set labels
                aa.set_xlabel("Time (s)" if use_times else "Index")
                aa.set_ylabel(f"{cc.name} ({units})")
                aa.set_title(f"Plot of {cc.name}")
            # if creating separate figure for each measurement
            else:
                f,axis = plt.subplots(tight_layout=True)
                axis.plot(data.index.values,data.values.flatten())
                f.suptitle(f"Measurement {cc.name} from group {gg.name}")
                axis.set_xlabel("Time (s)" if use_times else "Index")
                axis.set_ylabel(f"{cc.name} ({units})")
                figs[gg.name][cc.name] = f
    # returns figures
    return figs

def plotTDMSCWS(path,use_times=True,scales=None,wavelet='morl',T = 1/1e5,period=1):
    '''
        Plot the data stored in a TDMS file

        Loads a TDMS file and iterates over the channels and plots them.

        Inputs:
            path : Path to TDMS file or TdmsFile file object
            scales : Scales to use in wavelet
            wavelet : Target wavelet to use
            T : Sampling period
            period : Window period

        Returns list of matplotlib.Figure objects
    '''
    import nptdms
    # if the user gave a path open the file
    if isinstance(path,str):
        file = nptdms.TdmsFile(path)
    # if the user gave the file object directly
    elif isinstance(path,nptdms.tdms.TdmsFile):
        file = path
    coi = {
        'alpha':0.5,
        'hatch':'/',
    }
    # iterate over upper groups
    for gg in file.groups():
        # process group name for saving
        gg_name = gg.name.replace('\\','-')
        # iterate over channels
        for cc,units in zip(gg.channels(),[cc.properties["unit_string"] for cc in gg.channels()]):
            # process channel name for saving
            cc_name = cc.name.replace('\\','-')
            # convert to dataframe
            data = cc.as_dataframe(time_index=use_times)
            # create time vector
            time = np.arange(0.0,len(data)*T,T,dtype='float16')
            # chunks
            tbits = list(range(int(np.floor(time.min())),int(np.ceil(time.max())),period))
            # iterate over windows pairwise
            for A,B in zip(tbits,tbits[1:]):
                # get time indices for target time period
                tchunk = time[(time>=A) & (time<=B)]
                # get data indices for target time period
                chunk = data.values.flatten()[(time>=A) & (time<=B)]
                # create axes
                f,axis = plt.subplots(tight_layout=True)
                # set title                
                f.suptitle(f"Measurement {cc.name} from group {gg.name}")
                # plot CWS
                axis = scg.cws(tchunk,chunk,scales,wavelet=wavelet,
                             xlabel="Time (s)" if use_times else "Index",ylabel="Scale",yaxis="scale",title=f"Plot of {cc.name}",
                             yscale='log',coikw=coi,ax=axis)
                # yield the figure for saving or viewing
                yield f,(A,B),cc_name,gg_name

def icwt(coef,scales,wtype='morl'):
    ''' inverse continuous wavelet from https://github.com/PyWavelets/pywt/issues/328 '''
    import pywt
    mwf = pywt.ContinuousWavelet(wtype).wavefun()
    y_0 = mwf[0][np.argmin(np.abs(mwf[1]))]

    r_sum = np.transpose(np.sum(np.transpose(coef)/ scales ** 0.5, axis=-1))
    return r_sum * (1 / y_0)

def filterCWS(data,filters,time=None,scales=None,wavelet='morl',T = 1/1e5,period=1,new_mag='min',periods=None):
    from scaleogram.wfun import fastcwt
    # check that filters are given
    if not filters:
        raise ValueError("Filters must be not None!")
    # check if length of filters is non-zero
    if len(filters)==0:
        raise ValueError("List of filters must have entries!")
    if time is None:
        # creat time vector
        time = np.arange(0.0,len(data)*T,T,dtype='float16')
    # if scales aren't specified
    if scales is None:
        # create scales based on time
        scales = np.arange(1, min(len(time)/10, 100))
    # calculate CWT
    coefs,_ = fastcwt(data,scales,wavelet,T)
    # iterate over the target scales
    for sc in filters:
        # if the user gave a two element range
        # filter using the range
        if len(sc) ==2:
            si=np.where((scales >=sc[0]) & (scales <= sc[1]))[0]
        # if a single element then filter from value onwards
        else:
            si=np.where(scales >=sc[0])[0]
        # set the target scale range to the target value
        if new_mag == 'min':
            coefs[si,:] = coefs.min()
        elif new_mag == 'max':
            coefs[si,:] = coefs.max()
        else:
            coefs[si,:] = new_mag
        coefs[si,:] = 0.5*coefs[si,:]
    # perform inverse
    return icwt(coefs,scales)

def filterDWT(data,wavelet='db6',mode='symmetric',level='max',axis=-1):
    import pywt
    if level == 'max':
        level = pywt.dwt_max_level(len(data),pywt.Wavelet(wavelet).dec_len)
    coefs = pywt.wavedec(data,wavelet,mode,level,axis)
    for i in range(1,level+1):
        coefs[i] = np.zeros_like(coefs[i])
    return pywt.waverec(coefs,wavelet,mode,axis)

def plotChangePoints(data,**kwargs):
    '''
        Plot changepoints stored in the given data structure

        Supports generated pandas dataframe and numpy array

        Inputs:
            data : Data array loaded from a MAT file
            single_plot : Plot all on a single plot
            hole_plot : Plot each hole on a separte figure
            plot_title : Figure title. Default Change points or Change points for hole hh, coupon cc depending on other flags
            ylabel : Label used on y axis for each axis. Default Signal.

        Returns figure or list of figures depending on 
    '''
    # get number of holes and columns
    if isinstance(data,pd.DataFrame):
        nh,nc = (data.index.get_level_values('Hole'),
              data.index.get_level_values('Coupon'))
        is_df = True
    elif isinstance(data,np.ndarray):
        nh,nc = data.shape[:2]
        nh,nc = (np.arange(nh),np.arange(nc))
        is_df = False
    else:
        raise ValueError("Don't know how to unpick given data structure!")
    # if user wants a single of all hole and coupon change points
    if kwargs.get('single_plot',False):
        print("single plot")
        # create single giant plot
        # share x-axis so it can be compared
        # default giant figsize
        f,ax = plt.subplots(nrows=int(nh.max())*int(nc.max()),ncols=int(nc.max()),figsize=(20,10),tight_layout=True,sharex=True)
        for hh,cc,aa in zip(nh,nc,ax.flatten()):
            if is_df:
                data_clip = data.loc[(hh,cc)]
                xdata,ydata = data_clip["Time"].values,data_clip["CP"].values
            else:
                data_clip = data[hh,cc,:,:]
                xdata,ydata = data_clip[0,:],data_clip[1,:]
            # iterate over data in pairs
            aa.step(xdata,ydata,where="post")
            # set labels for axes
            aa.set_xlabel("Time (s)")
            aa.set_ylabel(kwargs.get("ylabel","Signal"))
            aa.set_title(f"Hole {hh}, Coupon {cc}")
        plt.legend()
        # set title for figure
        f.suptitle(kwargs.get("plot_title","Change points"))
        return f
    # if the user wants separate plots for each hole
    elif kwargs.get('hole_plots',False):
        print("hole plots")
        # create list to hold figures
        figs = []
        # iterate over the holes
        for hh in nh:
            # create axis
            # coupon axes arranged as 
            f,ax = plt.subplots(nrows=int(nc.max()),tight_layout=True,sharex=True,figsize=(12,15))
            # iterate over coupons and axes
            for cc,aa in zip(nc,ax.flatten()):
                # get data for target hole and coupon
                if is_df:
                    data_clip = data.loc[(hh,cc)]
                    xdata,ydata = data_clip["Time"].values,data_clip["CP"].values
                else:
                    data_clip = data[hh,cc,:,:]
                    xdata,ydata = data_clip[0,:],data_clip[1,:]
                # iterate over data in pairs
                aa.step(xdata,ydata,where="post")
                # set labels for axes
                aa.set_xlabel("Time (s)")
                aa.set_ylabel(kwargs.get("ylabel","Signal"))
                aa.set_title(f"Coupon {cc}")
            # set figure title
            f.suptitle(kwargs.get("plot_title",f"Change points for hole {hh}"))
            figs.append(f)
        return figs
    # if the user wants separate plots for each coupon
    else:
        # create list of figures
        figs = []
        # iterate over holes, coupons and 
        for hh,cc in zip(nh,nc):
            # create axes
            f,ax = plt.subplots(tight_layout=True)
            # get data for target hole and coupon
            if is_df:
                data_clip = data.loc[(hh,cc)]
                xdata,ydata = data_clip["Time"].values,data_clip["CP"].values
            else:
                data_clip = data[hh,cc,:,:]
                xdata,ydata = data_clip[0,:],data_clip[1,:]
            # iterate over data in pairs
            aa.step(xdata,ydata,where="post",label=f"H{int(hh)}C{int(cc)}")
            # set labels for axes
            ax.set_xlabel("Time (s)")
            ax.set_ylabel(kwargs.get("ylabel","Signal"))
            ax.set_title(f"Hole {hh}, Coupon {cc}")
            plt.legend()
            f.suptitle(kwargs.get("plot_title",f"Change points for hole {hh}, coupon {cc}"))
            figs.append(f)
        return f
        
def plot_depthest(dest,**kwargs):
    hh,cc = dest.shape
    # create axes
    f,ax = plt.subplots(figsize=kwargs.get("figsize",(12,10)),constrained_layout=True)
    # get x and y values for axes
    # if not specified, then values are set from matrix dimensions
    xt = [0,] + kwargs.get('xticks',np.arange(0,hh))
    yt = [0,] + kwargs.get('yticks',np.arange(0,cc))
    #ext = [min(xt),max(xt),max(yt),min(yt)]
    # plot the data as an image
    # aspect ration is set to equal to create a square matrix
    im = ax.imshow(dest,cmap=kwargs.get("cmap","Blues"),aspect="equal")
    ax.set_xticklabels([str(ii) for ii in xt])
    ax.set_yticklabels([str(ii) for ii in yt])
    # set labels
    # default labels are Coupon and Hole as that was their original purpose
    ax.set_xlabel(kwargs.get("xlabel","Coupon"))
    ax.set_ylabel(kwargs.get("ylabel","Hole"))
    f.suptitle(kwargs.get('title',"Stats about depth estimation"))
    # mark each cell of the image with the values
    # written in white font, 2 dp
    #dh = np.diff(xt)[0]//2
    #dc = np.diff(yt)[0]//2
    for i in range(hh):
        for j in range(cc):
            ax.text(i, j, f"{dest[i, j]:.2f}",ha="center", va="center", color=kwargs.get("col",'k'))
    if kwargs.get('flipy',False):
        ax.invert_yaxis()
    # add colorbar
    divider = make_axes_locatable(ax)
    cax = divider.append_axes("right", size="5%", pad=0.05)
    f.colorbar(im, cax=cax)
    f.tight_layout()
    return f,ax

def sortFiles(fn):
    fpath = os.path.splitext(os.path.basename(fn))[0]
    pts = fpath.split('_')
    if (int(pts[-1]) == 1) and (int(pts[-2]) != 1):
        return int(pts[-2])
    else:
        return int(pts[-1])
    return int(pts[-3]) if len(pts)>4 else int(pts[-2])

def plotSummary(paths,sort=True,**kwargs):
    '''
        Plot a summary of a set of new files
        
        This function creates a series of plots of useful statistics for new files that need to be analysed.
        All statistics are plotted as a scatter and histogram plot
        
        Current statistics
            - Energy
            - Max Torque
            - Max Position
            - Peak Position
            
        More may be added as needs arise
        
        Filenames are based on the directory and wildcard portion of the path
        
        e.g.
        plotSummary(r"\\FR0-VSIAAS-5706\StNazaire\SETITEC\370A-CAB-CFAR-2-1-D\*108-V3-102*.xls")
        
        Would produces files containing 370A-CAB-CFAR-2-1-D-108-V3-102 followed by the respective metric and hist if it's a histogram
        
        Inputs:
            paths : Wildcard path
            sort : Attempt to sort files
    '''
    import energy_estimation as em
    
    dirname = os.path.split(os.path.dirname(paths))[-1]
    fsearch = os.path.splitext(os.path.basename(paths))[0].strip('*')
    title = dirname + ' ' + fsearch
    fname = dirname + '-' + fsearch
    
    paths = glob(paths)
    if sort:
        paths = list(filter(lambda x : sortFiles(x)!=0,sorted(paths,key=sortFiles)))
    # create figures
    ee = []
    for fn in paths:
        data_all = loadSetitecXls(fn,"auto_data")
        ee.append(em.get_energy_estimation(data_all['I Torque (A)'].values.flatten(),data_all['I Torque Empty (A)'].values.flatten(),data_all['I Thrust (A)'].values.flatten()))
    f,ax = plt.subplots()
    ax.plot(ee,'x')
    ax.set(xlabel="Hole Number",ylabel="Energy (J)",title=title +" Energy")
    f.savefig(f"{fname}-energy.png")
    plt.close(f)
    
    f,ax = plt.subplots()
    ax.hist(ee,20,facecolor='g',alpha=0.5,edgecolor='black',linewidth=1.2)
    ax.set(xlabel="Energy (J)",ylabel="Counts",title=title +" Energy")
    f.savefig(f"{fname}-energy-hist.png")
    plt.close(f)
    
    ee.clear()
    for fn in paths:
        data_all = loadSetitecXls(fn,"auto_data")
        ee.append((data_all['I Torque (A)'].values.flatten()+data_all['I Torque Empty (A)'].values.flatten()).max())
    f,ax = plt.subplots()
    ax.plot(ee,'x')
    ax.set(xlabel="Hole Number",ylabel="Max Torque (A)",title=title +" Max Torque")
    f.savefig(f"{fname}-max-torque.png")
    plt.close(f)
    
    f,ax = plt.subplots()
    ax.hist(ee,20,facecolor='g',alpha=0.5,edgecolor='black',linewidth=1.2)
    ax.set(xlabel="Max Torque (A)",ylabel="Counts",title=title +" Max Torque")
    f.savefig(f"{fname}-max-torque-hist.png")
    plt.close(f)
    
    ee.clear()
    for fn in paths:
        data_all = loadSetitecXls(fn,"auto_data")
        ee.append(np.abs(data_all['Position (mm)'].values.flatten()).max())
    f,ax = plt.subplots()
    ax.plot(ee,'x')
    ax.set(xlabel="Hole Number",ylabel="Max Position (mm)",title=title +" Max Position")
    f.savefig(f"{fname}-max-position.png")
    plt.close(f)
    
    f,ax = plt.subplots()
    ax.hist(ee,'auto',facecolor='g',alpha=0.5,edgecolor='black',linewidth=1.2)
    ax.set(xlabel="Max Position (mm)",ylabel="Counts",title=title +" Max Position")
    f.savefig(f"{fname}-max-position-hist.png")
    plt.close(f)
    
    ee.clear()
    for fn in paths:
        data_all = loadSetitecXls(fn,"auto_data")
        ee.append(np.abs(data_all['Position (mm)'].values.flatten())[(data_all['I Torque (A)'].values.flatten()+data_all['I Torque Empty (A)'].values.flatten()).argmax()])
    f,ax = plt.subplots()
    ax.plot(ee,'x')
    ax.set(xlabel="Hole Number",ylabel="Peak Position (mm)",title=title +" Peak Position")
    f.savefig(f"{fname}-peak-position.png")
    plt.close(f)
    
    f,ax = plt.subplots()
    ax.hist(ee,20,facecolor='g',alpha=0.5,edgecolor='black',linewidth=1.2)
    ax.set(xlabel="Peak Position (mm)",ylabel="Counts",title=title +" Peak Position")
    f.savefig(f"{fname}-peak-position-hist.png")
    plt.close(f)

if __name__ == "__main__":
##    f_loc,fmax = plotBatchSetitecSpectrogram("8B life test/*.xls",fclip=0.3,clip_mode='lt',opath='8B life test/plots')
##    f_loc.savefig("8B life test/plots/8B-life-test-stft-max-response-fclip-lt-0.3.png")
##    fmax.savefig("8B life test/plots/8B-life-test-signal-vs-stft-max-response-fclip-lt-0.3.png")
##
##    f_loc,fmax = plotBatchSetitecSpectrogram("8B life test/*.xls",fclip=5.0,clip_mode='gt',opath='8B life test/plots')
##    f_loc.savefig("8B life test/plots/8B-life-test-stft-max-response-fclip-gt-5.0.png")
##    fmax.savefig("8B life test/plots/8B-life-test-signal-vs-stft-max-response-fclip-gt-5.0.png")
##    plt.show()
    pass
