import pandas as pd
import numpy as np
import json
import os
from _io import TextIOWrapper
import openpyxl
import h5py
from nptdms import TdmsFile
from statistics import stdev
import warnings
from glob import glob

############################################## UTILITIES ##############################################
def printHDF5Struct(path,to_file=True):
    '''
        Pretty print the structure and contents of a HDF5 file

        Also works on >=v7.3 MAT files as they can be read like HDF5 files

        For known large files or MAT files, it is recommended to redirect the contents to a log file for easier reading

        Inputs:
            path : Path to target MAT or HDF5 file
            to_file : Flag or string. If True, then the structure is dumped to a text file with the same name as source.
                    If string, then that is used as the output file path
    '''
    # if the user wants to dump the structure to a file
    if to_file:
        # if the user specified an output file
        if isinstance(to_file,str):
            opath = to_file
        # if the user simply specified True
        # infer output path from file
        elif to_file:
            # if the source is a h5py.File object
            if isinstance(path,h5py._hl.files.File):
                opath = os.path.splitext(os.path.basename(path.filename))[0]+".txt"
            elif isinstance(path,str):
                opath = os.path.splitext(os.path.basename(path))[0]+".txt"
        # create log filepointer
        fp = open(opath,'w')
        # create version of function where it's written to a file
        def _print_item(name, item):
            fp.write(name+"\n")
            # Format item attributes if any
            if item.attrs:
                fp.write('\tattributes:\n')
                for key, value in item.attrs.items():
                    fp.write('\t\t{}: {}\n'.format(key, str(value).replace('\n', '\n\t\t')))

            # Format Dataset value
            fp.write('\tValue:\n')
            fp.write('\t\t' + str(item).replace('\n', '\n\t\t') + '\n')
    else:
        def _print_item(name, item):
            print(name)
            # Format item attributes if any
            if item.attrs:
                print('\tattributes:')
                for key, value in item.attrs.items():
                    print('\t\t{}: {}'.format(key, str(value).replace('\n', '\n\t\t')))

            # Format Dataset value
            print('\tValue:')
            print('\t\t' + str(item).replace('\n', '\n\t\t'))

    # if the user gave a file path
    if isinstance(path,str):
        with h5py.File(path,'r') as hdf_file:
            # Here we first print the file attributes as they are not accessible from File.visititems()
            _print_item(hdf_file.filename, hdf_file)
            # Print the content of the file
            hdf_file.visititems(_print_item)
    # if the user has given something that's HDF5 accessible
    else:
        # if it's something that has a filename attribute
        if hasattr(path,'filename'):
            # Here we first print the file attributes as they are not accessible from File.visititems()
            _print_item(path.filename, path)
        # else print the path as object followed bu the string representation of it
        else:
            _print_item("Object",path)
        # Print the content of the file
        path.visititems(_print_item)

def convertSizeToUnits(sz):
    '''
        Takes the given size in bytes and converts it to the largest memory denominator

        Input:
            sz : Total size in bytes

        Returns size and units
    '''
    sfx=['B','KB','MB','GB','TB']
    total_sz = sz
    sfi = 0
    while (total_sz>1024) and (sfi < (len(sfx)-1)):
        sfi += 1
        total_sz = total_sz/1024.0
    return total_sz,sfx[sfi]

def printTDMSStruct(path,del_file=True,to_log=True):
    '''
        Print the TDMS file structure when parsed to a HDF5 file structure

        Wrapper around printHDF5Struct

        Loads TDMS file, creates HDF5 file and then passes the ref to printHDF5Struct

        Inputs:
            del_file : Delete HDF5 file
            to_log : Flag or string for sending result to log file. If a Flag and True, then the
                    filepath is used for the logfile. If a string, then that's used as the log filepath
    '''
    # get filename
    fname = os.path.splitext(os.path.basename(path))[0]
    # load in file
    with TdmsFile.read(path) as file:
        # convert to hdf5
        with file.as_hdf(fname+".hdf5") as source:
            printHDF5Struct(source,to_log)
    # delete file if wanted
    if del_file:
        if os.path.exists(fname+".hdf5"):
            os.remove(fname+".hdf5")

def printTDMSInfo(path,to_file=True):
    '''
        Calls the tdmsinfo command on the target file and dumps the info to
        either the terminal or a file.

        The tdmsinfo command is installed with the nptdms package

        Inputs:
            path : File path to TDMS file
            to_file : Flag or file path to place the results.
                if False, then it is returned to the user
                if a str, then it is treated as the output file path
                if True, then is is placed in a text file with the same name as the file
    '''
    import subprocess
    # get filename
    fname = os.path.splitext(os.path.basename(path))[0]
    td_struct = subprocess.check_output(["tdmsinfo","-p",path])
    td_str = td_struct.decode("utf-8")
    if not to_file:
        return td_str
    else:
        # if the user specified an output file
        if isinstance(to_file,str):
            open(to_file,'w').write(td_str)
        elif isinstance(to_file,bool):
            open(f"{fname}.txt",'w').write(td_str)

def get_hole_coupon(path,tool):
    '''
        Get hole and coupon from the path

        Extracts the coupon and hole from the file path, assuming the
        standardised XLS file paths

        Inputs:
            path : Input file paths
            tool : Target tool. Used in parsing the coupon. If None,
                    then the first two characters are used.

        Returns the coupon and hole as integers
    '''
    # get the filename without the file extension
    fname = os.path.splitext(os.path.basename(path))[0]
    # split into parts by underscores
    pts = fname.split('_')
    # part 1 has the coupon number
    # 2nd part has the hole number
    return int(pts[0].split(tool)[1]), int(pts[1].split('H')[1])

def max_hole_coupon(path,tool):
    '''
        Find the max hole and coupon values from the wildcard paths.

        Iterates over the paths building a list of hole and coupons. The max
        values of each are then returned.

        WARNING: This is designed to work with the Manchester dataset where the info
        is stored in the pathname and the tool IDs (e.g. UB, UC, UD etc) are known and
        supported.

        Inputs:
            path : Input wildcard path
            tool : Tool used. Passed to get_hole_coupon
    '''
    all_hc = list([get_hole_coupon(pp,tool) for pp in glob(path)])
    return max(all_hc,key=lambda x : x[0]),max(all_hc,key=lambda x : x[1])

def iterTDMSFile(path,repchar='-',ret_units=True):
    '''
        Iterate over the datasets in the target TDMS file

        The TDMS file is opened using a with statement and each group, channel is iterated
        over. The found dataset along wtih group name and channel name are returned.

        Slashes in group and channel name are replaced with repchar so they can be used for plotting, saving etc.
        If repchar is None, then the unprocessed group and channel names are returned

        Inputs:
            path : Filepath to TDMS file
            repchar : Char to replace \\ in group and channel names. Default '-'
            ret_units : Returns units string from metadata. Default True.

        Yields processed group name, processed channel name and the datasets as a Pandas dataframe. If ret_units is True,
        the units string from metadata.
    '''
    with TdmsFile(path) as file:
        # iterate over each group        
        for gg in file.groups():
            gg_name = gg.name
            # format the group name
            if repchar is not None:
                gg_name = gg_name.replace('\\',repchar)
            # iterate over each channel
            for cc,units in zip(gg.channels(),[cc.properties["unit_string"] for cc in gg.channels()]):
                cc_name = cc.name
                # format the channel name
                if repchar is not None:
                    cc_name = cc_name.replace('\\',repchar)
                if ret_units:
                    yield gg_name,cc_name,cc.as_dataframe(time_index=True),units
                else:
                    yield gg_name,cc_name,cc.as_dataframe(time_index=True)
############################################## SPREADSHEETS ##############################################
def _replaceLast(ss,c,r):
    # if the character is not in the string to begin with
    # return the original string
    if not (c in ss):
        return ss
    # find where last c occured
    ii = ss[::-1].find(c)
    # convert string to list of characters
    tt = list(ss)
    # replace last comma with dot
    tt[-ii-1] = r
    # reform to string
    return "".join(tt)

def _formatValue(ss,c=',',r='.'):
    res = _replaceLast(ss,c,r).replace(',','.')
    try:
        return float(res)
    except ValueError:
        return res

def getSetitecXLSGlobalHead(open_file):
    '''
        Search metadata of Setitec XLS for global head counter

        The global head counter is stored in the first batch of metadata in the file.
        The function searches for the column 'Head Global Counter' and returns its value

        Inputs:
            open_file : Path to Setitec XLS file

        Returns integer representing global head counter
    '''

    with open(open_file,'r',encoding='latin-1') as of:
        # skip first line
        of.readline()
        for ll in of:
            # read headers
            head = ll.strip()
            if not ('Counter' in head):
                continue
            head = head.split('\t')
            # find head global counter column number
            for hi,hh in enumerate(head):
                if 'Head Global Counter' in hh:
                    return int(of.readline().strip().split('\t')[hi])

def getSetitecXLSLocalHead(open_file):
    '''
        Search metadata of Setitec XLS for local head counter

        The function searches for the column 'Head Local Counter' and returns its value in the next row

        Inputs:
            open_file : Path to Setitec XLS file

        Returns integer representing global head counter
    '''
    with open(open_file,'r',encoding='latin-1') as of:
        # skip first line
        of.readline()
        for ll in of:
            # read headers
            head = ll.strip()
            if not ('Counter' in head):
                continue
            head = head.split('\t')
            # find head global counter column number
            for hi,hh in enumerate(head):
                if 'Head Local Counter' in hh:
                    return int(of.readline().strip().split('\t')[hi])

def getSetitecXLSFirmwareVers(open_file):
    '''
        Search metadata of Setitec XLS for firmware version

        The function searches for the column 'Firmware Version' and returns its value in the next row

        Inputs:
            open_file : Path to Setitec XLS file

        Returns string representing firmware version
    '''
    with open(open_file,'r',encoding='latin-1') as of:
        # skip first line
        of.readline()
        for ll in of:
            # read headers
            head = ll.strip()
            if not ('Firmware' in head):
                continue
            head = head.split('\t')
            # find head global counter column number
            for hi,hh in enumerate(head):
                if 'Firmware' in hh:
                    return of.readline().strip().split('\t')[hi]

def getSetitecXLSHeadTag(open_file):
    '''
        Search metadata of Setitec XLS for Head TAG

        The function searches for the column 'Head TAG' and returns its value in the next row

        Inputs:
            open_file : Path to Setitec XLS file

        Returns string representing Head TAG
    '''
    with open(open_file,'r',encoding='latin-1') as of:
        # skip first line
        of.readline()
        for ll in of:
            # read headers
            head = ll.strip()
            if not ('Head TAG' in head):
                continue
            head = head.split('\t')
            # find head global counter column number
            for hi,hh in enumerate(head):
                if hh == 'Head TAG':
                    return of.readline().strip().split('\t')[hi]

def getSetitecXLSHeadName(open_file):
    '''
        Search metadata of Setitec XLS for Head Name

        The function searches for the column 'Head Name' and returns its value in the next row

        Inputs:
            open_file : Path to Setitec XLS file

        Returns string representing Head Name
    '''
    with open(open_file,'r',encoding='latin-1') as of:
        # skip first line
        of.readline()
        for ll in of:
            # read headers
            head = ll.strip()
            if not ('Head Name' in head):
                continue
            head = head.split('\t')
            # find head global counter column number
            for hi,hh in enumerate(head):
                if hh == 'Head Name':
                    return of.readline().strip().split('\t')[hi]

def getSetitecXLSBoxName(open_file):
    '''
        Search metadata of Setitec XLS for BOX Name

        The function searches for the column 'BOX Name' and returns its value in the next row

        Inputs:
            open_file : Path to Setitec XLS file

        Returns string representing Head Name
    '''
    with open(open_file,'r',encoding='latin-1') as of:
        # skip first line
        of.readline()
        for ll in of:
            # read headers
            head = ll.strip()
            if not ('BOX Name' in head):
                continue
            head = head.split('\t')
            # find head global counter column number
            for hi,hh in enumerate(head):
                if hh == 'BOX Name':
                    return of.readline().strip().split('\t')[hi]

def getSetitecXLSSearch(open_file,search):
    '''
        Search metadata of Setitec XLS for matching target

        The function searches for the column that matches the search term
        and returns the value immediately below it

        Inputs:
            open_file : Path to Setitec XLS file
            search : Search term to search EXACTLY for

        Returns a list of strings representing BOX data
    '''
    found = {}
    with open(open_file,'r',encoding='latin-1') as of:
        # skip first line
        of.readline()
        for ll in of:
            # read headers
            head = ll.strip()
            # if the head doesn't contain any of the target search terms
            # move on
            if all([not (ss in head) for ss in search]):
                continue
            # split the header into parts
            head = head.split('\t')
            # iterate over the parts of teh head
            # find head global counter column number
            for hi,hh in enumerate(head):
                # if the header matches any of the search terms
                if hh == search:
                    found[hh] = of.readline().strip().split('\t')[hi]
                
    return found

def sortSetitecFiles(path,skip_air_files=True):
    '''
        Sort Setitec files by hole number

        Sorts the files by global and local hole number.

        If skip_air_files is True, then the files which have 3 hole numbers (e.g. 21_22_1) are skipped. If False,
        then it is still stored by the global and local hold number (e.g. 21_22_1 -> 21,22)

        Inputs:
            path : Wildcard path or iterator returning paths
            skip_air_files : Skip files that contain 3 hole numbers

        Returns sorted list
    '''
    if isinstance(path,str):
        path = glob(path,recursive=True)
    def is_not_air(f):
        return any([not c.isnumeric() for c in os.path.splitext(os.path.basename(f))[0].split('_')[:-3]])
    if skip_air_files:
        path = filter(is_not_air,path)
    def getHNs(f):
        if is_not_air(f):
            return os.path.splitext(os.path.basename(f))[0].split('_')[-2],os.path.splitext(os.path.basename(f))[0].split('_')[-1]
        else:
            return os.path.splitext(os.path.basename(f))[0].split('_')[-3],os.path.splitext(os.path.basename(f))[0].split('_')[-2]
    return sorted(path,key=getHNs)

def findFirmwareVersion(path):
    '''
        Find the Firmware Version in the file using RE

        Scans file for the first match of a substring matching firmware standard

        Input:
            path: Full path to file

        Returns first string that matches pattern
    '''
    import re
    mt = re.compile('V\s[1-9][.][0-9][.][0-9][.][0-9]*')
    with open(path,'r') as of:
        for line in of:
            finds = mt.findall(line.strip('\n').strip('\t'))
            if finds:
                return finds[0]

def findWriteDate(path,to_dt=False):
    '''
        Finds the first occurance of a datetime that matches the formst

        %Y-%m-%d:%H:%M:%S or %d/%m/%y : %I:%M %p

        This tries to find the write datetime at the top of the file and return it

        Input:
            path : Full path to file
            to_dt : Flag to convert found string to datetime object. Default False.

        Returns found matching string or datetime object depending on to_dt.
    '''
    import re
    from datetime import datetime
    mode = 0
    dt = None
    with open(path,'r') as of:
        # read each line of the file
        for line in of:
            # try and match target expression
            finds = re.compile(r'\d{4}-\d?\d-\d?\d:(?:2[0-3]|[01]?[0-9]):[0-5]?[0-9]:[0-5]?[0-9]').search(line.strip('\n').strip('\t'))
            if finds:
                mode = 0
                dt = finds[0]
                break
            finds = re.compile(r'\d{2}(\/)+\d{2}(\/)+\d{2}(\s)+(:)+(\s)+\d{2}(:)+\d{2}(\s)+\w[AM,PM]+').search(line.strip('\n').strip('\t'))
            if finds:
                mode = 1
                dt = finds[0]
                break
    # if it returned nothing
    if (dt is None):
        return None
    if not to_dt:
        return dt
    return datetime.strptime(dt,"%d/%m/%y : %I:%M %p" if mode==1 else "%Y-%m-%d:%H:%M:%S")

def sortFilesByWriteDate(path,reverse=False):
    '''
        Sort files by the file write datetime at the top of the file

        Input:
            path: Wildcard path or iterator that returns file paths
            reverse : Flag to return the paths in descending order

        Returns paths sorted by write datetime
    '''
    if isinstance(path,str):
        path = glob(path)
    return sorted(path,key=lambda x : findWriteDate(x,True),reverse=reverse)

def getStepCodeSampleCount(fn,sc='all'):
    '''
        Get the number of data samples that have the target step code (sc)

        Step Codes are used to denote stages of the program used in the run data.
        Checking the number of samples in each step code, especially at the start,
        is a decent indicator of if the transition to the next step is happening
        too early/late.

        The target step code can be single value or an iterable collection of
        target values. If it's a single target, then only the number of samples
        is returned. If multiple targets are given, then a dictionary of counts
        and their corresponding values are returned

        Inputs:
            fn : File path to Setitec XLS file
            sc : Single or list of floating numerical step codes to search for.
                Also accepts string all to indicate all step codes in the file.
                Default 1.

        Return a dictionary of target-count pairs if multiple are given
    '''
    data = loadSetitecXls(fn,"auto_data")
    # get unique step codes
    steps_uq = data['Step (nb)'].unique()
    if sc == 'all':
        sc = steps_uq
    elif isinstance(sc,(int,float)):
        if not (sc in steps_uq):
            return {sc : None}
        sc = [float(sc),]
    return {u : data[data['Step (nb)']==u].shape[0] for u in sc}
        
def getStepCodeStartPos(fn,sc='all'):
    '''
        Get the first position value where the target step code (sc) occurs.

        Step Codes are used to denote stages of the program used in the run data.
        Checking the number of samples in each step code, especially at the start,
        is a decent indicator of if the transition to the next step is happening
        too early/late.

        The target step code can be single value or an iterable collection of
        target values. If it's a single target, then only the number of samples
        is returned. If multiple targets are given, then a dictionary of counts
        and their corresponding values are returned

        Inputs:
            fn : File path to Setitec XLS file
            sc : Single or list of floating numerical step codes to search for.
                Also accepts string all to indicate all step codes in the file.
                Default 1.

        Return a dictionary of target-position pairs if multiple are given
    '''
    data = loadSetitecXls(fn,"auto_data")
    # get unique step codes
    steps_uq = data['Step (nb)'].unique()
    if sc == 'all':
        sc = steps_uq
    elif isinstance(sc,(int,float)):
        if not (sc in steps_uq):
            return {sc : None}
        sc = [float(sc),]
    # find overlap between the step codes in the file and the ones requested
    sc = list(set(data['Step (nb)'].unique()).intersection(set(sc)))
    # if there's no overlap then return emopty dict
    if len(sc)==0:
        return {sc:None}
    res = {}
    for u in sc:
        df = data[data['Step (nb)']==u]
        res[u] = None if df.shape[0] == 0 else df['Position (mm)'].values.flatten()[0]
    return res

def getStepCodeFinalPos(fn,sc='all'):
    '''
        Get the final position value where the target step code (sc) occurs.

        Step Codes are used to denote stages of the program used in the run data.
        Checking the number of samples in each step code, especially at the start,
        is a decent indicator of if the transition to the next step is happening
        too early/late.

        The target step code can be single value or an iterable collection of
        target values. If it's a single target, then only the number of samples
        is returned. If multiple targets are given, then a dictionary of counts
        and their corresponding values are returned

        Inputs:
            fn : File path to Setitec XLS file
            sc : Single or list of floating numerical step codes to search for.
                Also accepts string all to indicate all step codes in the file.
                Default 1.

        Return a dictionary of target-position pairs if multiple are given
    '''
    data = loadSetitecXls(fn,"auto_data")
    # get unique step codes
    steps_uq = data['Step (nb)'].unique()
    if sc == 'all':
        sc = steps_uq
    elif isinstance(sc,(int,float)):
        if not (sc in steps_uq):
            return {sc : None}
        sc = [float(sc),]
    # find overlap between the step codes in the file and the ones requested
    sc = list(set(data['Step (nb)'].unique()).intersection(set(sc)))
    # if there's no overlap then return emopty dict
    if len(sc)==0:
        return {sc:None}
    # iterate over each target step code
    res = {}
    for u in sc:
        df = data[data['Step (nb)']==u]
        res[u] = None if df.shape[0] == 0 else df['Position (mm)'].values.flatten()[-1]
    return res

def getStepCodeFinalSignal(fn,sc='all'):
    '''
        Get the final torque + empty value where the target step code (sc) occurs.

        Step Codes are used to denote stages of the program used in the run data.
        Checking the number of samples in each step code, especially at the start,
        is a decent indicator of if the transition to the next step is happening
        too early/late.

        The target step code can be single value or an iterable collection of
        target values. If it's a single target, then only the number of samples
        is returned. If multiple targets are given, then a dictionary of counts
        and their corresponding values are returned

        Inputs:
            fn : File path to Setitec XLS file
            sc : Single or list of floating numerical step codes to search for.
                Also accepts string all to indicate all step codes in the file
                Default 1.

        Return a dictionary of final torque and thrust values inc empty if available
    '''
    data = loadSetitecXls(fn,"auto_data")
    # get unique step codes
    steps_uq = data['Step (nb)'].unique()
    if sc == 'all':
        sc = steps_uq
    elif isinstance(sc,(int,float)):
        if not (sc in steps_uq):
            return {sc : None}
        sc = [float(sc),]
    sc = list(set(data['Step (nb)'].unique()).intersection(set(sc)))
    if len(sc)==0:
        return {sc : None}
    res = {}
    for u in sc:
        df = data[data['Step (nb)']==u]
        if df.shape[0] == 0:
            res[u] = (None,None)
        else:
            tq = df['I Torque (A)'].values.flatten()[-1]
            if 'I Torque (Empty A)':
                tq += df['I Torque (Empty A)'].values.flatten()[-1]
            tt = df['I Thrust (A)'].values.flatten()[-1]
            if 'I Thrust (Empty A)':
                tt +=df['I Thrust (Empty A)'].values.flatten()[-1]
            res[u] = (tq,tt)
    return res

def getStepCodeFirstSignal(fn,sc='all'):
    '''
        Get the first torque + empty value where the target step code (sc) occurs.

        Step Codes are used to denote stages of the program used in the run data.
        Checking the number of samples in each step code, especially at the start,
        is a decent indicator of if the transition to the next step is happening
        too early/late.

        The target step code can be single value or an iterable collection of
        target values. If it's a single target, then only the number of samples
        is returned. If multiple targets are given, then a dictionary of counts
        and their corresponding values are returned

        Inputs:
            fn : File path to Setitec XLS file
            sc : Single or list of floating numerical step codes to search for.
                Also accepts string all to indicate all step codes in the file
                Default 1.

        Return a dictionary of final torque and thrust values inc empty if available
    '''
    data = loadSetitecXls(fn,"auto_data")
    # get unique step codes
    steps_uq = data['Step (nb)'].unique()
    if sc == 'all':
        sc = steps_uq
    elif isinstance(sc,(int,float)):
        if not (sc in steps_uq):
            return {sc : None}
        sc = [float(sc),]
    sc = list(set(data['Step (nb)'].unique()).intersection(set(sc)))
    if len(sc)==0:
        return {sc : None}
    res = {}
    for u in sc:
        df = data[data['Step (nb)']==u]
        if df.shape[0] == 0:
            res[u] = (None,None)
        else:
            tq = df['I Torque (A)'].values.flatten()[-1]
            if 'I Torque (Empty A)':
                tq += df['I Torque (Empty A)'].values.flatten()[0]
            tt = df['I Thrust (A)'].values.flatten()[-1]
            if 'I Thrust (Empty A)':
                tt +=df['I Thrust (Empty A)'].values.flatten()[0]
            res[u] = (tq,tt)
    return res
    
def getBatchStepCodeSampleCount(path,sc=1,use_mp=True,flatten=True):
    '''
        Wrapper around getStepCodeSampleCouunt to apply to a batch of files

        The use_mp argument is to use multiprocessing.Pool to speed up processing large amount of files.
        If use_mp is True, then 3 processes are used else it is treated as the integer number of processes
        to use.

        By default, the results dictionary is organised by filename and their respective counts. If flatten
        is True, then the dictionary is reorganised into unique step codes and a list of the counts for each file.
        The filenames are added under the fn key to provide a means to track down anomalous files.

        Inputs:
            path : Wildcard path to set of files or pre-sorted list of file paths
            sc : Single or iterable list of target step codes to search for. Default 1.
            use_mp : Flag than when True uses 3 processes to parallelize the collection or an integer number of processes to use.
                    Default True.
            flatten : Flag to re-arrange the results dictionary into unique step codes and list of counts with filenames under fn.

        Returns dictionary of results
    '''
    if isinstance(path,str):
        search = glob(path)
    else:
        search = path
    if use_mp:
        import multiprocessing as mp
        res = mp.Pool(3 if (mp is True) else int(use_mp)).starmap(getStepCodeSampleCount,[(fn,sc) for fn in search])
        res_dict = {fn : rr for fn,rr in zip(search,res)}
    else:
        res_dict = {fn : getStepCodeSampleCount(fn,sc) for fn in search}
    if not flatten:
        return res_dict
    # get unique step codes from all files
    res_flat = {'fn': search}
    for cts in res_dict.values():
        for uq,count in cts.items():
            if not (uq in res_flat):
                res_flat[uq] = []
            res_flat[uq].append(count)
    return res_flat

def getBatchStepCodeStartPos(path,sc=1,use_mp=True,flatten=True):
    '''
        Wrapper around getStepCodeStartPos to apply to a batch of files

        The use_mp argument is to use multiprocessing.Pool to speed up processing large amount of files.
        If use_mp is True, then 3 processes are used else it is treated as the integer number of processes
        to use.

        By default, the results dictionary is organised by filename and their respective positions. If flatten
        is True, then the dictionary is reorganised into unique step codes and a list of the positions for each file.
        The filenames are added under the fn key to provide a means to track down anomalous files.

        Inputs:
            path : Wildcard path to set of files or pre-sorted list of file paths
            sc : Single or iterable list of target step codes to search for. Default 1.
            use_mp : Flag than when True uses 3 processes to parallelize the collection or an integer number of processes to use.
                    Default True.
            flatten : Flag to re-arrange the results dictionary into unique step codes and list of positions with filenames under fn.

        Returns dictionary of results
    '''
    if isinstance(path,str):
        search = glob(path)
    else:
        search = path
    if use_mp:
        import multiprocessing as mp
        res = mp.Pool(3 if (mp is True) else int(use_mp)).starmap(getStepCodeStartPos,[(fn,sc) for fn in search])
        res_dict = {fn : rr for fn,rr in zip(search,res)}
    else:
        res_dict = {fn : getStepCodeStartPos(fn,sc) for fn in search}
    if not flatten:
        return res_dict
    # get unique step codes from all files
    res_flat = {'fn': search}
    for cts in res_dict.values():
        for uq,count in cts.items():
            if not (uq in res_flat):
                res_flat[uq] = []
            res_flat[uq].append(count)
    return res_flat

def getProgramTriggers(fn):
    '''
        Scan Setitec XLS file for the triggers that cause the program to move onto the next step

        The program steps in the Setitec run program are set to trigger when a particular value
        passes a threshold. There are a number of thresholds that can be used e.g. torque,
        thrust, DEP etc.

        The program metadata is collected from the file into a dictionary. It is then
        filtered to where Step On/Off is 1 as this means those steps are active. The
        filtered dictionary is then scanned for any non-zero trigger values.

        Any columns containing Torque, Thrust, Gap, Peck, Delay and Stroke are considered
        triggers.

        A dictionary is returned containing the step code and a list of triggers and their values

        Input:
            fn : File path to Setitec XLS

        Returns dictionary of step codes and corresponding list of tuples of trigger column names and non-zero
        trigger values
    '''
    with open(fn,'r') as open_file:
        # vector of indicies where Step On/Off is 1
        si = []
        # counter for current row
        fi = 0
        for line in open_file:
            # if the header does not contain Feed or AV get next line
            if (not ('Feed' in line)) and (not ('AV' in line)):
                continue
            head = line.strip('\n').split('\t')
            # initialize dictionary
            av = {hh:[] for hh in head}
            # read and process first line of values
            line = open_file.readline().strip('\n').split('\t')
            # if it's not empty and each element has something
            while line and all([bool(v) for v in line]):
                # iterate over keys and values converted to floating point values
                for kk,vv in zip(head,[_formatValue(v) for v in line]):
                    # update vector in dict
                    av[kk].append(vv)
                    # keep track of which rows are active in the program
                    if (kk == 'Step On/Off') and (vv==1):
                        si.append(fi)
                fi += 1
                # get next line
                line = open_file.readline().strip('\n').split('\t')
            break
        else:
            raise ValueError(f"Reached EOF in {fn}! Couldn't find Step (Nb)")
    # filter to where step on/off is 1
    av = {kk:[vv[i] for i in si] for kk,vv in av.items()}
    # create dictionary to hold triggers
    triggers = {v : [] for v in av['Step Nb']}
    # get all keys containing one of the phrases that indicates it's a trigger
    for col in filter(lambda x : any([k in x for k in ['Torque','Thrust','Gap','Peck','Delay','Stroke','DEP']]),av.keys()):
        # skip trigger if it's not in the file
        if not col in av:
            continue
        # update list of triggers for step code if value in row is non zero
        for code,row in zip(triggers.keys(),av[col]):
            if row>0:
                triggers[code].append((col,row))
    return triggers

def getAV(fn):
    '''
        Retrieve the Feed Rate values from the given object

        If fn is a string then it is treated as a file path and the
        Program metadata is extracted by reading the file.

        If fn is something else, it is assumed to be a list of objects
        as if the results of loadSetitecXls with version set to 'auto'
        were passed. The collection is filtered to the dictionary
        containing Feed or AV in any of the keys. If none can be found,
        then None is returned.

        As the Feed rate is stored in both mm/s and mm/t, the purpose of use_key
        is to provide a way to specify which one you want. The key has to
        be exact.

        e.g. Extracting from a file
        av = getAV(fn)

        e.g.2 Extracting from already loaded results
        data = loadSetitecXls(fn,'auto')
        av = getAV(fn)

        Inputs
            fn : File path string, dictionary of list of objects containing the target dictionary. 
            use_key: Key or list of keys used to reference specific values in the dictionary. Default Feed (mm/s).

        Returns dict, list of floating point feed rate values or None.
    '''
    if isinstance(fn,str):
        with open(fn,'r') as open_file:
            # scan until we've found header
            line = open_file.readline()
            for line in open_file:
                if ('AV' in line):
                    break
            else:
                return None
            # break into parts
            head = line.strip('\n').split('\t')
            # initialize dictionary of values
            av = {hh:[] for hh in head if ('Feed' in hh) or ('AV' in hh) or ('Step' in hh)}
            # get target keys
            targets = {i:k for i,k in enumerate(av.keys())}
            # iterate over lines
            for line in open_file:
                # process first line of values
                line = line.strip('\n').split('\t')
                # if empty or each element is empty then end of section
                if (not line) or (all([not bool(v) for v in line])):
                    break
                # iterate over keys and values converted to floating point values
                for i,vv in enumerate(line):
                    if i in targets:
                        # update vector in dict
                        av[targets[i]].append(_formatValue(vv))
        # filter to where Step On/Off is 1
        return {k : [v for i,v in enumerate(vals) if av[f'Step On/Off'][i]!=0] for k,vals in av.items() if k in targets and (k != 'Step On/Off')}
    # if the user has given a list of objects
    else:
        # filter list to dictionary that contains Step On/Off
        filt = list(filter(lambda x : 'Step On/Off' in x,fn))
        # if there's no match
        if len(filt)==0:
            return None
        # extract first element
        av = filt[0]
        # dictionary of vals
        vals = {}
        for k,data in av.items():
            if ('feed' in k.lower()) or ('av' in k.lower()) or (k == 'Step Nb'):
                vals[k] = [v for i,v in enumerate(data) if (av['Step On/Off'][i]!=0)]
        return vals

def getDEP(fn):
    '''
        Get the values of the DEP (mm) column in the Program Data

        The DEP column is the distance a program step lasts for. The values
        are filtered to where Step On/Off is 1 as that means those steps are
        being used.

        The results of loadSetitecXls with version set to auto can also be passed
        to save double loading the contents of the file

        Input:
            fn : Input file path to Setitec file or list of objects loaded from loadSetitecXls.

        Returns list of step codes and corresponding DEP values
    '''
    if isinstance(fn,str):
        with open(fn,'r') as open_file:
            # vector of indicies where Step On/Off is 1
            si = []
            # counter for current row
            fi = 0
            for line in open_file:
                line =  line.strip('\n').split('\t')
                # if the header does contain DEP (mm)
                if any([k=='DEP (mm)' for k in line]):
                    break
            else:
                return None
            # initialize list of DEP values
            deps = []
            sk = None
            dk = None
            # get which column is Step On/Off
            for ii in range(len(line)):
                if line[ii] == 'Step On/Off':
                    sk = ii
                if line[ii] == 'DEP (mm)':
                    dk = ii
                if (sk is not None) and (dk is not None):
                    break
            # if it's not empty and each element has something
            for line in open_file:
                line = line.strip('\n').split('\t')
                if all([not bool(v) for v in line]):
                    break
                # get value of row in step on/off
                # if it's zero then that row isn't being used
                if _formatValue(line[sk])==0:
                    continue
                # append value to list
                deps.append((_formatValue(line[0]),_formatValue(line[dk])))
                # get next line
                line = open_file.readline().strip('\n').split('\t')
            else:
                return None
        return deps
    # if the user has given a list of objects
    else:
        # filter list to dictionary that contains at least one key with Feed in it's name
        filt = list(filter(lambda x : 'DEP (mm)' in x,fn))
        # if the length is 0 return None
        if len(filt)==0:
            return None
        # extract first element
        av = filt[0]
        # filter to values where step on/off is 1
        return [(av['Step Nb'][i],av['DEP (mm)'][i]) for i in range(len(av['DEP (mm)'])) if av['Step On/Off'][i]!=0]

def getProgramValues(fn,key):
    '''
        Retrieve values for target keys from program data in Setitec file

        The input fn is either a path to a target Setitec file or a list of
        objects loaded from using loadSetitecXls.

        The input key is the search phrase used to indentify columns. For e.g.
        if key is thrust, the the values for all the columns containin thrust
        (case insensitive) are returned.

        The values returned are only those Step On/Off value is 1 meaning it was actually used.

        e.g. get all Thrust triggers
        # returns the values for Thrust Max (A), Thrust Min (A), Thrust Safety (A), Thrust Limit (A)
        # as they all contain thrust
        vals = getProgramValues(fn,'thrust')

        Inputs:
            fn : Filepath to target Setitec XLS file or a list of objects loaded from loadSetitecXls
            key : Search phrase used to identify columns. If all or None, then all values are returned.

        Returns dictionary of matching keys and active values
    '''
    # convert key to lowercase for easier searching
    key = key.lower()
    # if target is a filename
    if isinstance(fn,str):
        # open file
        with open(fn,'r') as open_file:
            # search file for header
            for line in open_file:
                line =  line.strip('\n').split('\t')
                # if the header contains Step On/Off then break
                if any([k=='Step On/Off' for k in line]):
                    break
            # if EOF return None
            else:
                return None
            # filter line to column headers that contain target phrase
            if (key=='all') or (key is None):
                vals = {k : [] for k in line}
            else:
                vals = {k : [] for k in line if key in k.lower()}
            # get target columns
            targets = list(vals.keys())
            # column number for Step On/Off
            sk = None
            # list of locations for each target key
            dk = {}
            # iterate over each column
            for ii,ll in enumerate(line):
                # get which column is Step On/Off
                if ll == 'Step On/Off':
                    sk = ii
                if ll in targets:
                    dk[ii] = ll
            # read and process first line of values
            line = open_file.readline().strip('\n').split('\t')
            # if it's not empty and each element has something
            for line in open_file:
                line = line.strip('\n').split('\t')
                if not (line and all([bool(v) for v in line])):
                    break
                # if step on/off is zero then that row isn't being used
                if _formatValue(line[sk])==0:
                    continue
                # iterate over each value in the line
                for ii,ll in enumerate(line):
                    # if the column is a target store value
                    if ii in dk:
                        vals[dk[ii]].append(_formatValue(ll))
            # if EOF return None
            else:
                return None
        return vals
    # if the user has given a list of objects
    else:
        # filter list to dictionary that contains Step On/Off
        filt = list(filter(lambda x : 'Step On/Off' in x,fn))
        # if there's no match
        if len(filt)==0:
            return None
        # extract first element
        av = filt[0]
        # dictionary of vals
        vals = {}
        # filter to values where Step On/Off is 1
        for k,data in av.items():
            if key in k.lower():
                vals[k] = [v for i,v in enumerate(data) if (av['Step On/Off'][i]!=0)]
        return vals

def loadSetitecXls(open_file,version="auto"):
    '''
        Parse the Setitec data into the different parts and then return them

        if version == manchester or stnazaire
            - General parameters -> dict
            - Cycle Paramters -> dict
            - Program parameters -> Pandas DataFrame
            - Data parameters -> dict
            - Run Data -> Pandas DataFrame

        if version == counterspark
            - General Parameters -> dict
            - Control Box -> dict
            - Motor Datas -> dict
            - Head Datas -> dict
            - Pset -> dict
            - Cycle Params -> dict
            - Program -> dict
            - Resuts -> dict
            - Data -> Pandas DataFrame

        if version is auto, then it returns a list of dictionaries representing the metadata with the last element
        being a Pandas DataFrame containing the recorded data. Results can be affectd by file firmware version

        If version is auto_data, then it searches for Position (mm) and reads in the remaining rows as data. This
        is meant to be version-independent and used on new files. If the header is missing then the columns are guessed
        based on the number of columns and their contents.

        If version is auto_params, then it collects the metadata into a list dictionaries as normal but stops when it
        reaches the row containing Position (mm) and returns the list of dictionaries
        
        Inputs:
            open_file : Path to target Setitec File
            version : String representing which method used to parsing.

        Returns the separate parts of the dataset as a list of dictionaries with a Pandas dataset at the end
    '''
    # check version
    if not (version in ["manchester","counterspark","stnazaire","auto","auto_data","auto_params"]):
        raise ValueError(f"Unsupported version {version} specified!")
    # open file
    with open(open_file,'r',encoding='latin-1') as open_file:
        if version == "manchester":
            # *** General Parameters ***
            l = open_file.readline()
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            general_parameters = dict(zip(l_head, l_tail))
            l = open_file.readline()
            # *** Cycle Parameters ***
            l = open_file.readline()
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            d1 = dict(zip(l_head, l_tail))
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            d2 = dict(zip(l_head, l_tail))
            cycle_parameters = {**d1, **d2}

            l = open_file.readline()
            l = open_file.readline()

            # *** Program ***
            ll = []
            column_names = None
            while (True):
                l = open_file.readline().strip("\n").split("\t")
                if (l[-1] == ''): break
                if column_names is None:
                    column_names = l
                    continue
                # decimals are comma format e.g. 3,000 -> 3.000 or 6,000,000 -> 6000.000
                # replace last comma with dots
                # remove all others
                ll.append([_replaceLast(tt,',','.').replace(',','') for tt in l])
            program = pd.DataFrame(ll, columns=column_names)

            l = open_file.readline()
            
            # *** Data ***
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            data1 = dict(zip(l_head, l_tail))

            ll = []
            while (True):
                l = open_file.readline().strip("\n").split("\t")
                if (l[-1] == ''): break
                ll.append(l)
            column_names = ll.pop(0)
            data = pd.DataFrame(ll, columns=column_names)
            
            for column in data.columns:
                data[column] = data[column].str.replace(',','.').astype(float)
            open_file.close()
            return general_parameters, cycle_parameters, program, data1, data
        # if reading counerspark data
        elif version == "counterspark":
            # *** General Parameters ***
            l = open_file.readline()
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            general_parameters = dict(zip(l_head, l_tail))
            l = open_file.readline()
            
            # *** Control Box Datas ***
            # skip star header
            open_file.readline()
            # read line header
            l_head = open_file.readline().strip("\n").split("\t")
            # there are a bunch of empty lines
            # skip them
            for l in open_file:
                if l.strip('\n'):
                    break
            # for some reason the keys are actually split
            l_head.extend(l.strip('\n').split('\t'))
            # all the values are in the next line
            l_tail = open_file.readline().strip('\n').split('\t')
            control_box_datas = dict(zip(l_head,l_tail))
            # skip the next empty line
            open_file.readline()
            
            # *** Motor Datas ***
            # skip section head
            open_file.readline()
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            motor_datas = dict(zip(l_head,l_tail))
            # skip empty line
            open_file.readline()

            # *** Head Datas ***
            # skip section head
            open_file.readline()
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            head_datas = dict(zip(l_head,l_tail))
            # skip empty line
            open_file.readline()
            
            # *** Pset ***
            # skip section head
            open_file.readline()
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            Pset = dict(zip(l_head,l_tail))
            # skip empty line
            open_file.readline()

            # *** Cycle parameters ***
            # skip section head
            open_file.readline()
            # read first header
            l_head = open_file.readline().strip("\n").split("\t")
            # read first tail
            l_tail = [float(vv.replace(',','.')) for vv in open_file.readline().strip("\n").split("\t")]
            l_head.extend(open_file.readline().strip("\n").split("\t"))
            l_tail.extend([float(vv.replace(',','.')) for vv in open_file.readline().strip("\n").split("\t")])
            cycle_params = dict(zip(l_head,l_tail))
            # skip empty line
            open_file.readline()
            
            # *** Program ***
            open_file.readline()
            ll = []
            column_names = None
            while (True):
                l = open_file.readline().strip("\n").split("\t")
                if (l[-1] == ''): break
                if column_names is None:
                    column_names = l
                    continue
                # decimals are comma format e.g. 3,000 -> 3.000 or 6,000,000 -> 6000.000
                # replace last comma with dots
                # remove all others
                ll.append([_replaceLast(tt,',','.').replace(',','') for tt in l])
            program = pd.DataFrame(ll, columns=column_names)
            
            # *** Results ***
            open_file.readline()
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = [tt.replace(',','.') for tt in open_file.readline().strip("\n").split("\t")]
            results = dict(zip(l_head,l_tail))
            open_file.readline()

            # *** Datas ***
            open_file.readline()
            ll = []
            while (True):
                l = open_file.readline().strip("\n").split("\t")
                if (l[-1] == ''): break
                ll.append(l)
            column_names = ll.pop(0)
            data = pd.DataFrame(ll, columns=column_names)
            
            for column in data.columns:
                data[column] = data[column].str.replace(',','.').astype(float)

            return general_parameters,control_box_datas,motor_datas,head_datas,Pset,cycle_params,program,results,data
        elif version == "stnazaire":
            # *** General Parameters ***
            l = open_file.readline()
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            general_parameters = dict(zip(l_head, l_tail))
            l = open_file.readline()
            # *** Cycle Parameters ***
            l = open_file.readline()
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            d1 = dict(zip(l_head, l_tail))
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            d2 = dict(zip(l_head, l_tail))
            cycle_parameters = {**d1, **d2}

            l = open_file.readline()
            l = open_file.readline()

            # *** Program ***
            ll = []
            column_names = None
            while (True):
                l = open_file.readline().strip("\n").split("\t")
                if (l[-1] == ''): break
                if column_names is None:
                    column_names = l
                    continue
                # decimals are comma format e.g. 3,000 -> 3.000 or 6,000,000 -> 6000.000
                # replace last comma with dots
                # remove all others
                ll.append([_replaceLast(tt, ',', '.').replace(',', '') for tt in l])
            program = pd.DataFrame(ll, columns=column_names)

            l = open_file.readline()

            # *** Data ***
            #open_file.readline()
            #open_file.readline()
            l_head = open_file.readline().strip("\n").split("\t")
            l_tail = open_file.readline().strip("\n").split("\t")
            data1 = dict(zip(l_head, l_tail))

            ll = []
            while (True):
                l = open_file.readline().strip("\n").split("\t")
                if (l[-1] == ''): break
                ll.append(l)
            column_names = ll.pop(0)
            data = pd.DataFrame(ll, columns=column_names)

            for column in data.columns:
                data[column] = data[column].str.replace(',', '.').astype(float)
            open_file.close()
            return general_parameters, cycle_parameters, program, data1, data
        # attempt to autoparse the file
        elif version == "auto":
            # create the list of dictionaries to return
            section_list = []
            # iterate over lines in the file
            for ll in open_file:
                # remove trailing characters and break it into parts
                ll = ll.strip('\n').replace('\t','')
                # if the line is blank
                if not ll:
                    continue
                # if the line has stars in the line then it's the start of a section
                if "*" in ll:
                    # next line has the column headers
                    l_head = open_file.readline().strip("\n").split("\t")
                    # remove any blank parts
                    l_head = [kk for kk in l_head if kk]
                    rr = 0
                    # construct dictionary of values for section
                    section = {kk : [] for kk in l_head}
                    # iterate over the next lines these are taken as values for the column headers
                    for ln in open_file:
                        # split into parts
                        ln = ln.strip('\n').split('\t')
                        # remove any blank ones
                        l_tail = [vv for vv in ln if vv]
                        # if line is blank then it's the dead space between sections
                        if len(l_tail)==0:
                            break
                        # if all parts do NOT contain a comma and are not numbers if comma is removed
                        # then a new set of columns has been found
                        if (not all([',' in t for t in l_tail])) and all([all([not cc.isnumeric() for cc in t.replace(',','')] for t in l_tail)]) and rr>0:
                            # hidden row of columns
                            l_head = l_tail
                            # add section to list
                            section_list.append(section)
                            # add section to list
                            section = {kk : [] for kk in l_head}
                            rr=0
                            continue
                        # process the values into floats replacing last comma with dot
                        #l_tail = [_replaceLast(vv,',','.').replace(',','') for vv in l_tail]
                        l_tail = [_formatValue(vv,',','.') for vv in l_tail]
                        # update section values
                        for kk,val in zip(l_head,l_tail):
                            section[kk].append(val)
                        rr+=1
                    section_list.append(section)
            # convert last data set to a pandas array
            section_list[-1] = pd.DataFrame.from_dict(section_list[-1],dtype='float32')
            return section_list
        # only return run data at the end
        elif version == "auto_data":
            # read lines
            for line in open_file:
                l_head = line.strip("\n").split("\t")
                # if the row contains Position (mm) then we're at the data section
                if not (("Position (mm)" in l_head[0])):
                    continue
                # initialize dictionary
                data = {kk : [] for kk in l_head}
                # iterate over remaining lines
                for ll in open_file:
                    # process the lines
                    ll = ll.strip("\n").split("\t")
                    # if the line contains nothing
                    if len(ll)==0:
                        break
                    # process the values into valid floats
                    ll = [_replaceLast(vv,',','.').replace(',','') for vv in ll]
                    # update dictionary
                    for hh,tt in zip(l_head,ll):
                        data[hh].append(tt)
                # WH 2024.05.17 remove empty keys from dictionary
                data = {kk : vv for kk,vv in data.items() if vv}
                return pd.DataFrame.from_dict(data,dtype='float32')
            # if EOF and haven't found a header
            # then attempt to extract data based on knowledge of column headers
            else:
                # return to start
                open_file.seek(0)
                # read in document
                data = [[_formatValue(v) for v in line.strip("\n").split("\t")] for line in open_file]
                data = np.row_stack(data)
                # count number of columns
                nc = len(data[1])
                cols = ['Position (mm)','I Torque (A)','I Thrust (A)','I Torque Empty (A)']
                if nc == 8:
                    cols += ['Step (nb)','Stop code','Mem Torque min (A)','Mem Thrust min (A)']
                elif nc == 11:
                    # check 4th column is all integers
                    # if so then take that column to be the step code
                    if all([v.is_integer() for v in data[:,4]]) and np.unique(data[:,4]).shape[0]>0:
                        cols += ['Step (nb)','Stop code','Mem Torque min (A)','Mem Thrust min (A)','Rotation Speed (rpm)','Feed Speed (mm/s)','Drive Temperature (deg C)']
                    else:
                        cols += ['I Thrust Empty (A)','Step (nb)','Stop code','Mem Torque min (A)','Mem Thrust min (A)','Torque Power (W)','Gap Length (mm)']
                # first three columns are all the same
                # rest depend on firmware version so cols are set to Unknown
                else:
                    cols += (nc-len(cols))*['Unknown',]
                # form into dataframe using proxy columns
                return pd.DataFrame(data,columns=cols)
        # only return parameters
        elif version == "auto_params":
            # create the list of dictionaries to return
            section_list = []
            # iterate over lines in the file
            for ll in open_file:
                # remove trailing characters and break it into parts
                ll = ll.strip('\n').replace('\t','')
                # if the line is blank
                if not ll:
                    continue
                # if the line has stars in the line then it's the start of a section
                if "*" in ll:
                    # next line has the column headers
                    l_head = open_file.readline().strip("\n").split("\t")
                    # remove any blank parts
                    l_head = [kk for kk in l_head if kk]
                    # if we've reached a line with Position we want to stop scanning the file
                    if 'Position (mm)' in l_head:
                        return section_list
                    rr = False
                    # construct dictionary of values for section
                    section = {kk : [] for kk in l_head}
                    # iterate over the next lines these are taken as values for the column headers
                    for ln in open_file:
                        # split into parts
                        ln = ln.strip('\n').split('\t')
                        # remove any blank ones
                        l_tail = [vv for vv in ln if vv]
                        # if line is blank then it's the dead space between sections
                        if len(l_tail)==0:
                            break
                        # if ALL parts of the line are letters or punctuation, then a new set of columns
                        # has been found
                        #if all([all([cc.isalpha() or not cc.isalnum() for cc in t] for t in l_tail)]) and rr>1:
                        # if all parts do NOT contain a comma and are not numbers if comma is removed
                        if (not all([',' in t for t in l_tail])) and all([all([not cc.isnumeric() for cc in t.replace(',','')] for t in l_tail)]) and rr:
                            # hidden row of columns
                            l_head = l_tail
                            # add section to list
                            section_list.append(section)
                            # add section to list
                            section = {kk : [] for kk in l_head}
                            rr=False
                            continue
                        # process the values into floats replacing last comma with dot
                        l_tail = [_replaceLast(vv,',','.').replace(',','') for vv in l_tail]
                        # update section values
                        for kk,val in zip(l_head,l_tail):
                            section[kk].append(val)
                        rr=True
                    section_list.append(section)

def convertRecArrayToDict(arr):
    '''
        Convert numpy.recarray to a dictionary

        Dictionary organised by field/column names

        Return dictionary of field names and numpy arrays
    '''
    return {ff:arr[ff] for ff in arr.dtype.names}

def saveSetitecXlsAsNPZ(fp,**kwargs):
    '''
        Convert Setitec XLS files to a compressed NPZ

        Inputs:
            fp : File path or file pointer to input file
            **kwargs:
                sep_files : Generate separate NPZ files and JSON files
                no_struct : Don't use structured arrays in NPZ files
    '''
    if isinstance(fp,TextIOWrapper):
        fpath = os.path.splitext(os.path.basename(fp.name))[0]+'.npz'
    else:
        fpath = os.path.splitext(os.path.basename(fp))[0]+'.npz'
    # process file into numpy
    gen_dict,cycle_dict,program,data_dict,data = loadSetitecXlsAsNP(fp)
    # if the user wants separate files
    # create separate files 
    if kwargs.get('sep_files',False):
        # save dictionaries as JSON files
        np.savez_compressed(f"{fpath}-general-params.json",gen_dict)
        np.savez_compressed(f"{fpath}-cycle-params.json",cycle_dict)
        np.savez_compressed(f"{fpath}-process-params.json",data_dict)
        # if the user doesn't want structured arrays
        # break into dictionary
        if kwargs.get('no_struct',False):
            np.savez_compressed(f"{fpath}-program-data.npz",**convertRecArrayToDict(program))
            np.savez_compressed(f"{fpath}-process-data.npz",**convertRecArrayToDict(data))
        else:
            np.savez_compressed(f"{fpath}-program-data.npz",program)
            np.savez_compressed(f"{fpath}-process-data.npz",data)
    else:
        # if not saving it as a record
        if kwargs.get('no_struct',False):
            # each entry of the dictionaries is saved separately
            np.savez_compressed(fpath,**gen_dict,**cycle_dict,**data_dict,
                                # convert the rec arrays into dictionaries
                                **convertRecArrayToDict(program),**convertRecArrayToDict(data))
        else:
            # save structured arrays as a NPZ
            np.savez_compressed(fpath,**gen_dict,**cycle_dict,**data_dict,program=program,data=data)

def saveSetitecXlsAsHDF5(path,cp=9):
    '''
        Save Setitec Spreadsheet as a compressed HDF5 file

        Saves the dictionaries as attributes
        Saves each column of the DataFrames as a compressed dataset

        Inputs:
            path : Path to Setitec Spreadsheet
            cp : Compression level. Default 9
    '''
    if isinstance(path,TextIOWrapper):
        fpath = os.path.splitext(os.path.basename(path.name))[0]+'.hdf5'
    else:
        fpath = os.path.splitext(os.path.basename(path))[0]+'.hdf5'
    # process file into numpy arrays
    gen_dict,cycle_dict,program,data_dict,data = loadSetitecXls(path)
    # combine dictionaries into a giant one that can be easily iterated over
    params_dict = {kk:vv for kk,vv in list(gen_dict.items())+list(cycle_dict.items())+list(data_dict.items())}
    print(params_dict)
    # create destination file
    with h5py.File(fpath,'w') as dest:
        # iterate over paramters
        # as they're single values. no point compressing
        for kk,vv in params_dict.items():
            if kk:
                dest.attrs[kk] = vv
        # iterate over dataframes
        for df in [program,data]:
            for col,cells in df.items():
                print(col,cells.shape)
                dest.create_dataset(name=col,data=cells.values,chunks=cells.shape,dtype=cells.dtype,
                                    compression="gzip",compression_opts=cp)

def loadSetitecXlsAsNP(fp):
    '''
        Convert Setitec XLS files to a set of dicts + np.recarrays

        Inputs:
            fp : File path or file pointer to input file

        Returns set of 5 numpy arrays

        - General parameters -> dict
        - Cycle Paramters -> dict
        - Program parameters -> np.recarray
        - Data parameters -> dict
        - Run Data -> np
    '''    
    # process file into panda dataframes + dicts
    gen_params,cycle_params,program,datal,data = loadSetitecXls(fp)
    ## program data
    # split into columns
    pts = np.split(program.values,program.values.shape[1],axis=1)
    # convert to a numpy rec array
    # column names become field names
    prog_data = np.core.records.fromarrays(pts,names=program.columns.values.tolist())
    ## run data
    # split into columns
    pts = np.split(data.values,data.values.shape[1],axis=1)
    # convert to a numpy rec array
    # column names become field names
    data_data = np.core.records.fromarrays(pts,names=data.columns.values.tolist())
    # return the results
    return gen_params,cycle_params,prog_data,datal,data_data

def saveMeasurementsToNpz(path):
    '''
        Convert values in a measurements XLXS workbook to compressed NPZ file

        This uses openpyxl instead of pandas as it tended to fail during testing

        Created NPZ file has the same name as source

        e.g. measurementsALL.xlsx -> measurementsALL.npz

        Inputs:
            path : Path to filename
    '''
    # get filename
    fpath = os.path.splitext(os.path.basename(path))[0]
    # load workbook as read only + data only to get result of formulae
    wb = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    # check if only one sheet was present
    one_sheet = len(wb.sheetnames)==1
    # sheet dictionary
    sheet_dict ={}
    # iterate over sheets
    for sn in wb.sheetnames:
        # get sheet data
        sheet_data = wb[sn]
        # convert to Pandas DataFrame
        sheet_data = pd.DataFrame(sheet_data.values)
        # dictionary of values organised by column name
        # sheet name is added in case multiple sheets have the same named dictionary
        for col,cells in sheet_data.items():
            if one_sheet:
                sheet_dict[col] = cells.values
            else:
                sheet_dict[f"{sn.replace(' ','_')}_{col}"] = cells.values
    # save compressed file
    np.savez_compressed(fpath+'.npz',**sheet_dict)

def cleanupMeasurementsPD(data):
    '''
        Cleans up the Measurement Pandas DataFrame loaded from loadMeasurements

        The column names are in the first row of the DataFrame. Despite renaming the columns,
        the effect doesn't stick. This function is to perform the cleanup operation
    '''

    # column names are stored as the first row
    cols = data.iloc[0].values.astype("<U7")
    # get from the 1st row onwards
    sheet_clip = data.iloc[1:]
    # update column names
    sheet_clip.columns = cols.tolist()
    return sheet_clip

def loadMeasurementsAsNP(path):
    '''
        Convert values in a measurements XLXS workbook to dictionary of Numpy arrays

        This uses openpyxl instead of pandas as it tended to fail during testing

        Dictionary keys are sheet name followed by dataset name

        e.g. Sheet1_ALU_0

        Inputs:
            path : Path to filename

        Returns dictionary of arrays
    '''
    # load workbook as read only + data only to get result of formulae
    wb = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    one_sheet = len(wb.sheetnames)==1
    # sheet dictionary
    sheet_dict ={}
    # iterate over sheets
    for sn in wb.sheetnames:
        # get sheet data
        sheet_data = wb[sn]
        # convert to Pandas DataFrame
        sheet_data = pd.DataFrame(sheet_data.values)
        # dictionary of values organised by column name
        # sheet name is added in case multiple sheets have the same named dictionary
        for col,cells in sheet_data.items():
            if one_sheet:
                sheet_dict[col] = cells.values
            else:
                sheet_dict[f"{sn.replace(' ','_')}_{col}"] = cells.values
    return sheet_dict

def loadMeasurementsAsPD(path):
    '''
        Convert values in a measurements XLXS workbook to dictionary of Pandas DataFrames

        This uses openpyxl instead of pandas as it tended to fail during testing

        Dictionary keys are sheet name

        Inputs:
            path : Path to filename

        Returns dictionary of DataFrames
    '''
    # load workbook as read only + data only to get result of formulae
    wb = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    # sheet dictionary
    sheet_dict ={}
    # iterate over sheets
    for sn in wb.sheetnames:
        # get sheet
        sheet_data = wb[sn]
        # convert values to DataFrame
        sheet_data = pd.DataFrame(sheet_data.values)
        # store in array
        sheet_dict[sn] = sheet_data
    # return dictionary
    return sheet_dict

############################################## MAT FILES ##############################################
def loadMatCellMetaData(path):
    '''
        Attempt to get the metadata from the cells data stored in MAT file

        Inputs:
            path : Path to MAT file

        Returns dict of field names
    '''
    metadata = {}
    with h5py.File(path,'r') as file:
        # get refs grop
        refs = file['#refs#']
        # itertate over it
        for nn in refs.keys():
            # if it has a MATLAB_fields attribute
            if 'MATLAB_fields' in refs[nn].attrs.keys():
                # get the fields
                fields = refs[nn].attrs['MATLAB_fields']
                # iterate over converting each entry to string
                metadata[nn] = [ff.tobytes().decode('utf-8') for ff in fields]
    return metadata

def repackMATToHDF5(path,**kwargs):
    '''
        Repack a MAT file to a new file

        Only works for MAT versions >=7.3 as they can be read in as HDF5s

        The source file is iterated over and the same structure is recreated with a couple of exceptions.
            - Datasets are set to be compressed using GZIP to the compression level set by the user
            - The MATLAB_fields attribute converted from array of byte arrays to list of strings to make it
              easier to read snd more Python friendly

        NOTE: The MATLAB_fields change is implemented as attempting to read in the file using scipy.io.loadmat
        generated an error citing that attribute as the problem

        NOTE 2: GZIP is used for compression as it's available in all h5py installations and is the most portable.

        NOTE 3: Data is chunked to improve compression

        Inputs:
            path : Path to MAT file
            kwargs:
                cp : Compression level. Default 9.
                cp_algo : Compression algorithm. Default gzip
                csize : Chunk size used. Defaults to True.
    '''
    # open source
    with h5py.File(path,'r') as source:
        # open destination
        with h5py.File(os.path.splitext(os.path.basename(path))[0] + ".hdf5",'w') as dest:
            # start iterating over the source MAT file
            def repack_dset(name,item):
                # only datasets have dtype attribute
                if isinstance(item,h5py.Dataset):
                    # create a dataset with compression opts
                    ds = dest.create_dataset(name,chunks=kwargs.get('csize',True),dtype=item.dtype,data=source[name][()],
                                        compression=kwargs.get('cp_algo','gzip'),compression_opts=kwargs.get('cp',9))
                    # update the attributes
                    ds.attrs.update(**source[name].attrs)
                # if a group
                elif isinstance(item,h5py.Group):
                    gp = dest.require_group(name)
                    # get attributes
                    attrs = {kk:vv for kk,vv in source[name].attrs.items()}
                    # get the MATLAB fields
                    if 'MATLAB_fields' in attrs:
                        # get fields
                        fields = attrs.pop('MATLAB_fields')
                        # iterate over fields converting entries to string
                        fields = [ff.tobytes().decode('utf-8') for ff in fields]
                        # update field
                        attrs['MATLAB_fields'] = fields
                    # copy attributes to dest
                    gp.attrs.update(attrs)
            source.visititems(repack_dset)

def repackMAT(path,**kwargs):
    '''
        Repack MAT file to a new file format

        Only works for MAT versions >=7.3 as they can be read in as HDF5s

        Keyword arguments control filetype and compression options

        Inputs:
            path : File path to source MAT file
            kwargs: Keywords controlling target filetype and compression opts
                as_hdf5 : Compresses to HDF5 file. Uses cp keyword to specify compression level.
                cp : Compression level. If not specified, defaults to 9 if writing to hdf5 and None if writing to Parquet.
                    Effect of value varies depending on compression algorithm.
                cp_algo : Compression algorithm. Check filetype docs for supported algorithms. Default to gzip (lowercase).
                as_npz : Repacks to compressed NPZ file. Organised by variable names and stores a dictionary where keys are tuples of hole and coupon
                as_parq : Repack to compressed Parquet file. Uses cp to control compression level. Uses cp_algo to set compression algorithm. Converts
                            string to uppercase as pyarrow uses uppercase strings. Uses pyarrow.
    '''
    # if keywords are empty, don't do anything
    if not kwargs:
        return
    # open source
    with h5py.File(path,'r') as source:
        # if repacking to a hdf5
        if kwargs.get('as_hdf5',False):
            # start iterating over the source MAT file
            # open destination
            with h5py.File(os.path.splitext(os.path.basename(path))[0] + ".hdf5",'w') as dest:
                def repack_dset(name,item):
                    # only datasets have dtype attribute
                    if isinstance(item,h5py.Dataset):
                        # create a dataset with compression opts
                        ds = dest.require_dataset(name,chunks=item.shape,dtype=item.dtype,data=source[name][()],
                                            compression=kwargs.get('cp_algo',"gzip"),compression_opts=kwargs.get('cp',9))
                        # update the attributes
                        ds.attrs.update(**source[name].attrs)
                    # if a group
                    elif isinstance(item,h5py.Group):
                        gp = dest.require_group(name)
                        # get attributes
                        attrs = {kk:vv for kk,vv in source[name].attrs.items()}
                        # get the MATLAB fields
                        if 'MATLAB_fields' in attrs:
                            # get fields
                            fields = attrs.pop('MATLAB_fields')
                            # iterate over fields converting entries to string
                            fields = [ff.tobytes().decode('utf-8') for ff in fields]
                            # update field
                            attrs['MATLAB_fields'] = fields
                        # copy attributes to dest
                        gp.attrs.update(attrs)
            source.visititems(repack_dset)
        # repack to NPZ or parquet file
        elif kwargs.get('as_npz',False) or kwargs.get('as_parq',False):
            keys = list(source.keys())
            data_key = list(filter(lambda kk : not ('#' in kk),keys))[0]
            # get cell array shape
            if len(source[data_key][()].shape) == 2:
                pages = 1
                nholes,ncoupons = source[data_key][()].shape
            else:
                pages,nholes,ncoupons = source[data_key][()].shape
            # get paths to groups where the data references are stored
            sz_sort = {}
            def find_gps(name,item):
                # search for group
                if isinstance(item,h5py.Group):
                    # skip subsystem
                    if '#subsystem#' in name:
                        return
                    # if the group contains data
                    if 'data' in item:
                        # update data dictionary
                        # key values are groups
                        sz_sort[name] = [item['rowTimes']['origin'][()].tolist(), # get origin timestamp
                                    item['rowTimes']['sampleRate'][0,0], # get sample rate
                                    max(source[item['data'][()].flatten()[0]].shape), # get max size of the first dataset. Same set size for all
                                    max(item['data'].shape), # get number of data arrays
                                    item['varNames'][()].flatten()] # variable names
            source.visititems(find_gps)
            ## separate into groups by variable type
            # create a dictionary to associate number of variables with an index
            szs = {vv:ii for ii,vv in enumerate(set([vv[3] for vv in sz_sort.values()]))}
            # create list of key groups
            page_keys = [[] for _ in range(len(szs))]
            # iterate over the metadata dictionary
            for kk,vv in sz_sort.items():
                page_keys[szs[vv[3]]].append(kk)
            # get variable order
            vorder = {ii:vn for ii,vn in enumerate(set([decodeVarname(source[ref][()]) for ref in list(sz_sort.values())[0][4]]))}
            # get downsampling limit
            ds_lim = kwargs.get('dsf_lim',)
            # if the user gave a negative number raise error
            if ds_lim<=0:
                raise ValueError(f"Downsampling Limit has to be positive, non-zero (got {ds_lim}")
            # get downsampling factor
            dsf = int(kwargs.get('dsf',1))
            # function to check indicies
            def load_index(shape):
                if isinstance(shape,(list,tuple)):
                    mx_sz = max(shape)
                else:
                    mx_sz = shape
                return np.arange(0,mx_sz,dsf if mx_sz >= ds_lim else 1)
            # construct dictionary of data
            data_dict = {vn: # column names
                             {np.unravel_index(ii,(nholes,ncoupons)): # indexed by hole and coupon number
                                  # dataset that uses get_index top handle downsampling
                                  source[source[kk]['data'][()][ik,0]][()].flatten()[load_index(source[source[kk]['data'][()][ik,0]][()].shape[1])]
                              for ii,(kk,vv) in enumerate(page_keys)} # iterate over the keys in the group
                        for ik,vn in vorder.items()}
            # if user wants a NPZ file
            if kwargs.get('as_npz',False):
                ## make a nested dictionary by variable name
                np.savez_compressed(os.path.splitext(os.path.basename(path))[0] + ".npz",**data_dict)
            # if the user wants a parquet file
            elif kwargs.get('as_parq',False):
                from pyarrow import Table, parquet as pq
                # construct dictionary to Table
                tb = Table.from_dict(data_dict)
                # write the table to a parquet file
                pq.write_table(tb,compression=kwargs.get('cp_algo',"gzip").upper(),compression_level=kwargs.get('cp',None))
                
def loadUniqueFieldNames(path,return_counts=True):
    '''
        Get the unique MATLAB field names from a MAT file

        Uses np.unique

        Inputs:
            path : Filepath to MAT file
            return_counts: Flag to return the number of each occurence

        Returns unique field names + array of counts
    '''
    fields = []
    # open source
    with h5py.File(path,'r') as source:
        def find_fieldnames(name,item):
            if isinstance(item,h5py.Group):
                if 'MATLAB_fields' in item.attrs:
                    for ff in item.attrs['MATLAB_fields']:
                        fields.append(ff.tobytes().decode('utf-8'))
        source.visititems(find_fieldnames)
    return np.unique(fields,return_counts=return_counts)

def loadUniqueDsetNames(path,return_counts=True):
    '''
        Get the unique dataset names from a HDF5 file

        Uses np.unique

        Inputs:
            path : Filepath to HDF5 file
            return_counts: Flag to return the number of each occurence

        Returns unique field names + array of counts
    '''
    fields = []
    # open source
    with h5py.File(path,'r') as source:
        def find_fieldnames(name,item):
            if isinstance(item,h5py.Dataset):
                fields.append(item.name.split('/')[-1])
        source.visititems(find_fieldnames)
    return np.unique(fields,return_counts=return_counts)

def loadMatFieldsVals(path,field,join=False):
    '''
        Get the datasets from the MAT file whose name matches target

        Inputs:
            path : Path to MAT file
            field : Dataset name to search for
            join : Flag to join found datasets together into a single array. Default True

        Returns list of found datasets if join is False or a single numpy array if join is True
    '''
    vals = []
    with h5py.File(path,'r') as source:
        def find_fieldvals(name,item):
            # if target is a dataset
            if isinstance(item,h5py.Dataset):
                # get final part
                dname = name.split('/')[-1]
                if dname == field:
                    vals.append(item[()])
        source.visititems(find_fieldvals)
    # join arrays together
    if join:
        return np.vstack(vals)
    else:
        return vals

def loadUniqueDtypes(path,return_counts=True):
    '''
        Get the unique dataset dtypes from a HDF5 file

        Uses np.unique

        Inputs:
            path : Filepath to HDF5 file
            return_counts: Flag to return the number of each occurence

        Returns unique field names + array of counts
    '''
    fields = []
    # open source
    with h5py.File(path,'r') as source:
        def find_fieldnames(name,item):
            if isinstance(item,h5py.Dataset):
                fields.append(item.dtype)
        source.visititems(find_fieldnames)
    return np.unique(fields,return_counts=return_counts)

def loadMatData(path,sort_mode="vn",**kwargs):
    '''
        Extract the data stored in a Setitec or DAQ MAT file

        Finds all the datasets stored in the file and splits them into a max of two groups according to size. The size of the arrays in each page of measurements are similar
        (e.g. 695,696) so we can safely group them. Then the groups are sorted by timestamp and stored in either a nested list or DataFrame depending on as_df flag.

        If as_df is True, then the returned array is a Pandas DataFrame. If as_df is False,
        then a nested list is returned organised in a similar manner.

        Inputs:
            path : File path to MAT file
            sort_mode : Sorting mode for grouping datasets
                Supported modes:
                    sz,size : Datasets are grouped according to similarity in size. Grouped by standard deviation
                    vn,varNum: Number of variables
            dsf : Downsample factor. Rate at which to downsample the data e.g. 2 means every 2nd value is used.
            sample_rate : Sampling rate. Used to create time index
            rowbyrow : Flag to instead build the dataframe row by row instead of using dictionary comprehension

        Returns a list of panda DataFrames where each entry is the data for a page.
    '''
    with h5py.File(path,'r') as source:
        # search for the first key that does not have a hash symbol
        keys = list(source.keys())
        data_key = list(filter(lambda kk : not ('#' in kk),keys))[0]
        # get cell array shape
        if len(source[data_key][()].shape) == 2:
            pages = 1
            nholes,ncoupons = source[data_key][()].shape
        else:
            pages,nholes,ncoupons = source[data_key][()].shape
        # get paths to groups where the data references are stored
        sz_sort = {}
        def find_gps(name,item):
            # search for group
            if isinstance(item,h5py.Group):
                # skip subsystem
                if '#subsystem#' in name:
                    return
                # if the group contains data
                if 'data' in item:
                    # update data dictionary
                    # key values are groups
                    sz_sort[name] = [item['rowTimes']['origin'][()].tolist(), # get origin timestamp
                                item['rowTimes']['sampleRate'][0,0], # get sample rate
                                max(source[item['data'][()].flatten()[0]].shape), # get max size of the first dataset. Same set size for all
                                max(item['data'].shape), # get number of data arrays
                                item['varNames'][()].flatten()] # variable names
        source.visititems(find_gps)
        # group the entries by dataset size        
        if sort_mode in ["sz","size"]:
            # get keys
            sz_keys = list(sz_sort.keys())
            # create list of gaps between sizes
            # e.g. sz_gaps[0] -> sz[1] - sz[0]
            sz_gaps = [y - x for x, y in zip([sz_sort[kk][2] for kk in sz_keys[:-1]], [sz_sort[kk][2] for kk in sz_keys[1:]])]
            # calculate standard deviation of the gaps
            # single value
            sd = stdev(sz_gaps)
            # init list of keys for each page of values
            # groups are organised by number of standard deviations
            # i.e. group[0] -> within 1 std, group[1] -> more than 1 std
            # as we only need 2 groups this should be fine as the groups of sizes are only 1 difference from each other
            page_keys = [[sz_keys[0]]]
            # iterate over the keys of the sorted list
            # starting from element 1
            for kk in sz_keys[1:]:
                # get the size
                x = sz_sort[kk][2]
                # if the gap from the current item to the previous is more than 1 SD
                if (x - sz_sort[page_keys[-1][-1]][2]) / sd > 1:
                    # then start a new group
                    page_keys.append([])
                # add the current item to the last list in the list
                page_keys[-1].append(kk)
        # group the entries by number of variables 
        elif sort_mode in ["vn","varNum"]:
            # create a dictionary to associate number of variables with an index
            szs = {vv:ii for ii,vv in enumerate(set([vv[3] for vv in sz_sort.values()]))}
            # create list of key groups
            page_keys = [[] for _ in range(len(szs))]
            # iterate over the metadata dictionary
            for kk,vv in sz_sort.items():
                page_keys[szs[vv[3]]].append(kk)
        else:
            raise ValueError(f"Unsupported sorting algorithm {sort_mode}")
        # list of keys for each page
        pages = []
        # iterate over the groups
        for page in page_keys:
            # get the values for this group
            tbs_page = {kk:sz_sort[kk] for kk in page}
            # sort by origin timestamp
            tbs_page = {kk:vv for kk,vv in sorted(tbs_page.items(),key=lambda item: item[1][0])}
            # same variable name order for all so we just need var names from one entry
            vorder = {ii:vn for ii,vn in enumerate(set([decodeVarname(source[ref][()]) for ref in list(tbs_page.values())[0][4]]))}
            # get sample rate for page
            # should be single value
            #sample_rate = set([vv[1] for vv in tbs_page.values()]).pop()
            # find max measurement size
            max_sz = max(set([source[gg]['numRows'][0,0] for gg in tbs_page.keys()]))
            ## if user wants it as a pandas dataframe
            # creates indicies
            holes = np.arange(nholes)
            coupons = np.arange(ncoupons)
            # create full index list
            index = np.arange(0,max_sz,dtype="uint32")
            # if the max size is greater than the current size then downsample
            # get downsampling factor
            dsf = int(kwargs.get('dsf',1))
            if dsf <= 0:
                raise ValueError(f"Downsampling factor has to be positive (got {dsf})!")
            # downsample index
            if dsf >1:
                index = index[::dsf]
            # get variable names
            varNames = np.array(list(vorder.values()))
            # hideous dictionary comprehension to create pandas dataframe from the data
            hole_coup = pd.DataFrame({vn: # column names
                             {np.unravel_index(ii,(nholes,ncoupons)): # indexed by hole and coupon number
                                  # dataset that uses get_index top handle downsampling
                                  source[source[kk]['data'][()][ik,0]][0,index] # data type
                              for ii,(kk,vv) in enumerate(tbs_page.items())} # iterate over the keys in the group
                        for ik,vn in vorder.items()}, # iterate over variable names order
                index=pd.MultiIndex.from_product([holes, coupons, index],names=["Hole","Coupon","Index"]), # create multi index
                columns=varNames,dtype=np.float64) # set datatype to float64
            # if a sample rate is specified
            # create a time vector
            if 'sample_rate' in kwargs:
                sr = kwargs['sample_rate']
                if sr==0:
                    warnings.warn("Sample Rate cannot be 0! Skipping building time vector")
                else:
                    time = (1.0/sr)*np.arange(0,max_sz,dtype='float64')
                    time = time[index]
                    time = np.tile(time,nholes*ncoupons)
                    hole_coup['Time'] = time
            # add dataframe to list
            pages.append(hole_coup)
        # return list of dataframes
        return pages

def loadMatChangePoints(path,data_key=None,as_df=False):
    '''
        Extract the data from a changepoints array

        NOTE: This is different from getMatData as it doesn't need to unpick a cell_array, only a cell.

        If as_df is True, the returned DataFrame is organised by MultiIndex sorted by Hole and Coupon and the columns are
        Time and CP.

        Inputs:
            path : File path to target MAT file
            data_key : Path to location where cell is located. If None, it selects first key that does not have a # in it
            as_df : Flag to return the data as a Pandas DataFrame. If False, a 4D numpy array is returned instead. Row and column entry
                    is the change points and timestamp.

        Returns either a pandas array or a numpy array
    '''
    # open h5py file
    with h5py.File(path,'r') as source:
        #f data key is not specified
        if data_key is None:
            keys = list(source.keys())
            data_key = list(filter(lambda kk : not ('#' in kk),keys))[0]
        # access change locations
        data = source[data_key]
        # get shape
        nholes,ncoupons = data.shape
        # num of change points
        ncp = source[data[0,0]].shape[-1]
        # create index vectors
        holes = np.arange(nholes)
        coupons = np.arange(ncoupons)
        index = np.arange(source[data[0,0]].shape[-1])
        # column names
        cnames = ["Time","CP"]
        # if the user wants a dataframe
        if as_df:
            data_pack = pd.DataFrame({vn: # variable name
                                      {np.unravel_index(ii,(nholes,ncoupons))+(di,): # store multiindex by hole,coupon
                                           source[ref][vi,di] # resolve reference and store row
                                           for di in range(ncp) # unpack index
                                        for ii,ref in enumerate(data[()].flatten())} # iterate over cell
                                    for vi,vn in enumerate(cnames)}, # iterate over column names
                        columns=cnames,dtype=np.float64,index=pd.MultiIndex.from_product([holes, coupons,index],names=["Hole","Coupon","Index"]))
        # if the user wants a numpy array
        else:
            # create empty array to store values
            data_pack = np.zeros([item for sublist in [data.shape,source[data[0,0]].shape] for item in sublist],dtype=source[data[0,0]].dtype)
            # iterate over cell array
            for ii,ref in enumerate(data[()].flatten()):
                # find index
                hi,ci = np.unravel_index(ii,(nholes,ncoupons))
                # update array
                data_pack[hi,ci,:,:] = source[ref][()]
    return data_pack


def findArrOfSize(path,sz=[695,696]):
    '''
        Search a target MAT or HDF5 file for arrays whose largest dimension is within the target list

        Inputs:
            path : File path to target file
            sz : List of target sizes to serch for

        Returns dictionary of paths and identified datasets
    '''
    arr_dict = {}
    with h5py.File(path,'r') as source:
        # search groups that contain datasets that match the target conditions
        def find_arrays(name,item):
            if isinstance(item,h5py.Dataset) and (max(item.shape) in sz) and (item.dtype == np.float64):
                # get list of entries
                val = arr_dict.get(name,[])
                val.append(item)
                arr_dict[name.split('/')[-1]] = val
        source.visititems(find_arrays)
    return arr_dict

def loadMatVarNames(path):
    '''
        Get the unique variable names stored in the MAT file

        Inputs:
            path : File path to MAT file

        Returns set of variable names
    '''
    refs = []
    with h5py.File(path,'r') as source:
        def load_varnames(name,item):
            # find datasets whose path includes varNames
            # they are references to other arrays in the dataset
            if isinstance(item,h5py.Dataset):
                if 'data_array' in name:
                    return
                if '#subsystem#' in name:
                    return
                if "varNames" in name:
                    refs.append(item[()])
        # get the collection of references to datasets
        source.visititems(load_varnames)
        vnames = []
        # iterate over ref arrays found
        for rr in refs:
            # flatten for easier iteration
            for kk in rr.flatten():
                # sometimes variable names are 0
                # so skip them
                if kk==0:
                    continue
                # variable name is an array of character codes
                # conver to chars and join together to form string
                vn = ''.join([chr(cc) for cc in source[kk][()].flatten()])
                vnames.append(vn)
    return set(vnames)

def loadMatNumVars(path):
    '''
        Get the number of variables stored in the file according to numVars dataset

        Input:
            path : File path to MAT file

        Returns set of number of unique vars found
    '''
    refs = []
    with h5py.File(path,'r') as source:
        def load_varnames(name,item):
            # find datasets whose path includes varNames
            # they are references to other arrays in the dataset
            if isinstance(item,h5py.Dataset):
                if "numVars" in name:
                    refs.append(item[()][0,0])
        # get the collection of references to datasets
        source.visititems(load_varnames)
    return set(refs)

def loadMatNumRows(path):
    '''
        Get the number of variables stored in the file according to numRows dataset

        Input:
            path : File path to MAT file

        Returns set of number of rows
    '''
    refs = []
    with h5py.File(path,'r') as source:
        def load_numrows(name,item):
            # find datasets whose path includes varNames
            # they are references to other arrays in the dataset
            if isinstance(item,h5py.Dataset):
                if "numRows" in name:
                    refs.append(item[()][0,0])
        # get the collection of references to datasets
        source.visititems(load_numrows)
    return set(refs)

def loadMatNumDims(path):
    '''
        Get the number of variables stored in the file according to numDims dataset

        Input:
            path : File path to MAT file

        Returns set of number of dimensions
    '''
    refs = []
    with h5py.File(path,'r') as source:
        def load_numdims(name,item):
            # find datasets whose path includes varNames
            # they are references to other arrays in the dataset
            if isinstance(item,h5py.Dataset):
                if "numDims" in name:
                    refs.append(item[()][0,0])
        # get the collection of references to datasets
        source.visititems(load_numdims)
    return set(refs)

def loadMinCompatVersion(path):
    '''
        Get the minimum compatible version the MAT file is valid with

        Input:
            path : File path to MAT file

        Returns set of version numbers
    '''
    refs = []
    with h5py.File(path,'r') as source:
        def load_vers(name,item):
            # find datasets whose path includes varNames
            # they are references to other arrays in the dataset
            if isinstance(item,h5py.Dataset):
                if "minCompatibleVersion" in name:
                    refs.append(item[()][0,0])
        # get the collection of references to datasets
        source.visititems(load_vers)
    return set(refs)

def loadversSavesFrom(path):
    '''
        Get version of MATLAB the file was saved with

        Input:
            path : File path to MAT file

        Returns set of version numbers
    '''
    refs = []
    with h5py.File(path,'r') as source:
        def load_vers(name,item):
            # find datasets whose path includes varNames
            # they are references to other arrays in the dataset
            if isinstance(item,h5py.Dataset):
                if "versionSavedFrom" in name:
                    refs.append(item[()][0,0])
        # get the collection of references to datasets
        source.visititems(load_vers)
    return set(refs)

def loadMatShapes(path,rc=True):
    '''
        Get the unique shapes of datasets stored in teh MAT file

        Datasets stored in MATs are 1D arrays e.g. (1,6) so the shape stored
        is the larger value.

        Inputs:
            path : File path to MAT file
            rc : Return number of counts

        Returns the unique shapes and if rc is True the number of each one found
    '''
    sh = []
    # open h5py file
    with h5py.File(path,'r') as source:
        # function for searching
        def load_shapes(name,item):
            # ignore groups
            if isinstance(item,h5py.Dataset):
                sh.append(max(item.shape))
        source.visititems(load_shapes)
    # use np.unique to compress results down    
    return np.unique(sh,return_counts=rc)

def loadMatTimetables(path,sort=False,return_index=True):
    '''
        Find the MATLAB timetables in the MAT file

        inputs:
            path : File path to MAT file
            sort : Flag to sort the timetables in ascending order. Default False

        Returns a dictionary of HDF5 paths and the timetable
    '''
    time_dict = {}
    with h5py.File(path,'r') as source:
        def find_timetables(name,item):
            if isinstance(item,h5py.Dataset):
                if 'MATLAB_class' in item.attrs:
                    if b'timetable' in item.attrs['MATLAB_class']:
                        time_dict[name] = item[()]
        source.visititems(find_timetables)
    # sort array 
    if sort:
        time_dict = {k:v for k,v in sorted(time_dict.items(),key=lambda item:item[1].flatten()[-2])}
    return time_dict

def loadMatSampleRates(path):
    '''
        Get the unique sample rates stored in the MAT file

        Inputs:
            path : File path to MAT file

        Returns set of sample rates
    '''
    refs = []
    with h5py.File(path,'r') as source:
        def load_varnames(name,item):
            if isinstance(item,h5py.Dataset):
                if "sampleRate" in name:
                    refs.append(item[()])
        # get the collection of variable names
        source.visititems(load_varnames)
        srates = []
        # iterate over ref arrays found
        for rr in refs:
            srates.append(rr[0][0])
    return set(srates)

def decodeVarname(arr):
    '''
        Convert numpy array of character codes to a string

        Strings are stored as an array of character codes. This is a utility
        function for receiving variable names from a MAT file.

        Inputs:
            arr : Numpy array of integer codes

        Returns decoded string
    '''
    return ''.join([chr(cc) for cc in arr.flatten()])

def sortMatPaths(kk):
    '''
        Sort the MATLAB paths into the appropriate order

        MATLAB generates keys in alphabetical order starting with
        lowercase letters, then capital and then numbers

        This function is mean to be passed to sorted or similar functions
        as the key parameter
    '''
    # get final part of path
    pts = kk.split('/')
    last_pt = pts[-1]
    # get the first and 2nd characters of the part
    if len(last_pt)==1:
        pt = last_pt[-1]
        pt_2 = ''
    else:
        pt = last_pt[-2]
        pt_2 = last_pt[-1]
    # retun tuple of sorting conditions
    # sort from left to rights
    # True > False
    return (not pt.islower(), pt.isnumeric(), pt, pt_2)

def estimateMatDataSize(fname,as_bytes=False):
    '''
        Estimate the total memory size of the data arrays stored in a MAT file

        Searches for the datasets called 'data' stored under groups. The size of the array data in bytes is calculated
        and stored in a list. This is summed to get the total size. If as_bytes is False, the size is converted form Bytes
        to the largest file size denominator.

        Inputs:
            fname : Source filepath
            as_bytes : Flag to return size as bytes. If False, size is converted to largest file denominator

        Returns total data size and string representing unit.
    '''
    # dictionary of values
    metadata = []
    # open source file
    with h5py.File(fname,'r') as source:
        # function for searching for data arrays
        def find_arrays(name,item):
            # data arrays are stored under a group
            if isinstance(item,h5py.Group):
                if 'data' in item:
                    # update dictionary by 
                    metadata.extend([source[ref].size * source[ref][()].itemsize for ref in item['data'][()].flatten()])
        source.visititems(find_arrays)
    total_sz = sum(metadata)
    if as_bytes:
        return total_sz,"B"
    else:
        convertSizeToUnits(total_sz)

############################################## TDMS ##############################################
def convertTDMSToHDF5(path,cp=9):
    '''
        Wrapper for TdmsFile.read.as_hdf function

        Repacks the target TDMS file as a HDF5 file of the same name.

        Compresses datasets using GZIP to the level cp

        Inputs:
            path : Path to TDMS file
            cp : Compression level, 0-9
    '''
    # get filename
    fname = os.path.splitext(os.path.basename(path))[0]
    # open file using context manager
    with TdmsFile.read(path) as file:
        # create uncompressed file using context manager
        file.as_hdf(fname+".hdf5").close()
        # if the user doesn't want it compressed, return
        if not cp:
            return
        # reopen uncompressed in read mode
        with h5py.File(fname+".hdf5",'r') as source:
            # open destination file to hold compressed results
            with h5py.File(fname+"-compressed.hdf5",'w') as dest:
                # copy properties across
                for kk,vv in file.properties.items():
                    dest.attrs[kk] = vv
                # start iterating over the source MAT file
                def repack_dset(name,item):
                    # only datasets have dtype attribute
                    if isinstance(item,h5py.Dataset):
                        # create a dataset with compression opts
                        ds = dest.create_dataset(name,chunks=item.shape,dtype=item.dtype,data=source[name][()],
                                            compression="gzip",compression_opts=cp)
                        #ds = dest.create_dataset(name,shape=item.shape,dtype=item.dtype)
                        # update the attributes
                        ds.attrs.update(**source[name].attrs)
                    # if a group
                    elif isinstance(item,h5py.Group):
                        gp = dest.require_group(name)
                        # get attributes
                        attrs = {kk:vv for kk,vv in source[name].attrs.items()}
                        # copy attributes to dest
                        gp.attrs.update(attrs)
                # visit each item in source and repack it in the compressed file
                source.visititems(repack_dset)
        # delete uncompressed file
        if os.path.exists(fname+".hdf5"):
            os.remove(fname+".hdf5")

def convertTDMSToNPZ(path):
    '''
        Convert the contents of TDMS file to compressed NPZ

        Opens TDMS file, converts to dataframe, converts to dictionary and then saves as NPZ

        The array names are based off the original name. Slashes and various characters are replaced
        for easier indexing

        e.g. /'H1'/'PXI1Slot2\ai0' -> H1_PXI1Slot2_ai0

        Inputs:
            path : Filename to TDMS file
    '''
    # dictionary to hold values
    data_dict = {}
    # open TDMS file
    with TdmsFile.read(path) as file:
        # convert to dataframe
        data = file.as_dataframe()
        # iterate over entries
        for cc,ii in data.items():
            # format column names
            dname = cc.replace("/'",'').replace("'","_").replace('\\',"_")[:-1]
            # create entry in dictionary
            data_dict[dname] = ii.values
    # save to compressed file
    np.savez_compressed(os.path.splitext(os.path.basename(path))[0]+".npz",**data_dict)

def findNonEmptyChannels(path,find='all'):
    '''
        Search TDMS file for non-empty channels.

        Find parameter controls how far to search and what to return

        Assumes single-level depth (one layer of groups with one layer of channels)

        Inputs:
            path : File path to target TDMS file
            find : How to search the file.
                Supported strings:
                    all : Find all non-empty channels and return a list of tuples with their group and channel
                    first : Return the group and channel name of the first found non-empty channel
                    
        Returns a list. If find is 'first', then it's a 2-element list of group and channel name of the first non-empty channel.
        If find is 'all', then it's a list of tuples containing the group and channel names of all found non-empty channels
    '''
    if not (find in ['all','first']):
        raise ValueError(f"Invalid search mode {find}")
    nec = []
    with TdmsFile(path) as tdms_file:
        for gg in tdms_file.groups():
            for cc in gg.channels():
                if len(cc):
                    if find=='first':
                        return [gg.name,cc.name]
                    elif find=='all':
                        nec.append((gg.name,cc.name))
    return nec

def convertTDMSToParquet(path,inc_time=False):
    '''
        Convert the TDMS file to a Parquet compressed file

        Parses the data in a TDMS file to create a dictionary and writes it to a Parquet file of the same
        name. Only non-empty channels found are added to the 

        Inputs:
            path : Input path to TDMS file
            inc_time : Flag to include time data in the new file. Increases end file size.
    '''
    import pyarrow as pa
    import pyarrow.parquet as pq
    # find non-empty channels
    channels = findNonEmptyChannels(path)
    if not channels:
        print("Unable to find non-empty channels!")
        return
    # open file
    with TdmsFile(path) as tdms_file:
        # if the user wants timestamps to be included
        if inc_time:
            # construct dict of the data
            data_dict = {r'/'.join([gg,cc,'data']) : tdms_file[gg][cc] for gg,cc in channels}
            # consttruct dict of the timestamps
            time_dict = {r'/'.join([gg,cc,'time']) : tdms_file[gg][cc].time_track() for gg,cc in channels}
            data_dict.update(time_dict)
            # update data dictionary with timestamp dictionary
            # create table out of combined dictionary
            table = pa.table(data_dict)
        # if the user doesn't want time
        # only construct data dictionary
        else:
            table = pa.table({r'/'.join([gg,cc,'data']) : tdms_file[gg][cc] for gg,cc in channels})
        # write table to file
        pq.write_table(table,os.path.splitext(os.path.basename(path))[0]+'.parquet')

############################################## NPZ ##############################################    
def loadSetitecNPZ(path,columns,**kwargs):
    '''
        Load Setitec NPZ file and convert it to a Pandas DataFrame

        When building the files, the column names were placed in a separate text file where they were on separate lines.

        Inputs:
            path : File path to NPZ file
            columns : File path to columns text file or an iterable list of column names
            kwargs: Keyword arguments affecting how it's returned
                Supported keywords:
                    as_np: Return as a structured dtype
                    as_rec : Return as a np.recarray
                If no kwargs are given, then a pandas.DataFrame is given
    '''
    # if the user has given a path to a text file where the columns are
    # load it
    if os.path.isfile(columns):
        cols = np.genfromtxt(columns,dtype=str)
    # if they've given something else, it's assumed to be an iterable list of the names
    else:
        cols = columns
    # load the NPZ file
    data = np.load(path)['arr_0']
    # if the number of given columns does not match the data shape then rasie ValueError
    if len(cols)<data.shape[1]:
        raise ValueError(f"Number of columns does not match number of rows [{len(cols)} vs {data.shape[0]}]")
    # create dataframe passing data and columns
    if kwargs.get('as_np',False):
        # construct datatype from column names
        # assuming float for all
        dtype = np.dtype([(cc,'f4') for cc in cols])
        # data is transposed as numpy is row-major
        return data.T.astype(dtype)
    elif kwargs.get('as_rec',False):
        # construct datatype from column names
        # assuming float for all
        dtype = np.dtype([(cc,'f4') for cc in cols])
        # same behaviour as as_np flag but converts it to a np.recarray
        return data.T.astype(dtype).view(np.recarray)
    # else return pandas DataFrame
    else:
        return pd.DataFrame(data,columns=cols)

def repackSetitecNPZ(path,columns,**kwargs):
    '''
        Repack the Setitec NPZ file into another format

        Target filetype and some options are controlled by which flags are set in kwargs.

        Assumes the user has the required packages installed.

        If no kwargs are given, then the function returns immediately without even opening the file

        Inputs:
            path : Path to source file
            columns : File path to columns text file or an iterable list of column names
            kwargs: Flags to control what file it's written to. If no flags are given, nothing happens
                Supported flags:
                    as_hdf : Write to HDF5. Datasets are named after the columns, compression level is 9 and algorithm is gzip.
                    as_npz_cols : Repack to a new NPZ file. Data is organised by the columns instead. Filename has _repack appended on it
                    as_parq : Write to Parquet file. Relies on use_fast and use_arrow to set which engine. Compression is set to gzip
                    use_fast : Use fastparquet. Requires as_parq.
                    use_arrow : Use pyarrow. Requires as_parq.
    '''
    # if kwargs is empty
    # don't do anything
    if not kwargs:
        return
    # if the user has given a path to a text file where the columns are
    # load it
    if os.path.isfile(columns):
        cols = np.genfromtxt(columns,dtype=str)
    # if they've given something else, it's assumed to be an iterable list of the names
    else:
        cols = columns
    # load the NPZ file
    data = np.load(path)['arr_0']
    # if the number of given columns does not match the data shape then rasie ValueError
    if len(cols)<data.shape[1]:
        raise ValueError(f"Number of columns does not match number of rows [{len(cols)} vs {data.shape[0]}]")
    # get shape of array
    sh,cc = data.shape
    # if the user wants a hdf5
    if kwargs.get('as_hdf',False):
        # create file
        with h5py.File(os.path.splitext(os.path.basename(path))[0]+".hdf5",'w') as dest:
            # iterate over columns
            for ci,cc in enumerate(cols):
                # create dataset
                # require_dataset is used in case it's an existing file
                dest.require_dataset(cc,shape=(sh,),dtype=np.float64,data=data[:,ci], # dataset called column name
                                     compression="gzip",compression_opts=9) # set compression to max
    # repack to and NPZ where it's organised by columns instead of a single array
    elif kwargs.get('as_npz_cols',False):
        # repack to a NPZ where it's organised by column
        np.savez_compressed(os.path.splitext(os.path.basename(path))[0]+"_repack.npz",
                            **{cc:data[:,ci] for ci,cc in enumerate(cols)})
    # repack to a heavily compressed parquet file
    elif kwargs.get('as_parq',False):
        # set parquet engine
        if kwargs.get('use_fast',False):
            eng = 'fastparquet'
        elif kwargs.get('use_arrow',False):
            eng = 'pyarrow'
        else:
            eng = 'auto'
        # convert to pandas dataframe and then write to parquet file        
        pd.DataFrame(data,columns=cols).to_parquet(os.path.splitext(os.path.basename(path))[0]+".parquet",engine=eng,
                                                   compression="gzip")# use gzip compression
############################################## JSON ##############################################
def getAllFromJSON(path):
    '''
        Loads Setitec JSON file and collapses the process and program data into lists

        The program data is a list of dictionaries stating the parameter for each step.

        The process data is also a list of dictionaries and contains the recorded variables for each sample.

        The data is sorted by step number and the data is collapsed into a list.

        Inputs:
            path: File path to target JSON file

        Returns general parameters as dict, cycle parameters as dict, program data as a dict and process data as a dict
    '''
    # load json file
    source = json.load(path)
    # the program is a JSON string
    data = json.loads(source["program"])
    # the program is individual dictionaries of values
    # get the keys from the first one
    keys = set(data[0].keys())
    # collapse each step together where each key-value pair is a list of the values
    # the dictionary is sorted by the step number to guarantee correct order
    prog_dict = {kk : [step[kk] for step in sorted(data,key=lambda x : x["Step Nb"])] for kk in keys}
    # load process data
    # the program is a JSON string
    data = json.loads(source["process_data"])
    # the program is individual dictionaries of values
    # get the keys from the first one
    keys = set(data[0].keys())
    # collapse each step together where each key-value pair is a list of the values
    # the dictionary is sorted by the step number to guarantee correct order
    process_dict = {kk : [step[kk] for step in sorted(data,key=lambda x : x["Step Nb"])] for kk in keys}
    # return all
    return source["general_parameters"],source["cycle_parameters"],prog_dict,process_dict

def writeSetitecToPDF(path,opath=None,pdf=None,add_title_page=True,**kwargs):
    '''
        Plot the data in a Setitec file as a multipage PDF

        The file is loaded in for each column the data is plotted against the position data
        and the plot written to a page on a PDF

        Users can pass an existing pdf reference to write the plots to an already open file.
        This may be used when writing the plots for a specific setitec box to a single PDf

        When add_title_page is True, the first thing written is a blank page with the filename
        in size 24 font in the centre. This is useful when writing to an existing PDF and you
        want to separate the plots from different files.

        Example use for writing several files to a single PDF
        -----------------------------
        with PdfPages('summary.pdf') as pdf:
            for fn in glob(path):
                writeSetitecToPDF(fn,pdf=pdf,add_title_page=True)

        Input:
            path : Path to Setitec XLS file
            opath : Output path for PDF file. If None, it is set to the filename + -pdf-summary. Default None.
            pdf : PdfPages object to an already opened PDF. Default None.
            add_title_page: Write a title page with the filename to the PDF
    '''
    from matplotlib.backends.backend_pdf import PdfPages
    import matplotlib.pyplot as plt
    # get filename
    fname= os.path.splitext(os.path.basename(path))[0]
    # if output path is not defined
    if opath is None:
        # set it to the filename with pdf-summary added onto the end
        opath = fname+'-pdf-summary.pdf'
    if pdf is None:
        pdf = PdfPages(opath)
    # flag indicating if first page    
    first = True
    # load file including metadata
    vers = kwargs.get('use_version','auto')
    all_data = loadSetitecXls(path,version=vers)
    # plot data
    if vers != "auto_data":
        plot_data = all_data.pop(-1)
        md = {}
        for dd in all_data:
            if isinstance(dd,dict):
                md.update(dd)
    else:
        plot_data = all_data
    # if the user wants a title page to be added
    if add_title_page:
        # create figure
        titlepage = plt.figure()
        # remove default axes
        titlepage.clf()
        # set text to the filename
        txt = fname
        # write text to PDF in the middle of the page
        titlepage.text(0.5,0.5,txt, transform=titlepage.transFigure, size=18, ha="center")
        # send page to pdf
        pdf.savefig(titlepage)
        # cleanup figure
        plt.close(titlepage)
        # attach metdata notes to title page
        if vers != "auto_data":
            pdf.attach_note('\n'.join([f'{k}: {v}' for k,v in md.items()]))
        first = False
    # get position data to use as x-axis
    pos = plot_data['Position (mm)'].values.flatten()
    # iterate over columns
    for kk in plot_data.keys():
        # make plotting axes
        f,ax = plt.subplots()
        # plot the position against key data
        ax.plot(pos,plot_data[kk].values.flatten(),'b-')
        # set axes with title as key
        ax.set(xlabel='Position (mm)',ylabel=kk,title=kk)
        # set figure title as filename
        f.suptitle(fname)
        if first:
            pdf.attach_note('\n'.join([f'{k}: {v}' for k,v in md.items()]))
            first = False
        # write figure to pdf
        pdf.savefig(f)
        # close figure to save memory
        plt.close(f)
    
def getProcessFromJSON(path,as_pd=False):
    '''
        Similar to getAllFromJSON but only returns the process data dictionary

        Inputs:
            path : File path to target JSON file
            as_pd : Flag to return data as a Pandas DataFrame

        Returns the process data as a dictionary of lists for the data at each step or a pandas data frame if as_pd is True
    '''
    # load json file
    source = json.load(path)
    # the program is a JSON string
    data = json.loads(source["process_data"])
    # the program is individual dictionaries of values
    # get the keys from the first one
    keys = set(data[0].keys())
    # collapse each step together where each key-value pair is a list of the values
    # the dictionary is sorted by the step number to guarantee correct order
    process_dict = {kk : [step[kk] for step in sorted(data,key=lambda x : x["Step (nb)"])] for kk in keys}
    # return
    if as_pd:
        return pd.DataFrame.from_dict(process_dict)
    else:
        return process_dict
    
def getProgramFromJSON(path,as_pd=False):
    '''
        Similar to getAllFromJSON but only returns the program data dictionary

        Inputs:
            path : File path to target JSON file
            as_pd : Flag to return data as a Pandas DataFrame

        Returns the program data as a dictionary of lists for the data at each step or a pandas data frame if as_pd is True
    '''
    # load json file
    source = json.load(path)
    # the program is a JSON string
    data = json.loads(source["program"])
    # the program is individual dictionaries of values
    # get the keys from the first one
    keys = set(data[0].keys())
    # collapse each step together where each key-value pair is a list of the values
    # the dictionary is sorted by the step number to guarantee correct order
    prog_dict = {kk : [step[kk] for step in sorted(data,key=lambda x : x["Step Nb"])] for kk in keys}
    if not as_pd:
        return prog_dict
    else:
        return pd.DataFrame.from_dict(prog_dict)

def getHoleCouponFromJSON(path,as_id=False):
    '''
        Gets the hold and coupon from the JSON file.

        The dictionary contains two entries for hole and column, raw number nf ID. The ID is a string and the raw number is
        an integer. The value of the as_id flag controls what's returned

        Inputs:
            path : File path to target JSON file
            as_id : Flag controlling whether to return the hole ID or the raw data

        Returns the hole and coupon as a formatted string or an integer depending on the valid of as_id.
    '''
    # load json file
    source = json.load(path)
    return source["general_parameters"]["hole_id" if as_id else "hole_n"],source["general_parameters"]["coupon_id" if as_id else "coupon_n"] 


############################################## KISTLER ##############################################
# function to parse the text file
def loadKistlerText(fn,as_df=True):
    '''
        Parse Airbus Drilling Text Data File

        The Airbus drilling text file is a collection of data from a given trial.
        The data is the forces experienced by the Kissler plate.

        Inputs:
            fn : Input path
            as_df : Flag to return data as a Pandas DataFrame. Default True.

        Returns pd.DataFrame if as_df is True. If False, then a list of tuples is
        returned where each entry is the variable name, sample rate and array of values.
    '''
    # if the user wants it whole
    # set data as a dictionary to populate
    if as_df:
        data = {}
    # if the user wants it in bits
    # set data as a list
    else:
        data = []
    # open text file
    with open(fn,'r') as text:
        # read first line
        # header info
        # can ignore it
        text.readline()
        # next 4 rows are the actual data values
        # read line
        for ll in text:
            # strip and split about tab character
            pts = ll.strip().split('\t')
            # get variable name
            var = pts.pop(0)
            # get sampling rate of variable
            sr = pts.pop(0)
            # if whole
            if as_df:
                # add to dictionary
                data[var] = {}
                data[var] = np.array(pts)
            # if as pts
            else:
                data.append((var,float(sr),np.array(pts,dtype='float32')))
    if as_df:
        return pd.DataFrame.from_dict(data,dtype='float32')
    else:
        return data

def loadKistlerSetitecXls(fn):
    '''
        Load Setitec data associated with the Kissler Plate data

        This is a different function to loadSetitecXls due to the different control box firmware versions producing a different format XLS file

        Inputs:
            fn : Filepath to to XLS file

        Returns the separate parts of the datset as a Panda datasets in the following order

        - General parameters -> dict
        - Contorl Box Parameters -> dict
        - Motor Paramters -> dict
        - Head Parameters -> dict
        - Pset -> dict
        - Cycle Paramters -> dict
        - Program parameters -> Pandas DataFrame
        - Data parameters -> dict
        - Run Data -> Pandas DataFrame
    '''
    # direct users over to loadSetitecXls functions due to differing firmware versions
    warnings.warn("Recommended to use loadSetitecXls with version set to auto due to differing firmware versions",category=DeprecationWarning)
    with open(fn,'r',encoding='latin-1') as open_file:
        # *** General Infos ***
        l = open_file.readline()
        l_head = open_file.readline().strip("\n").split("\t")
        l_tail = open_file.readline().strip("\n").split("\t")
        general_parameters = dict(zip(l_head, l_tail))
        # skip empty line
        open_file.readline()
        
        # *** Control Box Datas ***
        # skip star header
        open_file.readline()
        # read line header
        l_head = open_file.readline().strip("\n").split("\t")
        
        # there are a bunch of empty lines
        # skip them
        for l in open_file:
            if l.strip('\n'):
                break
        # for some reason the keys are actually split
        l_head.extend(l.strip('\n').split('\t'))
        # all the values are in the next line
        l_tail = open_file.readline().strip('\n').split('\t')
        control_box_datas = dict(zip(l_head,l_tail))
        # skip the next empty line
        open_file.readline()
        
        # *** Motor Datas ***
        # skip section head
        open_file.readline()
        l_head = open_file.readline().strip("\n").split("\t")
        l_tail = open_file.readline().strip("\n").split("\t")
        motor_datas = dict(zip(l_head,l_tail))
        # skip empty line
        open_file.readline()

        # *** Head Datas ***
        # skip section head
        open_file.readline()
        l_head = open_file.readline().strip("\n").split("\t")
        l_tail = open_file.readline().strip("\n").split("\t")
        head_datas = dict(zip(l_head,l_tail))
        # skip empty line
        open_file.readline()

        # *** Pset ***
        # skip section head
        open_file.readline()
        l_head = open_file.readline().strip("\n").split("\t")
        l_tail = open_file.readline().strip("\n").split("\t")
        Pset = dict(zip(l_head,l_tail))
        # skip empty line
        open_file.readline()

        # *** Cycle parameters ***
        # skip section head
        open_file.readline()
        # read first header
        l_head = open_file.readline().strip("\n").split("\t")
        # read first tail
        l_tail = [float(vv.replace(',','.')) for vv in open_file.readline().strip("\n").split("\t")]
        l_head.extend(open_file.readline().strip("\n").split("\t"))
        l_tail.extend([float(vv.replace(',','.')) for vv in open_file.readline().strip("\n").split("\t")])
        cycle_params = dict(zip(l_head,l_tail))
        # skip empty line
        open_file.readline()
        
        # *** Program ***
        open_file.readline()
        ll = []
        column_names = None
        while (True):
            l = open_file.readline().strip("\n").split("\t")
            if (l[-1] == ''): break
            if column_names is None:
                column_names = l
                continue
            # decimals are comma format e.g. 3,000 -> 3.000 or 6,000,000 -> 6000.000
            # replace last comma with dots
            # remove all others
            ll.append([_replaceLast(tt,',','.').replace(',','') for tt in l])
        program = pd.DataFrame(ll, columns=column_names)
        
        # *** Results ***
        open_file.readline()
        l_head = open_file.readline().strip("\n").split("\t")
        l_tail = [tt.replace(',','.') for tt in open_file.readline().strip("\n").split("\t")]
        results = dict(zip(l_head,l_tail))
        open_file.readline()

        # *** Datas ***
        open_file.readline()
        ll = []
        while (True):
            l = open_file.readline().strip("\n").split("\t")
            if (l[-1] == ''): break
            ll.append(l)
        column_names = ll.pop(0)
        data = pd.DataFrame(ll, columns=column_names)
        
        for column in data.columns:
            data[column] = data[column].str.replace(',','.').astype(float)

        return general_parameters,control_box_datas,motor_datas,head_datas,Pset,cycle_params,program,results,data

def _combineData(dir_path,procs=None):
    '''
        Combine all files found in the specified location into a single DataFrame

        Each file is loaded and a new column called HOLENUM is added representing the
        holenumber for the data. A unique value is applied to each batch of data so it
        cn be searched.

        Inputs:
            dir_path : Path to a location containing XLS files
            procs : Number of processes to use to parallelize loading the files. If None,
                    it's done sequentially. If True, 3 is passed.

        Returns a DataFrame combining all found values
    '''
    from multiprocessing import Pool
    if isinstance(dir_path,str):
        # if not a wildcard path make it one
        if not ('*.xls' in dir_path):
            dir_path=os.path.join(dir_path,'*.xls')
        dir_path = glob(dir_path)
    # create list of holenumbers to use and add hole number
    holenums = range(len(dir_path))
    # function to process files
    def _procPath(fn,hf=0):
        data = loadSetitecXls(fn)[-1]
        data['HOLENUM']=hf
        return data
    # set initial dataframe as first
    curr_dr = _procPath(dir_path[0],holenums[0])
    # if not multiprocessing
    if (procs is None) or (not procs):
        # iterate over holenumbers and paths
        for hn,fn in zip(holenums,dir_path):
            data = _procPath(fn,hn)
            # update dataframe
            curr_df = data if curr_df is None else pd.concat((curr_df,data))
    else:
        curr_df = pd.concat(Pool(procs).starmap(_procPath,[(fn,hf) for fn,hf in zip(dir_path,range(len(dir_path)))]))
    # return result
    return curr_df

def loadAllSetitecXLS(path,use_mt=False,**kwargs):
    '''
        Load all Setitec XLS files spread across multiple directories

        The data from each file is combined depending on the flags given.

        Supported flags:
            per_dir : Combine all the files found in each directory into its own DataFrame and add to a list.
                    Each DataFrame has the directory added under the attribute dirname
            as_one : Combine all the files into a single DataFrame. The directories are stored as a set under
                    the attribute dirnames

        Only ONE flag should be given! If multiple are given, a ValueError will be raised.

        e.g. loadAllSetitecXLS("AirbusData/*.xls",per_dir=True)

        Use per_dir if giving a path with multiple directories within it and as_one if giving a single directory

        Input:
            path : Wildcard path containing multiple directories
            use_mt : Flag to use a Pool to parallelize loading the files. If True then 3 cores are used. User
                    can also give an int representing number of cores to use
            Supported customisation flags:
                per_dir : Flag to store the files in each directory in its own DataFrame
                as_one : Flag to store all files under a single DataFrame

        Returns either a single pandas DataFrame os a list of them as per the set flags.
    '''
    # if no customisation flags were given
    if len(kwargs)==0:
        raise ValueError("No customisation flags are given!")
    # if more than one flag was given
    if len(kwargs)>1:
        raise ValueError("More than 1 customisation flag was given! Received {len(kwargs)}")
    if isinstance(use_mt,bool):
        use_mt = 3
    # create flags
    per_dir = False
    as_one = False
    # collect collection flags
    if 'per_dir' in kwargs:
        per_dir = kwargs['per_dir']
    elif 'as_one' in kwargs:
        as_one = kwargs['as_one']
    # if combining files per directory
    if per_dir:
        # iterate over each directory and combine the files in each to a dataframe adding to a list
        # skipping directories without a XLS
        dir_list = []
        for _,dirs,_ in os.walk(path):
            for dd in dirs:
                if len(glob(os.path.join(path,dd,'*.xls')))==0:
                    continue
                dir_list.append(_combineData(os.path.join(path,dd,'*.xls')),use_mt)
        return dir_list
    # if combining all files into one
    elif as_one:
        # if there are no files return None
        if len(glob(path,recursive=True))==0:
            return None
        return _combineData(glob(path,recursive=True),use_mt)

def findMinSetitecThickness(path):
    '''
        Search the Setitec file for Distance min (mm) and sum values in column

        In the Setitec file is a column called Distance min (mm) that refers to the min
        expected material thickness. This value could be used to set the lower end of the
        expected depth used when calculating depth.

        Inputs:
            path : Path to Setitec XLS file

        Returns floating point values
    '''
    with open(path,'r') as file:
        for line in file:
            # strip trailing tabs from the line
            line = line.strip()
            # check it isn't empty
            if not line:
                continue
            # check if min distance is in the list
            if not ('Distance min (mm)' in line):
                continue
            pts = line.split('\t')
            # find which column has Distance min (mm)
            for i,pp in enumerate(pts):
                if 'Distance min (mm)' == pp:
                    break
            sum_min = 0
            # iterate over the next lines
            for ll in file:
                ll = ll.strip()
                if not ll:
                    break
                # get the float value
                sum_min += float(ll.split('\t')[i].replace(',','.'))
            return sum_min

def findMaxSetitecThickness(path):
    '''
        Search the Setitec file for Distance max (mm) and sum values in column

        In the Setitec file is a column called Distance max (mm) that refers to the max
        expected material thickness. This value could be used to set the upper end of the
        expected depth used when calculating depth.

        Inputs:
            path : Path to Setitec XLS file

        Returns floating point values
    '''
    with open(path,'r') as file:
        for line in file:
            # strip trailing tabs from the line
            line = line.strip()
            # check it isn't empty
            if not line:
                continue
            # check if min distance is in the list
            if not ('Distance max (mm)' in line):
                continue
            pts = line.split('\t')
            # find which column has Distance min (mm)
            for i,pp in enumerate(pts):
                if 'Distance max (mm)' == pp:
                    break
            sum_max = 0
            # iterate over the next lines
            for ll in file:
                ll = ll.strip()
                if not ll:
                    break
                # get the float value
                sum_max += float(ll.split('\t')[i].replace(',','.'))
            return sum_max

def findSetitecThicknessRange(path):
    '''
        Search the Setitec file for Distance min (mm) and Distance max (mm)
        and sum values in each column

        In the Setitec file are columns called Distance min (mm) and Distance max (mm)
        that refers to the min and max expected material thickness. This value could
        be used to set the limits on calculating expected depth used when calculating depth.

        Inputs:
            path : Path to Setitec XLS file

        Returns two floating point values representing min and max values respectively.
    '''
    with open(path,'r') as file:
        for line in file:
            # strip trailing tabs from the line
            line = line.strip()
            # check it isn't empty
            if not line:
                continue
            # check if min distance is in the list
            if not ('Distance min (mm)' in line):
                continue
            pts = line.split('\t')
            # find which column has Distance min (mm)
            for i,pp in enumerate(pts):
                if 'Distance min (mm)' == pp:
                    break
            sum_min = 0
            sum_max = 0
            # iterate over the next lines
            for ll in file:
                # check that the line isn't blank
                ll = ll.strip()
                if not ll:
                    break
                pts = ll.split('\t')
                # get the float value
                sum_min += float(pts[i].replace(',','.'))
                # max value is always next to it
                sum_max += float(pts[i+1].replace(',','.'))
            return sum_min, sum_max
        else:
            return None,None

if __name__ == "__main__":
    from glob import glob
    import matplotlib.pyplot as plt
    from math import fsum
    import matplotlib.pyplot as plt
    data = loadSetitecXls(glob('8B Life Test/*.xls')[0],'auto')
    print(getAV(glob('8B Life Test/*.xls')[0]))
        
