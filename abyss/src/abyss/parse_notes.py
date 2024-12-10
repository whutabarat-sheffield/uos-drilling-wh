import openpyxl
from openpyxl.utils import get_column_letter
import os
import pandas as pd
import numpy as np

class TestNotes:
    def __init__(self,path):
        self._path = path
        self.wb = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
        self._hl = {}

    def searchCellRange(self):
        ''' get the range of columns for hole data '''
        self._hl.clear()
        for sheet in self.wb.worksheets:
            hc = None
            rc = None
            hend = None
            for ri,row in enumerate(sheet,1):
                # if all cells are empty
                if all([not cell.value for cell in row]):
                    continue
                # search for hole number title
                for i,c in enumerate(row,start=1):
                    # if the cell is not empty and contains a string
                    if c.value and (c.data_type == 's'):
                        # check if phrase Hole is in the value
                        if ('Hole' in c.value) and (not hc):
                            hc = i
                            rc = ri
                        elif any([t in c.value.lower() for t in ['comment','note']]):
                            hend = i
                            break
                # if the column range has been found stop iterating
                if hc and hend:
                    break
            # iterate dictionary
            self._hl[sheet.title] = [rc,(hc,hend)]
        return self._hl

    # from https://stackoverflow.com/a/37713627
    @staticmethod
    def getCellRef(row, column, zero_indexed=True):
        if zero_indexed:
            row += 1
            column += 1
        return get_column_letter(column) + str(row)

    def getData(self):
        if not self._hl:
            self.searchCellRange()
        data = {}
        for k,(r,(hc,hend)) in self._hl.items():
            if (r==None) or (hc==None) or (hend==None):
                continue
            # convert to cell ref
            refA = TestNotes.getCellRef(r,hc)
            refB = TestNotes.getCellRef(r+50,hend)
            # gets tuple-of-tuple of cell refs
            data[k] = []
            for r in self.wb[k][refA:refB]:
                data[k].append([])
                for c in r:
                    data[k][-1] = c.value
        return data
                
if __name__ == "__main__":
    path = "pauls_notes/MAPAL 8B CFRP-AL life re-test.xlsx"
    rr = TestNotes(path)
    data = rr.getData()
        
    
