import pandas as pd
import numpy as np
import abyss.dataparser as dp
import matplotlib
#matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from glob import glob
import os
from tslearn.metrics import dtw
import math
import multiprocessing as mp
import seaborn as sns

# load worksheet, convert each sheet into a dataframe and stack
def stackCPSheets(path):
    '''
        Load the Counterspark metadata file and stack the sheets to form a single Pandas datasheet

        Loads each sheet in the target Counterspark metadata file, converts it to a Pandas dataframe
        and stacks them to form a single dataframe.

        Columns labelled None are removed and the dataframe is clipped to the first NaN row in Test
        Number. This is to remove unused rows in the file.

        Input:
            path : Path to target Counterspark dataframe

        Returns single Pandas dataframe
    '''
    df = None
    # load worksheet in read only
    metadata = load_workbook(path,True)
    # iterating over sheets
    for sn in metadata.sheetnames:
        sheet = metadata[sn].values
        # get columns
        cols = next(sheet)
        # convert to pandas dataframe
        sdf = pd.DataFrame(sheet,columns=cols)
        # remove None columns
        if None in cols:
            sdf = sdf.drop(columns=[None])
        # find where Test Number is N
        mask = sdf.isna()['Test Number ']
        if (mask.values.any()):
            sdf.drop(index=range(mask[mask==True].index.min(),len(sdf)),inplace=True)
        if df is None:
            df = sdf
        else:
            df = pd.concat([df,sdf],ignore_index=True)   
    return df

def loadCPDataAsPD(path):
    '''
        Load the Counterspark metadata file, convert each sheet to a dataframe and return it

        Loads each sheet in the target Counterspark metadata file, converts it to a Pandas dataframe
        and adds them to a list.

        Columns labelled None are removed and the dataframe is clipped to the first NaN row in Test
        Number. This is to remove unused rows in the file.

        Input:
            path : Path to target Counterspark dataframe

        Returns list of dataframes for each sheet in the target file
    '''
    # load worksheet in read only
    metadata = load_workbook(path,True)
    dfs = []
    for sn in metadata.sheetnames:
        sheet = metadata[sn].values
        # get columns
        cols = next(sheet)
        # convert to pandas dataframe
        sdf = pd.DataFrame(sheet,columns=cols)
        # remove None columns
        if None in cols:
            sdf = sdf.drop(columns=[None])
        # mask values
        mask = sdf.isna()['Test Number ']
        if (mask.values.any()):
            sdf.drop(index=range(mask[mask==True].index.min(),len(sdf)),inplace=True)
        dfs.append(sdf)
    return dfs

def plotKistlerPlateHist(path,**kwargs):
    '''
        Plot the history of the Kistler plate files found in path

        For each Kistler text file found in the wildcard path given by user, each
        column is processed and results plotted on separate axis.

        By default the max value of each column is given, but the user can give
        a custom function to process each column using the process keyword.

        Inputs:
            path : Wildcard path to folder containing the Kistler text files.
            process : Method used to process each column and create a single value to plot.
                    Default max.
            ftitle : Figure title to use. Default Kistler Plate History

        Returns matplotlib figure
    '''
    from matplotlib import cm
    cols = None
    proc = kwargs.get('process',max)
    # if process is set to None
    if proc is None:
        # get the binary color map
        cmap = cm.get_cmap('binary')
        # get number of files
        nf = len(sorted(glob(path)))
    c = []
    max_data = {}
    for fn in sorted(glob(path)):
        # load text file
        data = dp.loadKistlerText(fn,True)
        # if columns haven't been set yet
        if cols is None:
            # get columns
            cols = data.columns.to_list()
            # make axes
            f,ax = plt.subplots(2,2,constrained_layout=True)
            # iterate over columns and associated axis
            for cc,aa in zip(cols,ax.flatten()):
                # assign labels
                aa.set(xlabel="File Name",ylabel=cc,title=f"{proc.__name__ if proc is not None else 'Full'} {cc}")
                # initialize data structure to hold max values, axis and filenames
                max_data[cc] = ([],aa,[])
                # generate a unique color for the column
                c.append(tuple(np.random.random(size=3)))
        # iterate over each unique column
        for cc in cols:
            # attempt to get data
            # req as Air Flow key is sometimes Air Flow ADU or AirFlow
            try:
                maxd = proc(data[cc]) if proc is not None else data[cc].values
            except KeyError as exp:
                if cc == 'Air Flow ADU':
                    maxd = data['AirFlow'].max()
                else:
                    raise exp
            # add max value to list
            max_data[cc][0].append(maxd)
            max_data[cc][-1].append(os.path.splitext(os.path.basename(fn))[0])
    # iterate over collected data and plot it
    for (md,aa,fnames),col in zip(max_data.values(),c):
        # if process is set to None
        if proc is None:
            # plot each signal as a line
            # setting the color to elements of the binary colormap
            for fi,dd in enumerate(md):
                aa.plot(dd,'-',c=cmap(fi/nf))
        else:
            aa.plot(md,'x',c=col)
        # set the xtick labels to the filenames
        aa.set_xlim(0,len(md))
        aa.set_xticks(range(len(md)),labels=fnames,rotation='vertical')
    f.suptitle(kwargs.get("ftitle","Kistler Plate History"))
    return f

def plotKistlerPlateFFTHist(path,htype='full',**kwargs):
    '''
        Plot the FFT history of Kistler plate

        The input htype controls what type of history is plotted.
        Supported:
            full : FFT magnitude signals
            max  : FFT magnitude max
            min  : FFT magnitude min

        Inputs:
            path : Wildcard input path
            htype : Type of history. Supported 'full','max' and 'min'.

        Returns generated matplotlib figure
    '''
    from scipy.fft import rfft, rfftfreq
    if not ('htype' in ['full','max','min']):
        raise ValueError(f"Unsupported history type {htype}")
    cols = None
    T = 1/5000.0
    axes = {}
    for fi,fn in enumerate(sorted(glob(path))):
        # load text file
        data = dp.loadKistlerText(fn,True)
        if cols is None:
            # get columns
            cols = data.columns.to_list()
            # make axes
            f,ax = plt.subplots(2,2,constrained_layout=True)
            # iterate over columns and associated axis
            for cc,aa in zip(cols,ax.flatten()):
                # assign labels
                axes[cc] = aa
                if htype == 'full':
                    aa.set(xlabel="Frequency (Hz)",ylabel="FFT Mag",title=f"FFT {cc}")
                elif htype == 'max':
                    aa.set(xlabel="File Index",ylabel="FFT Mag Max",title=f"FFT Max {cc}")
                elif htype == 'min':
                    aa.set(xlabel="File Index",ylabel="FFT Mag Min",title=f"FFT Min {cc}")
        # iterate over columns and axes
        for cc,aa in axes.items():
            try:
                vals = data[cc].values
                N = len(data[cc])
            except KeyError as exp:
                if cc == 'Air Flow ADU':
                    vals = data['AirFlow'].values
                    N = len(data['AirFlow'])
                else:
                    raise exp
            # plot full FFT history
            if htype == 'full':
                xf = rfftfreq(N,T)
                aa.plot(xf,np.abs(rfft(vals)))
                aa.set_yscale('log')
            elif htype == 'max':
                aa.plot(fi,np.abs(rfft(vals)).max(),'bx')
            elif htype == 'min':
                aa.plot(fi,np.abs(rfft(vals)).min(),'bx')
    return f

def plotKistlerPlateSTFTHist(path):
    '''
        Plot the STFT Kistler plate history

        For each signal, calculate the STFT and overlap them.
        Color limits are set from global limits

        Inputs:
            path : WIldcard path to text file
    '''
    from scipy import signal
    from mpl_toolkits.axes_grid1 import make_axes_locatable
    axes = {}
    cols = None
    for fi,fn in enumerate(sorted(glob(path))):
        # load text file
        data = dp.loadKistlerText(fn,True)
        if cols is None:
            # get columns
            cols = data.columns.to_list()
            # make axes
            fig,ax = plt.subplots()
            # iterate over columns and associated axis
            for cc,aa in zip(cols,ax.flatten()):
                # assign labels
                axes[cc] = [aa,[]]
                axes[cc][0].set(xlabel='Time (s)',ylabel='Frequency (Hz)',title=cc)
        # iterate over columns and axes
        for cc in cols:
            try:
                vals = data[cc].values
                N = len(data[cc])
            except KeyError as exp:
                if cc == 'Air Flow ADU':
                    vals = data['AirFlow'].values
                    N = len(data['AirFlow'])
                else:
                    raise exp
            # caclulate the spectrogram and add it to the list
            f,t,Sxx = signal.spectrogram(vals,5000)
            axes[cc][1].append([f,t,Sxx])
    # iterate over items
    for cc,(aa,data) in axes.items():
        vmin = min([d[2].min() for d in data])
        vmax = max([d[2].max() for d in data])
        for f,t,Sxx in data:
            pcm = aa.pcolormesh(t,f,Sxx,shading='gouraud',alpha=0.7,vmin=vmin,vmax=vmax)
    return fig
    
    
def plotCounterspark(data_path,mpath,xaxis='Tool Life',yaxis='Max Torque',target_d='4B',**kwargs):
    '''
        Plot an axis of Counterspark metadata against data from the Setitec files

        The metadata file is loaded using stackCPSheets and filtered to the target diameter specified by
        target_d. The Setitec hole count is extracted from the corresponding column. For each hole count,
        the target file is searched for in data_path. If it is not found it is skipped.

        xaxis is a column from the Counterspark metadata file and is used as the x-axis values.

        If the xaxis is for categorical data (e.g. Tool Life) values are generated for each unique value (i.e. 0,1,2 etc.)
        and the category labels are used as the tick labels
        
        yaxis is data extracted from each found file using the specified method.

        For e.g. if yaxis is the string 'Max Torque', then for each target file the max torque is extracted.
        If no_empty is True, then the Empty channel is not used in the max calculation.

        Supported strings are 'Max Torque','Min Torque','Max Thrust','Min Thrust'. This is also used as the
        y-axis label.

        If yaxis is not a string, it is treated as a callable and the extracted Setitec runtime data is
        passed to it
        i.e.
        data = dp.loadSetitecXls(fn,'auto_data')
        val = yaxis(data)

        The name of the function is used as the y-axis label unless the user has specified a label using
        ylabel keyword.

        Example of use
        -------------------------
        e.g. 1 : Plotting max torque for each hole against the Test Number column
        
        plotCounterspark(r"COUNTERSPARK/data/4B/*.xls","COUNTERSPARK/Counterspark drilling test.xlsx",xaxis='Test Number',yaxis='Max Torque')


        Inputs:
            data_path : Wildcard path for Setitec data files
            mpath : Direct path to Counterspark metadata file
            xaxis : Target column key of Counterspark metadata file to use as x-axis values. Default Tool Life.
            yaxis : Processed data from found target files to use on yaxis. String or callable. Default Max Torque.
            target_d : Target diameter filter for. Default 4B.
            no_empty : Do not use empty channel when calculating min or max value. Only used when supported string is passed to yaxis.
            ylabel : Y-axis label. If not, given yaxis is used if it's a string or the name of the function if yaxis is callable.
            ftitle : Figure title. Default f"Counterspark {xaxis} vs {yaxis}".
            no_lim : Flag to not set the x-limits. Default False.

        Returns a matplotlib figure.
    '''
    # check flag for no empty
    no_empty = kwargs.get('no_empty',False)
    # load metadata file
    metadata = stackCPSheets(mpath)
    # filter to target diameter
    if (target_d != None) and (target_d != 'all'):
        # check that the target is supported
        if not (target_d in metadata['Diameter'].unique()):
            raise ValueError(f"Invalid target diameter {target_d}! Only valid values {metadata['Diameter'].unique()}")
        metadata = metadata.loc[metadata['Diameter'] == target_d]
    # if the target xaxis is a string specifying a particular column
    if isinstance(xaxis,str):
        if not (xaxis in metadata.columns.tolist()):
            # some column names have an added space on the end for some reason
            if not (xaxis + ' ' in metadata.columns.tolist()):
                raise ValueError(f"Invalid target column {xaxis} for x-axis! Only valid values are {metadata.columns.tolist()}")
            else:
                xaxis += ' '
    # if the target yaxis is a string check against supported list
    proc = None
    if isinstance(yaxis, str):
        if not (yaxis in ['Max Torque','Min Torque','Max Thrust','Min Thrust']):
            raise ValueError(f"Invalid target column for y-axis! Only valid values are {metadata.columns}")
    else:
        proc = yaxis
        yaxis = kwargs.get('ylabel',yaxis.__name__)
    # get target file hole counts
    fns = []
    xdata = []
    # extract y data for plotting
    ydata = []
    for ii,uq in metadata['Setitec hole count'].items():
        # if the hole count is None for some reason
        # skip it informing user
        if (uq is None) or (math.isnan(uq)):
            print(f"Skipping hole number {metadata['Test Number '].iloc[ii]}")
            continue
        # search for target file that has the target count
        found = list(filter(lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[-1] == str(int(uq)), glob(data_path)))
        # if a matching file was found append the file path and corresponding target xaxis value to lists
        if found:
             # load data
            data = dp.loadSetitecXls(found[0],'auto_data')
            # if it's not a string assume it's a callable        
            if not (proc is None):
                val = proc(data)
            else:
                # extract target y value
                if yaxis == 'Max Torque':
                    val = data['I Torque (A)'].values.max()
                    # if inc empty in calculation
                    if not no_empty:
                        try:
                            val += data['I Torque Empty (A)'].values.max()
                        except KeyError:
                            pass
                elif yaxis == 'Min Torque':
                    val = data['I Torque (A)'].values.min()
                    # if inc empty in calculation
                    if not no_empty:
                        try:
                            val += data['I Torque Empty (A)'].values.min()
                        except KeyError:
                            pass
                elif yaxis == 'Max Thrust':
                    val = data['I Thrust (A)'].values.max()
                    # if inc empty in calculation
                    if not no_empty:
                        try:
                            val += data['I Thrust Empty (A)'].values.max()
                        except KeyError:
                            pass
                elif yaxis == 'Min Thrust':
                    val = data['I Thrust (A)'].values.min()
                    # if inc empty in calculation
                    if not no_empty:
                        try:
                            val += data['I Thrust Empty (A)'].values.min()
                        except KeyError:
                            pass
            ydata.append(val)
            xdata.append(metadata[xaxis].iloc[ii])
        else:
            print(f"Cannot find file with hole count {str(int(uq))}!")
    # if the target column for xaxis is known to be not a number
    if not (xaxis in ['Test Phase ','Test Number ','Setitec Hole Count ']):
        # convert data to be category data
        xlabels = pd.Series(xdata,dtype='category')
        # get unique codes for categories to be used as the x values
        xdata = xlabels.cat.codes.tolist()
        # get the values as a list
        xlabels = xlabels.astype("string").tolist()
    else:
        xdata = [int(x) for x in xdata]
        xlabels = [str(x) for x in xdata]   
    # make axis
    f,ax = plt.subplots(constrained_layout=True)
    # plot data
    ax.plot(xdata,ydata,'x-')
    # set axis labels and adjust limits
    ax.set(xlabel=xaxis,ylabel=yaxis)
    if not kwargs.get('no_lim',True):
        ax.set_xlim(min(xdata),max(xdata))
    # change the xtick labels and set rotation
    ax.set_xticks(xdata,xlabels,rotation=kwargs.get("label_ang",45))
    # set figure title
    f.suptitle(kwargs.get("ftitle",f"Counterspark {xaxis} vs {yaxis}"))
    return f

def plotToolLife(data_path,mpath,target_d = '4B',**kwargs):
    '''
        Plot the Max and Min Torque against interpolated Tool Life

        Tool Life is originally categorical with the values 0 - 33%, 33 - 66% and 66 - 100%.
        This function generates a range of values from 0-100 so the torque data can be plotted
        against an increasing value rather than categorical.

        By default it uses np.linspace to generate the values. The user can also supply a function
        using the wear_data keyword. The function must accept the number of torque values and return
        a length of tool wear values of the same length

        Inputs:
            data_path : Wildcard path to data files.
            mpath : Target path of metadata file
            target_d : Target diameter. Can be 4B, 8B, all or None. If all or None, then all values are used.
                        Default 4B.
    '''
    # check flag for no empty
    no_empty = kwargs.get('no_empty',False)
    # load metadata file
    metadata = stackCPSheets(mpath)
    # filter to target diameter
    if (target_d != None) and (target_d != 'all'):
        # check that the target is supported
        if not (target_d in metadata['Diameter'].unique()):
            raise ValueError(f"Invalid target diameter {target_d}! Only valid values {metadata['Diameter'].unique()}")
        metadata = metadata.loc[metadata['Diameter'] == target_d]

    # get target file hole counts
    targets = metadata['Setitec hole count']
    ydata = []
    ymin = []
    for ii,uq in metadata['Setitec hole count'].items():
        # if the hole count is None for some reason
        # skip it informing user
        if (uq is None) or (math.isnan(uq)):
            print(f"Skipping hole number {metadata['Test Number '].iloc[ii]}")
            continue
        # search for target file that has the target count
        found = list(filter(lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[-1] == str(int(uq)), glob(data_path)))
        # if a matching file was found append the file path and corresponding target xaxis value to lists
        if found:
            # load data
            data = dp.loadSetitecXls(found[0],'auto_data')
            val = data['I Torque (A)'].values
            # if inc empty in calculation
            if not no_empty:
                try:
                    val += data['I Torque Empty (A)'].values
                except KeyError:
                    pass
            ydata.append(val.max())
            ymin.append(val.min())
    # generate tool life data
    if kwargs.get('wear_data',None):
        twear = kwargs.get('wear_data')(len(ydata))
    else:
        twear = np.linspace(0,100,len(ydata)-1)
    f,ax = plt.subplots(ncols=2,constrained_layout=True)
    # plot data
    ax[0].plot(twear,ydata,'x-')
    ax[0].set(xlabel="Tool Wear (%)",ylabel="Max Torque (A)",title="Max Torque vs Tool Wear")

    ax[1].plot(twear,ymin,'x-')
    ax[1].set(xlabel="Tool Wear (%)",ylabel="Min Torque (A)",title="Min Torque vs Tool Wear")
    return f

def plotCounterspark3D(data_path,mpath,yaxis,xaxis='Setitec hole count',zaxis='Max Torque',target_d='4B',**kwargs):
    '''
        Plot a target datasource against the file name and max torque

        yaxis is the metadata column key used as the y-axis. It is plotted against
        the file name on the x-axis and max torque on the z-axis

        Inputs:
            data_path : Wildcard path to data files
            mpath : Path to metadata file
            yaxis : Metadata column key used as y data
            target_d : Target diameter. Default 8B
            no_empty : Flag to not use empty channel in max torque calculation. Default False
            ftitle : Figure title. Default f"File Name vs {yaxis} vs Max Torque"

        Returns figure object
    '''
    # check flag for no empty
    no_empty = kwargs.get('no_empty',False)
    # load metadata file
    metadata = stackCPSheets(mpath)
    # filter to target diameter
    if (target_d != None) and (target_d != 'all'):
        # check that the target is supported
        if not (target_d in metadata['Diameter'].unique()):
            raise ValueError(f"Invalid target diameter {target_d}! Only valid values {metadata['Diameter'].unique()}")
        metadata = metadata.loc[metadata['Diameter'] == target_d]
    # get target hole counts for files
    fns = []
    # get target y data
    ysource = metadata[yaxis].astype("category")
    ydata = []
    # generate lists for xdata and xlabels
    if xaxis != 'File Index':
        xsource = metadata[xaxis].astype("category")
    xdata = []
    xlabels = []
    # start list for zdata
    # used as labels too
    zdata = []
    # iterate over target hole counts
    for ii,uq in metadata['Setitec hole count'].items():
        # if the hole count is None for some reason
        # skip it informing user
        if (uq is None) or (math.isnan(uq)):
            print(f"Skipping hole number {metadata['Test Number '].iloc[ii]}")
            continue
        # search for target file that has the target count
        found = list(filter(lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[-1] == str(int(uq)), glob(data_path)))
        if not found:
            print(f"Skipping hole number {metadata['Test Number '].iloc[ii]}")
            continue
        # if a matching file was found append the file path and corresponding target xaxis value to lists
        if found:
            fns.append(found[0])
            ydata.append(ysource.cat.codes.iloc[ii])
            # load data
            data = dp.loadSetitecXls(found[0],'auto_data')
            if zaxis == 'Max Torque':
                val = data['I Torque (A)'].values
                # if inc empty in calculation
                if not no_empty:
                    try:
                        val += data['I Torque Empty (A)'].values
                    except KeyError:
                        pass
                zdata.append(val.max())
            else:
                zdata.append(zaxis(data))
            # if the target x-axis is file index
            if xaxis == 'File Index':
                xdata.append(fi)
                #xlabels.append(os.path.splitext(os.path.basename(fn))[0])
                xlabels.append(targets.iloc[ii])
            # if it's something else
            else:
                xlabels.append(xsource.iloc[ii])
                xdata.append(xsource.cat.codes.iloc[ii])     
    # make figure and axis
    f = plt.figure()
    ax = f.add_subplot(projection='3d')
    # plot scatter data
    ax.scatter(xdata,ydata,zdata)
    # set labels and xlimits
    ax.set(xlabel=kwargs.get("xlabel",xaxis),ylabel=kwargs.get("ylabel",yaxis),zlabel=zaxis.__name__ if callable(zaxis) else zaxis,xlim=[min(xdata),max(xdata)])
    # if the source of the ydata is categorical data
    # set the y ticks to the unique cat values and category labels
    if ysource.dtype.name == 'category':
        ax.set_yticks(ysource.cat.codes.unique(),ysource.unique().to_list())
    # set the x labels and values
    ax.set_xticks(xdata,xlabels)
    #ax.set_xticks(xdata,xlabels,rotation=kwargs.get("label_ang",45))
    f.suptitle(kwargs.get("ftitle",f"{ax.get_xlabel()} vs {ax.get_ylabel()} vs {ax.get_zlabel()}"))
    return f

def onlyVirgin(metadata):
    ''' Function to filter the metadata to only where Condition of Supply is Virgin '''
    return metadata[metadata['Condition of Supply']=='virgin']

def onlyPreHole(metadata):
    ''' Function to filter the metadata to only where Condition of Supply is Pre Hole '''
    return metadata[metadata['Condition of Supply']=='pre hole']

def makeInteractionPlot(data_path,mpath,xaxis='Tool Life',trace='Material Config',target_d = '4B',**kwargs):
    '''
        Plot an interaction plot between two variables using Max Torque as the response

        Creates an interaction plot between xaxis variable and trace variable using Max Torque as the response.
        The result is a plot to show how the mean Max Torque changes given combinations of xaxis and trace.
        The goal is to see if there's a relationship between xaxis and trace using Max Torque.

        If the line is flat, then there's no relation. If it's not flat then there's a relationship between
        the two variables.

        The input add_filter is a callable to take the stacked metadata DataFrame and filter it down further.
        The idea is to filter it down to where a column is a specific value. See onlyVirgin as an example

        Inputs:
            xaxis : Counterspark column used as the x-axis. Default Tool Life
            trace : Counterspark column used for each trace. Default Material Config
            target_d : Target diameter to use. Can be 4B, 8B, all or None. If all or None,
                       then all diameters are used. Default 4B.
            no_empty : Flag to not used Torque Empty in max calculatioin. Default False
            add_filter : Additional callable applied to the metadata dataframe to filter it down further.
            fix_xticks : Flag to set the xticks to the values of the xaxis. Sometimes better
                        over Matplotlib tick values. Doesn't always help. Default False.

        Returns the figure object
    '''
    from statsmodels.graphics.factorplots import interaction_plot
    import math
    # check flag for no empty
    no_empty = kwargs.get('no_empty',False)
    # load metadata file
    metadata = stackCPSheets(mpath)
    # filter to target diameter
    if (target_d != None) and (target_d != 'all'):
        # check that the target is supported
        if not (target_d in metadata['Diameter'].unique()):
            raise ValueError(f"Invalid target diameter {target_d}! Only valid values {metadata['Diameter'].unique()}")
        metadata = metadata.loc[metadata['Diameter'] == target_d]
    # apply additional filter to metadata dataframe
    filter_string = ''
    if kwargs.get('add_filter',None):
        metadata = kwargs.get('add_filter')(metadata)
        filter_string = f", Filtered using {kwargs.get('add_filter').__name__}"
    # get target filenames
    xdata = []
    trace_data = []
    ydata = []
    # iterate over hole counts
    for ii,uq in metadata['Setitec hole count'].items():
        # if the hole count is None for some reason
        # skip it informing user
        if (uq is None) or (math.isnan(uq)):
            print(f"Skipping Setitec hole count {uq}, idx {ii}")
            continue
        # search for target file that has the target count
        # don't need to check for air files
        found = list(filter(lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[-1] == str(int(uq)), glob(data_path)))
        if not found:
            print(f"Skipping Setitec hole count {uq}, idx {ii}")
            continue
        # if a matching file was found append the file path and corresponding target xaxis value to lists
        if found:
            # load data
            data = dp.loadSetitecXls(found[0],'auto_data')
            # get initial max
            val = data['I Torque (A)'].values.max()
            # if inc empty channnel in calculation
            if not no_empty:
                try:
                    val += data['I Torque Empty (A)'].values.max()
                except KeyError:
                    pass
            ydata.append(val)
            xdata.append(metadata[xaxis][ii])
            trace_data.append(metadata[trace][ii])
    # make figure
    f,ax = plt.subplots(constrained_layout=True,figsize=(9,9))
    #f = interaction_plot(x=metadata['Through Tool Air (Bar)'],trace=metadata[trace],response=np.array(ydata),ax=ax)
    f = interaction_plot(x=pd.Series(xdata),trace=pd.Series(trace_data),response=np.array(ydata),ax=ax)
    # fix the xticks to specific positions rather than using matplotlibs auto generated ones
    # can help with making the figure easier to read. Not always
    if kwargs.get('fix_xticks',False):
        ax.set_xticks(metadata[xaxis].values)
    # set labels and title
    ax.set(ylabel="Mean Max Torque (A)",title=f"Interaction Plot Between {xaxis} and {trace} using Max Torque (A)\nDiameter {target_d if target_d else 'all'}"+filter_string)
    return f

def makeAllInteractionPlots(opath):
    from itertools import combinations
    cols = ['Material Config','Condition of Supply','Drilling Orientation','Tool Life','Through Tool Air (Bar)','Vacuum Extraction','Vacuum Extraction Dwell','Machine Surface Contact','Exit pre strip']
    for d,p in zip(['4B','8B'],[r"COUNTERSPARK/data/4B/*.xls",r"COUNTERSPARK/data/8B/*.xls"]):
        print(f"Processing {d}")
        for x,y in combinations(cols,2):
            f = makeInteractionPlot(p,"COUNTERSPARK/Counterspark drilling test.xlsx",target_d=d,xaxis=x,trace=y)
            f.savefig(os.path.join(opath,f"{d}-{x}-{y}-interaction-plot-max-torque.png"))
            plt.close(f)
            # filter to virgin and pre holes creating different plots
            # if a column only has a single value as a result of filtering, interaction_plot throws an error so this is to prevent that
            if (x!='Condition of Supply') and (y!='Condition of Supply'):
                for sup,fun in zip(['virgin','pre hole'],[onlyVirgin,onlyPreHole]):
                    f = makeInteractionPlot(p,"COUNTERSPARK/Counterspark drilling test.xlsx",target_d=d,xaxis=x,trace=y,add_filter=fun)
                    f.savefig(os.path.join(opath,f"{d}-{x}-{y}-interaction-plot-max-torque-supply-{sup}.png"))
                    plt.close(f)

def makeEdgeUpdate(k):
    u,v = k
    return (u,v,{"weight":1./dtw(dp.loadSetitecXls(u,"auto_data")["I Torque (A)"].values,dp.loadSetitecXls(v,"auto_data")["I Torque (A)"].values)})

def plotDTWGraph(data_path,draw=False):
    '''
        Convert series of data files into a undirected complete graph where each edge weight is DTW similarity

        Inputs:
            data_path : Wildcard path to data files
            draw : Flag to draw the graph using matplotlib method

        Returns graph
    '''
    import networkx as nx
    from multiprocessing  import Pool
    # make a complete graph with blank edges
    # node names are file paths
    G = nx.complete_graph(glob(data_path))
    #G.update(edges=[(u,v,{"weight":dtw(dp.loadSetitecXls(u,"auto_data")["I Torque (A)"].values,dp.loadSetitecXls(v,"auto_data")["I Torque (A)"].values)}) for u,v in G.edges().keys()])
    G.update(Pool(6).map(makeEdgeUpdate,G.edges().keys()))
    if draw:
        ax = plt.subplot(111)
        #pos = nx.random_layout(G)
        pos = nx.spectral_layout(G)
        nx.draw(G,pos=pos,ax=ax,with_labels=False)
        # format labels to 2 dp
        labels = {(u, v): f'{d["weight"]:.2f}' for u, v, d in G.edges(data=True)}
        nx.draw_networkx_edge_labels(G,pos,edge_labels=labels,ax=ax)
    return G

def integrateSignal(data_path):
    '''
        Integrate each found signal and plot the area for each file on 2D axis

        Creates a 2D axis of hole number against integrated area.

        scipy.integrate.trapz is used as simpson returned a series of NaNs in testing

        Inputs:
            data_path : Wildcard path for all data files

        Returns list of integrated area
    '''
    from scipy.integrate import trapz
    area = []
    for fn in glob(data_path):
        data = dp.loadSetitecXls(fn,'auto_data')
        area.append(trapz(data['I Torque (A)'].values,np.abs(data['Position (mm)'].values)))
    return area

def integMats(mpath,data_path,plot=False,norm=False):
    '''
        For each material configuration integrate each found signal and plot the area for each file on 2D axis#

        scipy.integrate.trapz is used as simpson returned a series of NaNs in testing

        Each material config is plotted as a separate trace

        Each diameter has it's own axis

        Inputs:
            mpath : Path to metadata file
            data_path : Wildcard path for all data files referenced in metadata file

        Returns a dictionary of signal area organised by material and then by diameter
    '''
    from scipy.integrate import trapz
    metadata = stackCPSheets(mpath)
    area = {}
    for m in metadata['Material Config'].unique():
        config = metadata[metadata['Material Config'] == m]
        area[m] = {}
        for d in config['Diameter'].unique():
            config_d = config[config['Diameter'] == d]
            print(f"{m} {d} {config_d.shape}")
            area[m][d] = ([],[])
            for ii,uq in config_d['Setitec hole count'].items():
                # if the hole count is None for some reason
                # skip it informing user
                if (uq is None) or (math.isnan(uq)):
                    print(f"Skipping Setitec hole count {uq}, idx {ii}")
                    continue
                # search for target file that has the target count
                # don't need to check for air files
                found = list(filter(lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[-1] == str(int(uq)), glob(data_path)))
                if not found:
                    print(f"Skipping Setitec hole count {uq}, idx {ii}")
                    continue
                # if a matching file was found append the file path and corresponding target xaxis value to lists
                if found:
                    # load data
                    data = dp.loadSetitecXls(found[0],'auto_data')
                    area[m][d][0].append(ii)
                    tq = data['I Torque (A)'].values
                    if norm:
                        tq = (tq-tq.min())/(tq.max()-tq.min())
                    area[m][d][1].append(trapz(tq,np.abs(data['Position (mm)'].values)))
    if plot:
        f,ax = plt.subplots(ncols=2)
        for k,data in area.items():
            for i,(d,(x,v)) in enumerate(data.items()):
                ax[i].plot(x,v,'-x',label=k)
                ax[i].legend()
                ax[i].set(xlabel='Hole Number',ylabel='Integrated Area',title=f"Diameter {d}")
        f.suptitle(f"Integrated Area Norm={norm}")
    return area
                
def signalEnergyMats(mpath,data_path,plot=False):
    '''
        For each material configuration integrate each found signal and plot the energy for each file on 2D axis

        scipy.integrate.trapz is used as simpson returned a series of NaNs in testing

        Each material config is plotted as a separate trace

        Each diameter has it's own axis

        Inputs:
            mpath : Path to metadata file
            data_path : Wildcard path for all data files referenced in metadata file
            plot : Flag to plot the results. Default False.

        Returns a dictionary of signal energy organised by material and then by diameter
    '''
    metadata = stackCPSheets(mpath)
    energy = {}
    for m in metadata['Material Config'].unique():
        config = metadata[metadata['Material Config'] == m]
        energy[m] = {}
        for d in config['Diameter'].unique():
            config_d = config[config['Diameter'] == d]
            print(f"{m} {d} {config_d.shape}")
            energy[m][d] = ([],[])
            for ii,uq in enumerate(config_d['Setitec hole count']):
                # if the hole count is None for some reason
                # skip it informing user
                if (uq is None) or (math.isnan(uq)):
                    print(f"Skipping Setitec hole count {uq}, idx {ii}")
                    continue
                # search for target file that has the target count
                # don't need to check for air files
                found = list(filter(lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[-1] == str(int(uq)), glob(data_path)))
                if not found:
                    print(f"Skipping Setitec hole count {uq}, idx {ii}")
                    continue
                # if a matching file was found append the file path and corresponding target xaxis value to lists
                if found:
                    # load data
                    data = dp.loadSetitecXls(found[0],'auto_data')
                    energy[m][d][0].append(ii)
                    energy[m][d][1].append(np.sum(data['I Torque (A)'].values**2))
    if plot:
        f,ax = plt.subplots(ncols=2)
        for k,data in energy.items():
            for i,(d,(x,v)) in enumerate(data.items()):
                ax[i].plot(x,v,'-x',label=k)
                ax[i].legend()
                ax[i].set(xlabel='Hole Number',ylabel='Signal Energy',title=f"Diameter {d}")
    return energy

def signalEnergyMatsAgainst3D(mpath,data_path,zaxis,plot=False,swap_axis=True):
    '''
        For each material configuration integrate each found signal and plot the energy for each file against a
        3rd axis

        scipy.integrate.trapz is used as simpson returned a series of NaNs in testing

        Each material config is plotted as a separate trace

        Each diameter has it's own plotting axis

        The 3rd axis must be a column from metadata file e.g. Vacuum Extraction

        Inputs:
            mpath : Path to metadata file
            data_path : Wildcard path for all data files referenced in metadata file
            zaxis : Additional axis to plot against.
            plot : Flag to plot the results
            swap_axis : Swap the zaxis and y-axis. Default True.

        Returns a dictionary of signal energy organised by material and then by diameter
    '''
    import math
    metadata = stackCPSheets(mpath)
    zdata = metadata[zaxis].astype("category").cat.codes
    energy = {}
    for m in metadata['Material Config'].unique():
        config = metadata[metadata['Material Config'] == m]
        energy[m] = {}
        for d in config['Diameter'].unique():
            config_d = config[config['Diameter'] == d]
            print(f"{m} {d} {config_d.shape}")
            energy[m][d] = ([],[],[]) 
            for ii,uq in config_d['Setitec hole count'].items():
                # if the hole count is None for some reason
                # skip it informing user
                if (uq is None) or (math.isnan(uq)):
                    print(f"Skipping Setitec hole count {uq}, idx {ii}")
                    continue
                # search for target file that has the target count
                # don't need to check for air files
                found = list(filter(lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[-1] == str(int(uq)), glob(data_path)))
                if not found:
                    print(f"Skipping Setitec hole count {uq}, idx {ii}")
                    continue
                # if a matching file was found append the file path and corresponding target xaxis value to lists
                if found:
                    # load data
                    data = dp.loadSetitecXls(found[0],'auto_data')
                    energy[m][d][0].append(ii)
                    energy[m][d][1].append(np.sum(data['I Torque (A)'].values**2))
                    #energy[m][d][2].append()
                    cat = zdata[config_d[config_d['Setitec hole count']==uq].index.values[0]]
                    #print(f"{uq} -> {metadata[zaxis][idx]} -> {idx}")
                    #input()
                    energy[m][d][2].append(cat)
    if plot:
        f = plt.figure()
        ax = [f.add_subplot(121,projection='3d'),f.add_subplot(122,projection='3d')]
        for mat,data in energy.items():
            for i,(d,(x,v,z)) in enumerate(data.items()):
                if swap_axis:
                    ax[i].plot(x,z,v,'-x',label=mat)
                    ax[i].set(xlabel='Hole Number',zlabel='Signal Energy',ylabel=zaxis,title=f"Diameter {d}")
                else:
                    ax[i].plot(x,v,z,'-x',label=mat)
                    ax[i].set(xlabel='Hole Number',ylabel='Signal Energy',zlabel=zaxis,title=f"Diameter {d}")
                ax[i].legend()
        if swap_axis:
            ax[0].set_yticks(zdata.unique(),metadata[zaxis].astype("string").unique())
            ax[1].set_yticks(zdata.unique(),metadata[zaxis].astype("string").unique())
        else:
            ax[0].set_zticks(zdata.unique(),metadata[zaxis].astype("string").unique())
            ax[1].set_zticks(zdata.unique(),metadata[zaxis].astype("string").unique())
    return energy

def makeTrainDataset(mpath,data_path,axis,split=0.5,**kwargs):
    '''
        Create the training and testing datasets

        The metadata is loaded and the Setitec hole counts are used to identify and find the
        data files in data_path.

        Axis is the column name used as the output label data. The column is cast to a category data type
        and the codes used as the output labels.

        The data features are currently as follows:
            - Max Torque
            - Avg Torque
            - Signal Energy (from torque)
            - Integrated Signal Area (from torque)

        When inc_labels is True, then other labels are used as inputs in the feature set. Currently used labels are
        Condition of Supply, Drilling Orientation, Tool Life, Through Tool Air (Bar), Vacuum Extraction, Vacuum Extraction Dwell, Machine Surface Contact and Exit pre strip
        
        The list of feature labels returned can be used as tick labels or as a unique identifier for each label

        If split is 1.0, then only the full set, full labels and feature labels are returned
    
        Inputs:
            mpath : Path to metadata file
            data_path : Wildcard path to data files.
            axis : Target axis in metadata file to predict
            split : Single or iterable collection of train/test split ratio. Float between 0 and 1.
            inc_labels : Include other labels as input features. Default True.

        Returns X_train, X_test, y_train, y_test and feature labels
    '''
    from scipy.integrate import trapz
    from sklearn.model_selection import train_test_split
    from sklearn.preprocessing import StandardScaler, LabelEncoder
    import math
    # list of supported columns
    cols = sorted(['Condition of Supply','Drilling Orientation','Tool Life','Through Tool Air (Bar)','Vacuum Extraction','Vacuum Extraction Dwell','Machine Surface Contact','Exit pre strip'])
    # load metadata
    metadata = stackCPSheets(mpath)
    # get output label
    ldata = metadata[axis].astype("category").cat.codes
    output_labels = [str(v) for v in metadata[axis].values.tolist()]
    # setup lists    
    X = []
    X_labels = []
    label = []
    feat_labels = ['Max Torque','Avg Torque','Signal Energy','Integ Signal']
    # get other labels
    if kwargs.get('inc_labels',False):
        if axis in cols:
            cols.remove(axis)
        feat_labels += cols
        other_labels = metadata[cols]
        # label enode the other labels
        for c in cols:
            other_labels[c] = LabelEncoder().fit_transform(other_labels[c].values)
            
    # collect data features
    for i,uq in enumerate(metadata['Setitec hole count'].values):
        if (uq is None) or (math.isnan(uq)):
            if not kwargs.get('quiet',True):
                print(f"Skipping Setitec hole count {uq}")
            continue
        # search for target file that has the target count
        # don't need to check for air files
        found = list(filter(lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[-1] == str(int(uq)), glob(data_path)))
        if not found:
            if not kwargs.get('quiet',True):
                print(f"Skipping Setitec hole count {uq}")
            continue
        # if a matching file was found append the file path and corresponding target xaxis value to lists
        if found:
            # load data
            data = dp.loadSetitecXls(found[0],'auto_data')
            pos = np.abs(data['Position (mm)'].values)
            torque = data['I Torque (A)'].values
            # create features
            if kwargs.get('inc_labels',False):
                X.append([torque.max(),torque.mean(),np.sum(torque**2),trapz(torque,pos)])
                X_labels.append(other_labels.iloc[i].values.tolist())
            else:
                X.append([torque.max(),torque.mean(),np.sum(torque**2),trapz(torque,pos)])
            # get output label value
            label.append(ldata[metadata[metadata['Setitec hole count']==uq].index[0]])
    # form into array    
    X = np.row_stack(X)
    # norm per column
    for i in range(X.shape[1]):
        norm = StandardScaler().fit_transform(X[:,i].reshape(-1,1))
        X[:,i] = norm.flatten()
    # encode the extra labels
    if kwargs.get('inc_labels',False):
        X_labels = np.row_stack(X_labels)
        for i in range(X_labels.shape[1]):
            X_labels[:,i] = LabelEncoder().fit_transform(X_labels[:,i])
        # add extra lebels
        X = np.column_stack((X,X_labels))
    # encode all labels
    label = LabelEncoder().fit_transform(label)
    # get split data and features
    return train_test_split(X,label,test_size=1.0-split,random_state=0),feat_labels,output_labels

def getAllFeatures(mpath,data_path):
    '''
        Create the full feature set without splitting

        The metadata is loaded and the Setitec hole counts are used to identify and find the
        data files in data_path.

        The data features are currently as follows:
            - Max Torque
            - Avg Torque
            - Signal Energy (from torque)
            - Integrated Signal Area (from torque)

        The list of feature labels returned can be used as tick labels or as a unique identifier for each label

        Inputs:
            mpath : Path to metadata file
            data_path : Wildcard path to data files

        Returns full feature set, fitting labels and feature names
    '''
    import math
    from scipy.integrate import trapz
    # load metadata
    metadata = stackCPSheets(mpath)
    # get output label
    X = []
    feat_labels = ['Max Torque','Avg Torque','Signal Energy','Integ Signal']
    # collect data features
    for uq in metadata['Setitec hole count'].values:
        if (uq is None) or (math.isnan(uq)):
            print(f"Skipping Setitec hole count {uq}")
            continue
        # search for target file that has the target count
        # don't need to check for air files
        found = list(filter(lambda x : os.path.splitext(os.path.basename(x))[0].split('_')[-1] == str(int(uq)), glob(data_path)))
        if not found:
            print(f"Skipping Setitec hole count {uq}")
            continue
        # if a matching file was found append the file path and corresponding target xaxis value to lists
        if found:
            # load data
            data = dp.loadSetitecXls(found[0],'auto_data')
            pos = np.abs(data['Position (mm)'].values)
            torque = data['I Torque (A)'].values
            # create features
            X.append([torque.min(),torque.max(),torque.mean(),np.sum(torque**2),trapz(torque,pos)])
    # form into array    
    X = np.row_stack(X)
    return X,['Min Torque','Max Torque','Avg Torque','Signal Energy','Integ Signal']

def plotFeatures(mpath,data_path):
    '''
        Build feature set using getAllFeatures and plot them in separate graphs

        Inputs:
            mpath : Path to metadata file
            data_path : Wildcard path to data files

        Returns created matplotlib figure object
    '''
    X,labels = getAllFeatures(mpath,data_path)
    k = np.sqrt(len(labels))
    f = plt.figure(constrained_layout=True)
    r = int(np.floor(k))
    c = int(np.ceil(k))
    for i,l in zip(range(len(labels)),labels):
        aa = f.add_subplot(r,c,i+1)
        aa.plot(X[:,i],'x')
        aa.set(xlabel="Hole Number",ylabel=l,title=l)
    return f

# from https://stackoverflow.com/a/61037626
def plotRandomTree(tree,feature_names,class_names):
    '''
        Iterate over a RandomTreeClassifier and plot the individual classifier trees

        RandomTreeClassifiers are composed of several classifiers fitted to sub samples of the dataset.
        This function iterates over those classifiers and plots them using sklearn.tree.plot_tree in a square
        matrix of subplots

        WARNING: This creates a giant matrix and is best saved and inspected afterwards rather than viewed

        Inputs:
            tree : RandomForestTree classifier
            feature_names : Iterable list of feature names
            class_names : Iterable list of output class names

        Returns matplotlib
    '''
    from sklearn.tree import plot_tree
    # calculate the number of rows and columns needed
    nq = np.sqrt(len(tree.estimators_))
    nr = int(np.ceil(nq))
    nc = int(np.floor(nq))
    # construct figure
    fig, axes = plt.subplots(nrows = nr,ncols = nc,figsize = (18,10), dpi=900,constrained_layout=True)
    axes = axes.flatten()
    # iterate over estimators
    for index in range(len(tree.estimators_)):
        plot_tree(tree.estimators_[index],
                       feature_names = feature_names, 
                       class_names=class_names,
                       filled = True,
                       ax = axes[index],
                      rounded = False,
                      fontsize=None)
    return fig

def attemptMultiPredict(mpath,data_path,axis,opath,split=0.5,inc_labels_as_feats=True, draw_classifiers=False):
    '''
        Train a set of classifiers to predict the target axis using a set of data features

        Split is the ratio of data used in the training set when sepaarating the data into training and testing dataset.
        Passed to makeTrainDataset

        Several plots are generated and saved
            - Feature importance according to MIDI
            - Feature importance accorcding to full permutation
            - Training accuracy & testing accuracy for each split ratio
            - Average training accuracy and testing accuracy over split ratio

        Input:
            mpath : Path to metadata file
            data_path : Wildcard path to data files.
            axis : Target column to predict
            opath : Output path to save figures in
            split : Single or iterable collection of train/test split ratio

        Returns dictionary of training data, testing data and all testing results
        organised by split value.
    '''
    from sklearn import svm
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.neighbors import KNeighborsRegressor
    from sklearn.tree import DecisionTreeClassifier, plot_tree
    from sklearn.discriminant_analysis import LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis
    from sklearn.neural_network import MLPClassifier
    from sklearn.gaussian_process import GaussianProcessClassifier
    from sklearn.metrics import accuracy_score
    from sklearn.inspection import permutation_importance
    from sklearn.model_selection import train_test_split
    #from sklearn.preprocessing import OneHotEncoder
    from sklearn.preprocessing import LabelEncoder
    from sklearn.compose import make_column_transformer
    from sklearn.model_selection import train_test_split

    # create figures
    facc,axacc = plt.subplots(ncols=2,constrained_layout=True,figsize=(20,6))
    # ensure split is a list
    try:
        for _ in split:
            break
    except TypeError:
        split = [split,]

    all_test_acc = {sp : [] for sp in split}
    all_train_acc = {sp : [] for sp in split}
    all_res = {sp : [] for sp in split}
    classifiers = []
    all_clf_label = []
    # collect accuracy and results
    test_acc = []
    train_acc = []
    res = []
    print(axis)
    # for each split value
    for sp in split:
        (X_train, X_test,y_train,y_test),feature_names,output_labels = makeTrainDataset(mpath,data_path,axis,sp,inc_labels=inc_labels_as_feats)
        classifiers.clear()

        # train random forest classifier
        forest = RandomForestClassifier(random_state=0)
        forest.fit(X_train,y_train)
        classifiers.append(forest)
        all_clf_label.append('Random Forest')

        # train svm
        clf = svm.SVC(decision_function_shape='ovo',max_iter=int(1e5))
        clf.fit(X_train,y_train)
        classifiers.append(clf)
        all_clf_label.append('SVC')

        # train LinearSVC
        clf_lin = svm.LinearSVC(max_iter=int(1e5))
        clf_lin.fit(X_train,y_train)
        classifiers.append(clf_lin)
        all_clf_label.append('Linear SVC')

        # linear discriminator
        lindisc = LinearDiscriminantAnalysis()
        lindisc.fit(X_train,y_train)
        classifiers.append(lindisc)
        all_clf_label.append('Linear Disc.')

        # linear discriminator
        quaddisc = QuadraticDiscriminantAnalysis()
        quaddisc.fit(X_train,y_train)
        classifiers.append(quaddisc)
        all_clf_label.append('Quad Disc.')

        # MLP
        mlp = MLPClassifier(solver='lbfgs',max_iter=int(1e5))
        #mlp = MLPClassifier()
        mlp.fit(X_train,y_train)
        classifiers.append(mlp)
        all_clf_label.append('MLP')

        # Gaussian Process
        gp = GaussianProcessClassifier()
        gp.fit(X_train,y_train)
        classifiers.append(gp)
        all_clf_label.append('Gaussian Proc.')

        # Decision Tree Classifier
        dtree = DecisionTreeClassifier(random_state=0)
        dtree.fit(X_train,y_train)
        classifiers.append(dtree)
        all_clf_label.append("D-Tree")

        # collect accuracy and results
        test_acc.clear()
        train_acc.clear()
        res.clear()
        
        for clf,label in zip(classifiers,all_clf_label):
            res.append(clf.predict(X_test))
            test_acc.append(accuracy_score(y_test,res[-1]))
            train_acc.append(accuracy_score(y_train,clf.predict(X_train)))
            # add to global dictionary
            all_test_acc[sp].append(test_acc[-1])
            all_train_acc[sp].append(train_acc[-1])
            all_res[sp].append(res[-1])
            # if it's the random forest classifier
            # draw the structure
            if draw_classifiers:
                if isinstance(clf,RandomForestClassifier):
                    continue
                    fclf = plotRandomTree(clf,feature_names,[str(c) for c in np.unique(y_train)])
                    fclf.savefig(os.path.join(opath,f"{label}-random-forest-drawing-using-perm-split-{sp:.2f}-axis-{axis}.png"))
                    plt.close(fclf)
                elif isinstance(clf,MLPClassifier):
                    vmin, vmax = clf.coefs_[0].min(), clf.coefs_[0].max()
                    nrows = int(np.ceil(np.sqrt(clf.coefs_[0].shape[0])))
                    ncols = int(np.floor(np.sqrt(clf.coefs_[0].shape[0])))
                    fclf,axclf = plt.subplots(nrows,ncols,constrained_layout=True)
                    for coef,ax in zip(clf.coefs_[0],axclf.ravel()):
                        ax.matshow(coef.reshape(10,10),cmap=plt.cm.gray)
                        ax.set_xticks(())
                        ax.set_yticks(())
                    fclf.suptitle("MLP Classifier Features")
                    fclf.savefig(os.path.join(opath,f"{label}-mlpclassifier-drawing-using-perm-split-{sp:.2f}-axis-{axis}.png"))
                    plt.close(fclf)
                elif isinstance(clf,DecisionTreeClassifier):
                    fclf,axclf = plt.subplots(constrained_layout=True,figsize=(35,10))
                    plot_tree(clf,feature_names=feature_names,class_names=list(set(output_labels)),filled=True,ax=axclf,fontsize=8)
                    fclf.savefig(os.path.join(opath,f"{label}-decision-tree-drawing-using-perm-split-{sp:.2f}-axis-{axis}.png"))
                    plt.close(fclf)
            
            # plotting testing feaure importance
            result = permutation_importance(clf, X_test, y_test, n_repeats=10, random_state=42, n_jobs=4)
            clf_importances = pd.Series(result.importances_mean, index=feature_names)
            fc, axc = plt.subplots()
            clf_importances.plot.bar(yerr=result.importances_std, ax=axc)
            axc.set_title(f"Testing Feature importances using permutation ({label})\n{axis}")
            axc.set_ylabel("Mean accuracy decrease")
            fc.tight_layout()
            fc.savefig(os.path.join(opath,f"{label}-feature-importances-using-perm-split-{sp:.2f}-axis-{axis}.png"))
            plt.close(fc)

##            # plotting training feature importance
##            # https://scikit-learn.org/stable/auto_examples/inspection/plot_permutation_importance_multicollinear.html#sphx-glr-auto-examples-inspection-plot-permutation-importance-multicollinear-py
##            # get coefficients
##            if type(clf) is RandomForestClassifier:
##                fimp = clf.feature_importances_
##            else:
##                try:
##                    fimp = clf.coef_
##                # cannot retrieve testing feature importance for non-linear kernels
##                except AttributeError:
##                    continue
##            print(label,"shape pre ravel ",fimp.shape)
##            # sometimes the feature importances come out as a 2d array (num classes x num features)
##            fimp = fimp.ravel()
##            # find feature importance in training data
##            result = permutation_importance(clf, X_train, y_train, n_repeats=10, random_state=42, n_jobs=4)
##            # sort so the bars are in descending order
##            perm_sorted_idx = result.importances_mean.argsort()
##            # sort feature importances
##            tree_importance_sorted_idx = np.argsort(fimp)
##            # yaxis coordinates for plotting
##            tree_indices = np.arange(0, len(fimp)) + 0.5
##            # make two axes to hold the value and the standard deviation
##            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 8))
##            print(fimp,type(fimp),tree_importance_sorted_idx,type(tree_importance_sorted_idx))
##            # plot bars indicating feature importance along the xaxis
##            ax1.barh(tree_indices, fimp[tree_importance_sorted_idx], height=0.7)
##            # set ticks and labels
##            ax1.set_yticks(tree_indices)
##            ax1.set_yticklabels(np.asarray(feature_names)[tree_importance_sorted_idx])
##            ax1.set_ylim((0, len(fimp)))
##            # make a box plot in the 2nd axis indicating std dev
##            ax2.boxplot(
##                result.importances[perm_sorted_idx].T,
##                vert=False,
##                labels=np.asarray(feature_names)[perm_sorted_idx],
##            )
##            # add labels
##            fig.suptitle(f"{axis} Training Feature Importance p={sp:.2f}")
##            fig.tight_layout()
##            fig.savefig(os.path.join(opath,f"{label}-training-feature-importances-using-perm-split-{sp:.2f}-axis-{axis}.png"))
##            plt.close(fig)

            # plotting testing feaure importance
            result = permutation_importance(clf, X_train, y_train, n_repeats=10, random_state=42, n_jobs=4)
            clf_importances = pd.Series(result.importances_mean, index=feature_names)
            fc, axc = plt.subplots()
            clf_importances.plot.bar(yerr=result.importances_std, ax=axc)
            axc.set_title(f"Training Feature importances using permutation ({label})\n{axis}")
            axc.set_ylabel("Mean accuracy decrease")
            fc.tight_layout()
            fc.savefig(os.path.join(opath,f"{label}-training-feature-importances-using-perm-split-{sp:.2f}-axis-{axis}.png"))
            plt.close(fc)

        # plot training and testing 
        axacc[0].plot(train_acc,'x',label=f"sp={sp:.2f}",markersize=10,markeredgewidth=5)
        axacc[1].plot(test_acc,'x',label=f"sp={sp:.2f}",markersize=10,markeredgewidth=5)

        # plot feature importance based on mean decrease in impurity
        importances = forest.feature_importances_
        std = np.std([tree.feature_importances_ for tree in forest.estimators_], axis=0)
        forest_importances = pd.Series(importances, index=feature_names)
        ff, axf = plt.subplots()
        forest_importances.plot.bar(yerr=std, ax=axf)
        axf.set_title("Feature importances using mean decrease in impurity")
        axf.set_ylabel("Mean decrease in impurity")
        ff.tight_layout()
        ff.savefig(os.path.join(opath,f"feature-importances-using-mdi-split-{sp:.2f}-axis-{axis}.png"))
        plt.close(ff)

        # Feature importance based on feature permutation
        result = permutation_importance(forest, X_test, y_test, n_repeats=10, random_state=42, n_jobs=4)
        forest_importances = pd.Series(result.importances_mean, index=feature_names)
        fi, axi = plt.subplots()
        forest_importances.plot.bar(yerr=result.importances_std, ax=axi)
        axi.set_title("Feature importances using permutation on full model")
        axi.set_ylabel("Mean accuracy decrease")
        fi.tight_layout()
        fi.savefig(os.path.join(opath,f"feature-importances-using-perm-split-{sp:.2f}-axis-{axis}.png"))
        plt.close(fi)

    fb,axb = plt.subplots(constrained_layout=True,figsize=(10,6))
    # re-arrange the data into lists
    train_acc_list = np.concatenate(list(all_train_acc.values()))
    test_acc_list = np.concatenate(list(all_test_acc.values()))
    nc = len(np.unique(all_clf_label))
    split_list = np.concatenate([nc*[k,] for k in all_test_acc.keys()])
    # combine into columns
    df_data = np.column_stack((all_clf_label,train_acc_list,test_acc_list,split_list))
    df = pd.DataFrame(df_data,columns=['Classifier','Training Accuracy','Testing Accuracy','Train/Test Split'])
    # force convert the accuracy columns to numeric
    # split can be left as object
    df["Training Accuracy"] = pd.to_numeric(df["Training Accuracy"])
    df["Testing Accuracy"] = pd.to_numeric(df["Testing Accuracy"])
    # create a classifier column of category numbers
    sns.barplot(data=df,x="Classifier",y="Testing Accuracy",hue="Train/Test Split",ax=axb)
    fb.suptitle(f"{axis} Testing Accuracy",fontsize=16)
    plt.xlabel('Classifier', fontsize=14)
    plt.ylabel('Testing Accuracy', fontsize=14)
    axb.tick_params(axis='both', which='major', labelsize=12)
    axb.tick_params(axis='both', which='minor', labelsize=10)
    fb.savefig(os.path.join(opath,f"counterspark-testing-barchart-predicting-classifier-axis-{axis}-acc.png"))
    plt.close(fb)

    # for training
    fb,axb = plt.subplots(constrained_layout=True,figsize=(10,6))
    sns.barplot(data=df,x="Classifier",y="Training Accuracy",hue="Train/Test Split",ax=axb)
    fb.suptitle(f"{axis} Testing Accuracy",fontsize=16)
    plt.xlabel('Classifier', fontsize=14)
    plt.ylabel('Testing Accuracy', fontsize=14)
    axb.tick_params(axis='both', which='major', labelsize=12)
    axb.tick_params(axis='both', which='minor', labelsize=10)
    fb.savefig(os.path.join(opath,f"counterspark-training-barchart-predicting-classifier-axis-{axis}-acc.png"))
    plt.close(fb)

    # format the tick labels to be a certain size
    for aa in axacc:
        aa.legend()
        aa.tick_params(axis='both', which='major', labelsize=12)
        aa.tick_params(axis='both', which='minor', labelsize=10)
    axacc[0].set_xticks(range(nc),all_clf_label[:nc],rotation=0)
    axacc[0].set(ylabel="Training Accuracy",title="Training")
    axacc[1].set(ylabel="Testing Accuracy",title="Testing")
    axacc[1].set_xticks(range(nc),all_clf_label[:nc],rotation=0)

    facc.suptitle(f"Predicting {axis}",fontsize=20)
    facc.savefig(os.path.join(opath,f"counterspark-training-predicting-classifier-axis-{axis}-acc.png"))

    # plot the average training and testing accuracy across all split levels
    # makes a simpler plot
    f,ax = plt.subplots(ncols=2,constrained_layout=True,figsize=(12,6))
    avg_acc = np.row_stack(list(all_train_acc.values())).mean(axis=0)
    ax[0].plot(avg_acc,'x',markersize=10,markeredgewidth=5)
    avg_acc = np.row_stack(list(all_test_acc.values())).mean(axis=0)
    ax[1].plot(avg_acc,'x',markersize=10,markeredgewidth=5)

    ax[0].set_xticks(range(nc),all_clf_label[:nc],rotation=0)
    ax[0].set(ylabel="Avg. Training Accuracy",title="Average Training")
    ax[1].set(ylabel="Avg. Testing Accuracy",title="Average Testing")
    ax[1].set_xticks(range(nc),all_clf_label[:nc],rotation=0)
    f.suptitle(f"Predicting {axis} (Average)",fontsize=20)
    f.savefig(os.path.join(opath,f"counterspark-training-predicting-classifier-axis-{axis}-average-acc.png"))
    plt.close('all')
    return (all_train_acc,all_test_acc),all_res

def formatProgram(prog,pad=5):
    '''
        Load and format program data into a rectangular matrix

        If prog is a path, then getProgramValues with key all is used to get data.
        If something else, then it's assumed to be the dictionary returned by getProgramValues

        The input pad adds rows so that the program is a target size. This is so that programs with different
        number of steps can be compared. The rows are all zeros.

        The material value is ignored as it's a category code that isn't guaranteed to be set. The other program values
        have a real-world value so are better in a distance metric

        Inputs:
            prog : Path string or result from getProgramValues
            pad : Number of zero-rows to add to make the matrix pad number of rows

        Returns a rectangular matrix of pad x 20 values
    '''
    if isinstance(prog,str):
        prog = dp.getProgramValues(prog,"all")
    # for each value pad to target
    for kk in prog.keys():
        prog[kk] += (pad-len(prog[kk]))*[0.0,]
    prog.pop('Step Nb')
    if 'Material' in prog:
        prog.pop('Material')
    # extract values and arrange into list
    # skipping the Step Nb value
    vals = [vv for vv in prog.values()]
    return np.column_stack(vals)

def computeProgDistance(pA,pB,pad=5,**kwargs):
    '''
        Compute the distances between two programs

        The program values are treated as samples in a multi-dimensional space.
        The distance is calculated as sum of distances between each step.
        e.g.
            total_dist = dist(prog A.step 0, prog A.step 0) +
                        dist(prog A.step 1, prog A.step 1) +
                        dist(prog A.step 2, prog A.step 2) +
                    ...
        The ideal being programs that are identical will havae a distance of 0.

        Inputs:
            pA : Path or data object from formatProgram.
            pB : Path or data object from formatProgram.
            pad : Padding applied to ensure all programs have pad number of steps
            **kwargs : Keyword arguments sent to scipy.distance.cdist

        Returns computed program distance
    '''
    from scipy.spatial import distance
    if isinstance(pA,str):
        pA = formatProgram(pA,pad)
    if isinstance(pB,str):
        pB = formatProgram(pB,pad)
    # computer distance between each step
    # e.g. step 0 vs step 0 etc.
    return sum([distance.cdist(pA[i].reshape(1,-1),pB[i].reshape(1,-1),**kwargs).flatten()[0] for i in range(pad)])

def makeProgEdge(k):
    '''
        Make an edge where weight is inverse of computeProgDistance

        If the distance is 0, then it's weight to a very high number to indicate a strong relationship

        Inputs:
            k : Pair of node keys from a graph

        Returns a tuple containing the node keys and a dict containing the weight
    '''
    u,v = k
    dist = computeProgDistance(u,v)
    dist = 1e5 if dist ==0 else 1./dist
    return (u,v,{"weight":dist})

def makeProgJSONEdge(k):
    '''
        Make a networkx edge from JSON created values

        The weight is based on the sum of distances between program steps

        Inputs:
            k : Nested tuple from JSON dictionary combination containing a pair of key and items

        Returns tuple of node keys and dictionary containing weight
    '''
    from scipy.spatial import distance
    (u,ud),(v,vd) = k
    pA = np.column_stack(ud)
    pB = np.column_stack(vd)
    # if one has no steps and the other has steps
    if (pA.shape[0] == 0) and (pB.shape[0] != 0):
        pA = np.zeros(pB.shape,dtype=pB.dtype)
    elif (pB.shape[0] == 0) and (pA.shape[0] != 0):
        pB = np.zeros(pA.shape,dtype=pA.dtype)
    # if both are empty assign a high weight because they're related
    else:
        return (u,v,{"weight":1e5})
    # skip over step number and enable columns
    pA = pA[:,2:]
    pB = pB[:,2:]
    # the data is padded to a certain size
    # but if there are still more steps ensure that they're the same size
    if pA.shape[0] < pB.shape[0]:
        pA = np.append(pA,np.zeros((pB.shape[0]-pA.shape[0],20)),0)
    elif pB.shape[0] < pA.shape[0]:
        pB = np.append(pB,np.zeros((pA.shape[0]-pB.shape[0],20)),0)
    # calculate distance between steps and sum them together
    sum_dist = [distance.cdist(A.reshape(1,-1),B.reshape(1,-1)).flatten()[0] for A,B in zip(pA,pB)]
    dist = sum(sum_dist)
    dist = 1e5 if dist ==0 else 1./dist
    return (u,v,{"weight":dist})

def getDistanceFromOrigin(value):
    from scipy.spatial import distance
    pA = np.column_stack(value)[:,2:]
    # construct origin
    origin = np.zeros(pA.shape,pA.dtype)
    sum_dist = [distance.cdist(A.reshape(1,-1),B.reshape(1,-1)).flatten()[0] for A,B in zip(pA,origin)]
    return sum(sum_dist)

def getStepDistanceFromOrigin(value):
    from scipy.spatial import distance
    pA = np.column_stack(value)[:,2:]
    # construct origin
    origin = np.zeros(pA.shape,pA.dtype)
    sum_dist = [distance.cdist(A.reshape(1,-1),B.reshape(1,-1)).flatten()[0] for A,B in zip(pA,origin)]
    return sum_dist

def plotStepDistance(path,lim=3):
    import json
    data = json.load(open(path))
    dists = mp.Pool(5).map(getStepDistanceFromOrigin,data.values())
    
    f = plt.figure()
    ax = f.add_subplot(111,projection='3d')
    # clip to lim number of steps
    for i in range(len(dists)):
        dists[i] = dists[i][:3]
    dists = np.row_stack(dists)
    ax.scatter(dists[:,0],dists[:,1],dists[:,2])
    ax.set(xlabel="Step 0 Distance",ylabel="Step 1 Distance",zlabel="Step 2 Distance")
    plt.show()

def plotAllStepDistance(path):
    import json
    f = plt.figure()
    ax = f.add_subplot(111,projection='3d')
    for fn in glob(path):
        data = json.load(open(fn))
        dists = mp.Pool(5).map(getStepDistanceFromOrigin,data.values())
        for i in range(len(dists)):
            dists[i] = dists[i][:3]
        try:
            dists = np.row_stack(dists)
        except ValueError:
            continue
        ax.scatter(dists[:,0],dists[:,1],dists[:,2])
    ax.set(xlabel="Step 0 Distance",ylabel="Step 1 Distance",zlabel="Step 2 Distance")
    plt.show()

def plotProgDistGraph(data_path,draw=False,save_gefx=True):
    '''
        Plot the files as a weighted graph where weight is
        based on program distance from computeProgDistance

        Inputs:
            data_path : Wildcard path to data files to compute distance between
            draw : Flag to draw the graph

        Returns the graph object
    '''
    import networkx as nx
    from networkx.drawing.nx_agraph import graphviz_layout
    from itertools import combinations
    from scipy.spatial import distance
    import json
    # make a complete graph with blank edges
    # node names are file paths
    if os.path.splitext(data_path)[1] == '.json':
        # load json file
        print("loading JSON")
        data = json.load(open(data_path))
        if not any([any(v) for v in vv] for vv in data.values()):
            return None
        G = nx.complete_graph(list(data.keys()))
        dist = []
        print("updating graph")
        edges = mp.Pool(4).map(makeProgJSONEdge,combinations(data.items(),2))
        print(f"{len(edges)} edges")
        G.update(edges)

    if save_gefx:
        nx.write_gexf(G,f"{os.path.splitext(os.path.basename(data_path))[0]}-gefx.gexf",prettyprint=False)
        return
        
    if draw:
        f,ax = plt.subplots(constrained_layout=True,figsize=(30,30))
        f.suptitle(os.path.splitext(os.path.basename(data_path))[0])
        #pos = nx.random_layout(G)
        #pos = nx.spectral_layout(G)
        ## FIX FOR GRAPHVIZ INSTALL https://github.com/ennauata/housegan/issues/15#issuecomment-776908939 ##
        #pos = graphviz_layout(G, prog='fdp')
        pos = nx.spring_layout(G)
        nx.draw(G,pos=pos,ax=ax,with_labels=False,edge_color='k')
        # format labels to 2 dp
        labels = {(u, v): f'{d["weight"]:.2E}' for u, v, d in G.edges(data=True)}
        nx.draw_networkx_edge_labels(G,pos,edge_labels=labels,ax=ax)
    return G

def plotDistanceFromOrigin(path):
    import json, seaborn
    data = json.load(open(path))
    dists = mp.Pool(5).map(getDistanceFromOrigin,data.values())
    # stack to make feature matrix
    X = np.array(dists).reshape(-1,1)
    seaborn.displot(X.flatten(),kde=True)
    ax = plt.gca()
    ax.set(xlabel="Program Distance from Origin",ylabel="Count",title=f"{os.path.splitext(os.path.basename(path))[0]}")
    plt.show()

def clusterProgDistance(path,**kwargs):
    from sklearn.cluster import OPTICS, cluster_optics_dbscan
    import seaborn
    import json
    # load json file
    data = json.load(open(path))
    dists = mp.Pool(5).map(getDistanceFromOrigin,data.values())
    X = np.array(dists).reshape(-1,1)
    # stack to make feature matrix
    clust = OPTICS(min_samples=50, xi=0.05, min_cluster_size=0.05)
    clust.fit(X)
    labels_200 = cluster_optics_dbscan(
        reachability=clust.reachability_,
        core_distances=clust.core_distances_,
        ordering=clust.ordering_,
        eps=0.5,
    )

    space = np.arange(len(X))
    reachability = clust.reachability_[clust.ordering_]
    labels = clust.labels_[clust.ordering_]
    print(f"labels {len(np.unique(labels))}")
    input()

def redrawProgDist(G):
    import pickle, seaborn
    if isinstance(G,str):
        path = os.path.splitext(os.path.basename(G))[0]
        G = pickle.load(open(G,'rb'))
    else:
        path = "Unknown Graph"
    print("loaded graph")
    print("drawing graph")
    f,ax = plt.subplots(constrained_layout=True,figsize=(14,14))
    f.suptitle(path)
    pos = nx.spring_layout(G)
    nx.draw(G,pos=pos,ax=ax,with_labels=False,edge_color='k')
    # format labels to 2 dp
    edges = G.edges(data=True)
    labels = {(u, v): f'{d["weight"]:.2E}' for u, v, d in edges}
    nx.draw_networkx_edge_labels(G,pos,edge_labels=labels,ax=ax)
    f.savefig(f"{path}.png")
    plt.close(f)

def hotEncodeCat(data,reverse='auto'):
    '''
        Hot encode the given list of category values

        pandas.get_dummies is used to hot encode the values into a 2d matrix
        e.g. tool life
        [0 - 33% ... ,34 - 66% ... , 67 - 100%] ->          0 - 33%  34 - 66%  67 - 100%
                                                    0          1         0          0
                                                    1          1         0          0
                                                    2          1         0          0
                                                    3          1         0          0
                                                    4          1         0          0
                                                    ..       ...       ...        ...
                                                    195        0         0          1
                                                    196        0         0          1
                                                    197        0         0          1
                                                    198        0         0          1
                                                    199        0         0          1
        Each row is then treated as bits of a binary number and converted to a corresponding integer
        e.g. 0 - 33%  34 - 66%  67 - 100%
               1         0          0     -> 0b100 -> 4
        The reverse flag controls the order in which the bits are encoded
        e.g
        If reverse is False
        1         0          0     -> 0b100 -> 4
        If reverse if True
        1         0          0     -> 0b001 -> 1

        This is available when order is important as when reverse is False, the hot encoded values would start
        from 4 so when applying it to a variable such as tool life, the results would be 0 - 33% (4), 34 - 66% (2)
        and 67 - 100% (1) which doesn't make sense.

        If reverse is auto, then it attempts to infer whether the input should be reversed or not.
        When the input is numbers, then reverse is set to True when the first number matches the max
        When the input is strings, then the number in the brackets is found and the same check as the numbers is performed.

        Inputs:
            data : Pandas Series of categorical values
            reverse : Flag to reverse the order the bits are read. If auto, the required state of the flag is inferred from the values
                    Default auto.

        Return np.array of hot encoded values
    '''
    if reverse == 'auto':
        # if the list is numbers
        if data.dtype in (np.float_, np.int_):
            # if the first number is equal to max then the order
            # needs to be reversed to maintain ordinality
            reverse = data[0] == data.max()
        # if it's a set of python strings then assume order doesn't matter
        elif data.dtype == object:
            import re
            # get all numbers from strings
            nums = []
            for val in data:
                match = re.search('\(\d\)',val)
                if not (match is None):
                    nums.append(int(val[match.start()+1:match.end()-1]))
            if len(nums)>0:
                # if the first value is max then set reverse flag
                reverse = nums[0] == max(nums)
            else:
                # attempt to convert first char of first value to an integer
                # for encoding tool life
                # e.g. 0 - 33% -> 0
                try:
                    int(data[0][0])
                except ValueError:
                    reverse = False
                else:
                    # if we can do that then 
                    nums = [int(val[0]) for val in data]
                    reverse = nums[0] != max(nums)
        else:
            reverse = False
    # convert to 2d array of hot encoded values 0s and 1s
    arr = pd.get_dummies(data).values
    # iterate over each row
    # convert 0s and 1s to chars and join them together into a string e.g. [0,1,0] -> '010'
    # convert binary string to an integer and add to a list
    # reverse flag is to control whether the bits are read in forward or reverse direction. e.g. [1,1,0] -> 110 or 011
    binary = [int(''.join([str(c) for c in aa[::(-1 if reverse else 1)].tolist()]),base=2) for aa in arr]
    return np.array(binary)

        
if __name__ == "__main__":
    import networkx as nx
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.preprocessing import LabelEncoder
    #from tqdm.contrib.concurrent import process_map
    import pickle
    #plotKistlerPlateHist(r"COUNTERSPARK/data/4B/*.txt")
    #plotKistlerPlateHist(r"COUNTERSPARK/data/4B/*.txt",process=None)
    #stack = stackCPSheets("COUNTERSPARK/Counterspark drilling test.xlsx")
    #sheets = loadCPDataAsPD("COUNTERSPARK/Counterspark drilling test.xlsx")
    #makeInteractionPlot(r"COUNTERSPARK/data/4B/*.xls","COUNTERSPARK/Counterspark drilling test.xlsx",trace='Through Tool Air (Bar)',add_filter=onlyVirgin)
    #plotCounterspark(r"COUNTERSPARK/data/4B/*.xls","COUNTERSPARK/Counterspark drilling test.xlsx",xaxis='Test Number ',label_ang=45)
    #plotCounterspark3D(r"COUNTERSPARK/data/4B/*.xls","COUNTERSPARK/Counterspark drilling test.xlsx",yaxis='Machine Surface Contact')
    data_path = "COUNTERSPARK/data/*/*.xls"
    #data,_ = getAllFeatures("COUNTERSPARK/Counterspark drilling test.xlsx",data_path)
    #plotProgDistGraph(data_path,True)
    #plotFeatures("COUNTERSPARK/Counterspark drilling test.xlsx",data_path)
    #f = integMats("COUNTERSPARK/Counterspark drilling test.xlsx",data_path,plot=True,norm=True)
    #f = integMats("COUNTERSPARK/Counterspark drilling test.xlsx",data_path,plot=True,norm=False)
    #G = plotDTWGraph(data_path,draw=True)
    cols = ['Condition of Supply','Drilling Orientation','Tool Life','Through Tool Air (Bar)','Vacuum Extraction','Vacuum Extraction Dwell','Machine Surface Contact','Exit pre strip']
##    for fn in glob(r"St-Naz-program-distance-dictionary\*.json"):
##        print(fn)
##        if os.path.basename(fn) == "complete-program-distance-dictionary.json":
##            continue
##        if len(glob(f"{os.path.splitext(os.path.basename(fn))[0]}-gefx.gexf"))>0:
##            continue
##       # plotStepDistance(fn)
##        plotProgDistGraph(fn,draw=False)
        #pickle.dump(G,open(fr"{os.path.splitext(os.path.basename(fn))[0]}-graph.pkl",'wb'))
##    for fn in glob("*.pkl"):
##        print(fn)
##        redrawProgDist(fn)
##        plt.gcf().savefig(f"St-Naz-program-distance-dictionary/{os.path.splitext(os.path.basename(fn))[0].replace(os.path.sep,'-')}.png")
##        plt.close('all')
##        break
##    #graphs = mp.Pool(4).map(plotProgDistGraph,glob(r"St-Naz-program-distance-dictionary\*.json"))
##    graphs = process_map(plotProgDistGraph, glob(r"St-Naz-program-distance-dictionary\*.json"), max_workers=4)
#    all_res = {}
    for cc in ['Tool Life',]:
        df = attemptMultiPredict("COUNTERSPARK/Counterspark drilling test.xlsx",data_path,cc,opath='COUNTERSPARK\classifier',split=np.arange(0.8,1.0,0.1),draw_classifiers=True)
    #signalEnergyMats("COUNTERSPARK/Counterspark drilling test.xlsx",data_path,True)
##    stack = stackCPSheets("COUNTERSPARK/Counterspark drilling test.xlsx")
##    for cc in cols:
##        print(LabelEncoder().fit_transform(stack[cc]))
##        print(stack[cc])
##    
##    for cc in cols:
##        energy = signalEnergyMatsAgainst3D("COUNTERSPARK/Counterspark drilling test.xlsx",data_path,cc,True)
##        f = plt.gcf()
##        f.suptitle(cc)
##        plt.show()
    #plt.plot(area)
    
    plt.show()
